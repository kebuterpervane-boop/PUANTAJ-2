import pandas as pd
from datetime import datetime, date
from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, 
                             QTableWidgetItem, QHeaderView, QPushButton, 
                             QLabel, QComboBox, QFileDialog, QMessageBox, QFrame, QSplitter, QGroupBox, QDialog)
from PySide6.QtWidgets import QProgressDialog  # WHY: show export progress without freezing the UI.
from PySide6.QtCore import Qt
from PySide6.QtCore import QThread, Signal, Slot, QObject  # WHY: background export worker support.
from PySide6.QtGui import QColor
import os
import calendar
from core.database import Database
from core.user_config import load_config, save_config
from core.hesaplama import hesapla_maktu_hakedis

class ExportWorker(QObject):  # WHY: generic worker for background export tasks.
    finished = Signal(object)  # WHY: return payload (path/status) to UI thread.
    error = Signal(str)  # WHY: surface errors safely without crashing UI.

    def __init__(self, task_fn):  # WHY: keep worker reusable for different export tasks.
        super().__init__()  # WHY: initialize QObject for signal/slot usage.
        self._task_fn = task_fn  # WHY: store export task callable.
        self._stop_requested = False  # WHY: allow cooperative cancel handling.

    def request_stop(self):  # WHY: allow UI to request a safe stop.
        self._stop_requested = True  # WHY: set flag without killing thread.

    def should_stop(self):  # WHY: shared stop check for long loops.
        return self._stop_requested or QThread.currentThread().isInterruptionRequested()  # WHY: respect both flags.

    @Slot()
    def run(self):  # WHY: thread entry point.
        try:
            result = self._task_fn(self)  # WHY: execute export task with stop-aware worker.
            self.finished.emit(result)  # WHY: notify UI on completion.
        except Exception as e:
            self.error.emit(str(e))  # WHY: forward exception to UI thread.

class DashboardPage(QWidget):
    def __init__(self, signal_manager):
        super().__init__()
        self.db = Database()
        self.firma_id = getattr(self.db, 'current_firma_id', 1)
        self.tersane_id = 0  # 0 = Tüm tersaneler
        self._needs_refresh = False  # NEW: lazy-load flag to avoid heavy refresh on hidden tabs.
        self.signal_manager = signal_manager
        self.current_data = []
        self._export_thread = None  # WHY: keep background export thread reference.
        self._export_worker = None  # WHY: keep export worker alive during run.
        self._export_dialog = None  # WHY: progress dialog for export operations.
        self._export_done_cb = None  # WHY: store export completion callback.
        self._export_cancelled = False  # WHY: track cancel to suppress success toast.
        self.setup_ui()
        
        today = datetime.now()
        cfg = load_config()
        year = cfg.get("dashboard_year", today.year)
        month = cfg.get("dashboard_month", today.month)
        try:
            year = int(year)
        except Exception:
            year = today.year
        try:
            month = int(month)
        except Exception:
            month = today.month
        if str(year) not in [self.combo_year.itemText(i) for i in range(self.combo_year.count())]:
            year = today.year
        if month < 1 or month > 12:
            month = today.month
        self.combo_year.setCurrentText(str(year))
        self.combo_month.setCurrentIndex(month - 1)
        
        self.combo_year.currentTextChanged.connect(self._on_period_changed)
        self.combo_month.currentIndexChanged.connect(self._on_period_changed)
        
        # --- Sinyal Gelince Hesapla (Lazy) ---
        self.signal_manager.data_updated.connect(self._on_data_updated)  # NEW: avoid heavy refresh on hidden tab.

    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        top_bar = QHBoxLayout()
        self.combo_year = QComboBox()
        self.combo_year.addItems([str(y) for y in range(2024, 2030)])
        self.combo_month = QComboBox()
        self.combo_month.addItems(["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran", 
                                   "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"])
        btn_export = QPushButton("📥 Excel Rapor İndir")
        btn_export.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 5px 15px;")
        btn_export.clicked.connect(self.export_excel)
        top_bar.addWidget(QLabel("Dönem:"))
        top_bar.addWidget(self.combo_month)
        top_bar.addWidget(self.combo_year)
        self.btn_this_month = QPushButton("Bu Ay")
        self.btn_this_month.setFixedWidth(70)
        self.btn_this_month.clicked.connect(lambda: self._set_period_relative(0))
        top_bar.addWidget(self.btn_this_month)
        self.btn_prev_month = QPushButton("Gecen Ay")
        self.btn_prev_month.setFixedWidth(80)
        self.btn_prev_month.clicked.connect(lambda: self._set_period_relative(-1))
        top_bar.addWidget(self.btn_prev_month)
        top_bar.addStretch()
        top_bar.addWidget(btn_export)
        layout.addLayout(top_bar)

        # BUGÜN KİM GELDİ/GELMEDİ WİDGET'I (YENİ)
        today_frame = QFrame()
        today_frame.setStyleSheet("background: #1a237e; border-radius: 8px; padding: 10px;")
        today_layout = QHBoxLayout(today_frame)
        
        self.lbl_today_present = QLabel("Bugün Gelen: 0")
        self.lbl_today_present.setStyleSheet("color: #4CAF50; font-size: 16px; font-weight: bold;")
        
        self.lbl_today_absent = QLabel("Gelmeyen: 0")
        self.lbl_today_absent.setStyleSheet("color: #f44336; font-size: 16px; font-weight: bold;")
        
        self.lbl_today_names = QLabel("")
        self.lbl_today_names.setStyleSheet("color: #ccc; font-size: 11px;")
        self.lbl_today_names.setWordWrap(True)
        
        today_layout.addWidget(QLabel("📍 BUGÜN:"))
        today_layout.addWidget(self.lbl_today_present)
        today_layout.addWidget(self.lbl_today_absent)
        today_layout.addStretch()
        today_layout.addWidget(self.lbl_today_names)
        
        layout.addWidget(today_frame)

        cards = QHBoxLayout()
        self.lbl_pay = self.create_card("Toplam Ödenecek", "0 ₺")
        self.lbl_avans = self.create_card("Toplam Avans", "0 ₺")
        self.lbl_ot = self.create_card("Toplam Mesai", "0 Saat")
        self.lbl_count = self.create_card("Personel", "0 Kişi")
        cards.addWidget(self.lbl_pay)
        cards.addWidget(self.lbl_avans)
        cards.addWidget(self.lbl_ot)
        cards.addWidget(self.lbl_count)
        layout.addLayout(cards)

        # Context menus for cards (sağ-tık -> Detay / Kopyala)
        for c in (self.lbl_pay, self.lbl_avans, self.lbl_ot, self.lbl_count):
            c.setContextMenuPolicy(Qt.CustomContextMenu)
            c.customContextMenuRequested.connect(self.card_context_menu)

        # Store last computed totals for detail dialogs
        self.last_totals = {'tot_pay': 0.0, 'tot_avans': 0.0, 'tot_ot_saat': 0.0, 'tot_ot_yev': 0.0, 'tot_ekstra': 0.0}

        splitter = QSplitter(Qt.Horizontal)
        self.table = QTableWidget()
        # Ekstra sütunu eklendi
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels(["Personel", "Ekip", "Maaş", "Normal", "Mesai", "Birim Üc.", "Brüt", "Ekstra", "Avans", "NET"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("QTableWidget { background-color: #2b2b2b; color: white; gridline-color: #444; border: none; alternate-background-color: #2a2a2a; }")
        
        self.table_teams = QTableWidget()
        self.table_teams.setColumnCount(2)
        self.table_teams.setHorizontalHeaderLabels(["Ekip", "Toplam Hakediş"])
        self.table_teams.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table_teams.setAlternatingRowColors(True)
        self.table_teams.setStyleSheet("""
            QTableWidget { background-color: #2b2b2b; color: white; gridline-color: #444; border: 1px solid #555; alternate-background-color: #2a2a2a; }
            QHeaderView::section { background-color: #4CAF50; color: white; font-weight: bold; }
        """)
        self.table_teams.setFixedWidth(300)

        splitter.addWidget(self.table)
        splitter.addWidget(self.table_teams)
        layout.addWidget(splitter)

    def create_card(self, title, val):
        f = QFrame()
        f.setStyleSheet("background: #333; border-radius: 8px; padding: 15px;")
        l = QVBoxLayout(f)
        l.addWidget(QLabel(title))
        val_lbl = QLabel(val)
        val_lbl.setStyleSheet("color: #4CAF50; font-size: 20px; font-weight: bold;")
        l.addWidget(val_lbl)
        return f

    def set_tersane_id(self, tersane_id, refresh=True):
        """Global tersane seçiciden gelen tersane_id'yi set eder ve verileri yeniler."""
        self.tersane_id = tersane_id
        self._needs_refresh = True  # NEW: mark dirty; refresh can be deferred.
        if refresh:
            self.update_view()  # WHY: only visible page refreshes to keep UI smooth.

    def update_view(self):
        """Görünür sayfa için güncel tersane verilerini yükle."""
        self._needs_refresh = False  # WHY: clear dirty flag after refresh.
        self.calculate()
        self.update_today_widget()

    def refresh_if_needed(self):
        """Lazy-load için: sayfa görünür olduğunda gerekiyorsa güncelle."""
        if self._needs_refresh:
            self.update_view()

    def _get_active_tersane_label(self):  # WHY: centralize export titles with active tersane name.
        if self.tersane_id and self.tersane_id > 0:  # WHY: include selected tersane in export metadata.
            tersane = self.db.get_tersane(self.tersane_id)  # WHY: fetch tersane name for display.
            return tersane['ad'] if tersane else f"ID {self.tersane_id}"  # WHY: fallback keeps export usable if name missing.
        return "Tüm Tersaneler"  # WHY: preserve global mode label when no tersane selected.

    def update_today_widget(self):
        """Bugünkü giriş-çıkış durumunu gösterir"""
        try:
            today_str = date.today().strftime("%Y-%m-%d")
            
            # Bugünkü kayıtları çek
            with self.db.get_connection() as conn:
                c = conn.cursor()
                # NEW: tersane filtresi varsa sadece o tersanenin kayıtlarını çek.
                if self.tersane_id and self.tersane_id > 0:
                    c.execute("""
                        SELECT ad_soyad, giris_saati, cikis_saati 
                        FROM gunluk_kayit 
                        WHERE tarih = ? AND firma_id = ? AND tersane_id = ?
                    """, (today_str, self.firma_id, self.tersane_id))
                else:
                    c.execute("""
                        SELECT ad_soyad, giris_saati, cikis_saati 
                        FROM gunluk_kayit 
                        WHERE tarih = ? AND firma_id = ?
                    """, (today_str, self.firma_id))
                today_records = c.fetchall()
                
                # Tüm aktif personeli çek
                if self.tersane_id and self.tersane_id > 0:
                    c.execute("SELECT ad_soyad FROM personel WHERE firma_id = ? AND tersane_id = ? ORDER BY ad_soyad", (self.firma_id, self.tersane_id))
                else:
                    c.execute("SELECT ad_soyad FROM personel WHERE firma_id = ? ORDER BY ad_soyad", (self.firma_id,))
                all_personnel = {row[0] for row in c.fetchall()}
            
            # Bugün gelenleri ayır
            present_today = {r[0] for r in today_records if r[1] or r[2]}  # Giriş veya çıkış var
            absent_today = all_personnel - present_today
            
            self.lbl_today_present.setText(f"Bugün Gelen: {len(present_today)}")
            self.lbl_today_absent.setText(f"Gelmeyen: {len(absent_today)}")
            
            if absent_today:
                absent_names = ", ".join(sorted(list(absent_today))[:10])
                if len(absent_today) > 10:
                    absent_names += f" (+{len(absent_today)-10} kişi)"
                self.lbl_today_names.setText(f"Gelmeyenler: {absent_names}")
            else:
                self.lbl_today_names.setText("✅ Herkes geldi!")
                
        except Exception as e:
            self.lbl_today_present.setText("Bugün: -")
            self.lbl_today_absent.setText("Hata")

    def _on_data_updated(self):
        """Veri değiştiğinde sadece görünürsek yenile (lazy)."""
        if not self.isVisible():
            self._needs_refresh = True  # WHY: defer heavy refresh until tab is visible.
            return
        self.update_view()

    def _set_period_relative(self, offset_months):
        today = datetime.now()
        y = today.year
        m = today.month + int(offset_months)
        while m <= 0:
            m += 12
            y -= 1
        while m > 12:
            m -= 12
            y += 1
        self.combo_year.blockSignals(True)
        self.combo_month.blockSignals(True)
        self.combo_year.setCurrentText(str(y))
        self.combo_month.setCurrentIndex(m - 1)
        self.combo_year.blockSignals(False)
        self.combo_month.blockSignals(False)
        self._on_period_changed()

    def _on_period_changed(self):
        cfg = load_config()
        cfg["dashboard_year"] = int(self.combo_year.currentText())
        cfg["dashboard_month"] = self.combo_month.currentIndex() + 1
        save_config(cfg)
        self.calculate()

    def calculate(self):
        try:
            y = int(self.combo_year.currentText())
            m = self.combo_month.currentIndex() + 1
            data = self.db.get_dashboard_data(y, m, tersane_id=self.tersane_id)
            self.current_data = data
            self.table.setRowCount(len(data))
            
            def format_mesai_summary(saat, yev):
                if yev > 0 and saat > 0:
                    return f"{saat:,.1f} Saat / {yev:,.1f} Yev."
                if yev > 0:
                    return f"{yev:,.1f} Yev."
                return f"{saat:,.1f} Saat"

            tot_pay, tot_ot_saat, tot_ot_yev = 0.0, 0.0, 0.0
            tot_avans = 0
            tot_ekstra = 0
            team_totals = {}
            
            # Ayın takvim gün sayısını belirle
            _, calendar_gun_sayisi = calendar.monthrange(y, m)
            
            for r, row in enumerate(data):
                maas = row['maas']
                norm = row['top_normal'] or 0
                mesai = row['top_mesai'] or 0
                avans = row['avans'] or 0
                ekstra = row.get('ekstra', 0.0) or 0.0
                ekip = row['ekip'] if row['ekip'] else "Diğer"
                yevmiyeci_mi = row.get('yevmiyeci_mi', 0)  # Yevmiyeci kontrolü
                birim_label = "Yevmiye" if yevmiyeci_mi else "Saat"
                
                # YEVMİYECİ HESAPLAMA
                if yevmiyeci_mi:
                    gunluk_yevmiye = maas
                    total_raw = norm + mesai
                    total_yevmiye_rounded = round(total_raw * 4) / 4.0
                    brut = total_yevmiye_rounded * gunluk_yevmiye
                    saat_ucreti = gunluk_yevmiye  # Saat ücreti yerine günlük yevmiye
                else:
                    # MAKTU ÜCRET SİSTEMİ — Bordro ile aynı formül
                    calisan_gun_sayisi_row = row.get('calisan_gun_sayisi', 0)
                    maktu = hesapla_maktu_hakedis(y, m, calisan_gun_sayisi_row, maas)
                    hakedis = maktu['hakedis']
                    saat_ucreti = maktu['gunluk_ucret']
                    # Mesai: mesai saat * (maas / 225)
                    mesai_saat_ucreti = maas / 225.0 if maas > 0 else 0
                    mesai_bonus = mesai * mesai_saat_ucreti
                    brut = hakedis + mesai_bonus
                net = brut + ekstra - avans
                
                tot_pay += net
                if yevmiyeci_mi:
                    tot_ot_yev += mesai
                else:
                    tot_ot_saat += mesai
                tot_avans += avans
                tot_ekstra += ekstra
                
                if ekip not in team_totals: team_totals[ekip] = 0
                team_totals[ekip] += net
                
                self.table.setItem(r, 0, QTableWidgetItem(row['ad_soyad']))
                self.table.setItem(r, 1, QTableWidgetItem(ekip))
                self.table.setItem(r, 2, QTableWidgetItem(f"{maas:,.0f}"))
                norm_item = QTableWidgetItem(f"{norm:.1f}")
                mesai_item = QTableWidgetItem(f"{mesai:.1f}")
                ucret_item = QTableWidgetItem(f"{saat_ucreti:.2f}")
                norm_item.setToolTip(f"Birim: {birim_label}")
                mesai_item.setToolTip(f"Birim: {birim_label}")
                ucret_item.setToolTip(f"Birim: {birim_label}")
                self.table.setItem(r, 3, norm_item)
                self.table.setItem(r, 4, mesai_item)
                self.table.setItem(r, 5, ucret_item)
                self.table.setItem(r, 6, QTableWidgetItem(f"{brut:,.2f}"))
                self.table.setItem(r, 7, QTableWidgetItem(f"{ekstra:,.2f}"))
                self.table.setItem(r, 8, QTableWidgetItem(f"{avans:,.2f}"))
                net_item = QTableWidgetItem(f"{net:,.2f} ₺")
                net_item.setBackground(QColor("#1B5E20"))
                self.table.setItem(r, 9, net_item)

            self.table_teams.setRowCount(len(team_totals))
            for i, (tm, val) in enumerate(team_totals.items()):
                self.table_teams.setItem(i, 0, QTableWidgetItem(tm))
                v_item = QTableWidgetItem(f"{val:,.2f} ₺")
                v_item.setBackground(QColor("#0D47A1"))
                self.table_teams.setItem(i, 1, v_item)

            # persist into last_totals for detail dialogs
            self.last_totals = {
                'tot_pay': tot_pay,
                'tot_avans': tot_avans,
                'tot_ot_saat': tot_ot_saat,
                'tot_ot_yev': tot_ot_yev,
                'tot_ekstra': tot_ekstra
            }

            # Update card labels so UI reflects calculated totals
            try:
                self.lbl_pay.layout().itemAt(1).widget().setText(f"{tot_pay:,.2f} ₺")
                self.lbl_avans.layout().itemAt(1).widget().setText(f"{tot_avans:,.2f} ₺")
                self.lbl_ot.layout().itemAt(1).widget().setText(format_mesai_summary(tot_ot_saat, tot_ot_yev))
                self.lbl_count.layout().itemAt(1).widget().setText(f"{len(data)} Kişi")

                # Tooltips for quick breakdown
                self.lbl_pay.setToolTip(f"Brüt toplam: {sum([row.get('maas',0) for row in data]):,.2f} ₺\nEkstra toplam: {tot_ekstra:,.2f} ₺\nAvans toplam: {tot_avans:,.2f} ₺")
                self.lbl_avans.setToolTip(f"Ay içindeki toplam avans tutarı: {tot_avans:,.2f} ₺")
                self.lbl_ot.setToolTip(f"Toplam mesai: {tot_ot_saat:.1f} Saat, {tot_ot_yev:.1f} Yevmiye")
            except Exception:
                # Silinmiş ya da yeniden düzenlenmiş widget'larda hata almamak için sessizce geç
                pass
        except Exception as e:
            from core.app_logger import log_error
            log_error(f"Dashboard calculate error: {e}")

    def card_context_menu(self, pos):
        """Show context menu for dashboard cards"""
        from PySide6.QtWidgets import QMenu
        card = self.sender()
        menu = QMenu(self)
        act_detay = menu.addAction("Detayları Göster")
        act_copy = menu.addAction("Kopyala")
        action = menu.exec_(card.mapToGlobal(pos))
        if action == act_copy:
            # copy the value text (second widget in layout)
            try:
                val_lbl = card.layout().itemAt(1).widget()
                QApplication.clipboard().setText(val_lbl.text())
            except Exception:
                pass
        elif action == act_detay:
            # show a dialog with breakdown
            vals = self.last_totals
            text = (
                f"Toplam Brüt (tüm personel): {vals.get('tot_pay',0):,.2f} ₺\n"
                f"Toplam Ekstra: {vals.get('tot_ekstra',0):,.2f} ₺\n"
                f"Toplam Avans: {vals.get('tot_avans',0):,.2f} ₺\n"
                f"Toplam Mesai (Saat): {vals.get('tot_ot_saat',0):.1f} Saat\n"
                f"Toplam Mesai (Yevmiye): {vals.get('tot_ot_yev',0):.1f} Yevmiye"
            )
            QMessageBox.information(self, "Dashboard Detay", text)

    def _export_excel_legacy(self):  # WHY: keep original sync export as reference; replaced by threaded version below.
        # Use ExportDialog to pick date range / team / person and options
        from page_records import ExportDialog
        dlg = ExportDialog(self.db, self)
        if dlg.exec() != QDialog.Accepted:
            return
        vals = dlg.get_values()
        try:
            try:
                import openpyxl  # required by export_df_to_file
            except Exception:
                QMessageBox.critical(self, "Hata", "openpyxl yüklü değil. Lütfen openpyxl yükleyin.")
                return
            firma_id = getattr(self.db, 'current_firma_id', 1)
            raw = self.db.get_records_between(vals['date_from'], vals['date_to'], team=vals['team'], person=vals['person'])
            rows = []
            for r in raw:
                rows.append({
                    'Tarih': r[1],
                    'Personel': r[2],
                    'Ekip': r[9] if r[9] else '',
                    'Giriş': r[3] or '',
                    'Çıkış': r[4] or '',
                    'Kayıp': r[5] or '',
                    'Normal': float(r[6] or 0),
                    'Mesai': float(r[7] or 0),
                    'Açıklama': r[8] or ''
                })

            if not rows:
                QMessageBox.information(self, "Bilgi", "Seçilen aralıkta kayıt bulunamadı.")
                return

            df = pd.DataFrame(rows)

            # Summary per person (dashboard style)
            persons = sorted(set(df['Personel'].tolist()))
            summary_rows = []
            avans_map = {}
            with self.db.get_connection() as conn:
                c = conn.cursor()
                c.execute("SELECT ad_soyad, SUM(CASE WHEN tur IN ('Avans', 'Kesinti') THEN tutar ELSE 0 END) FROM avans_kesinti WHERE tarih BETWEEN ? AND ? GROUP BY ad_soyad", (vals['date_from'], vals['date_to']))
                for r in c.fetchall():
                    avans_map[r[0]] = r[1]

            # Month context for maktu calculation (fallback to start date month)
            try:
                dt_from = datetime.strptime(vals['date_from'], "%Y-%m-%d")
                dt_to = datetime.strptime(vals['date_to'], "%Y-%m-%d")
                if dt_from.year == dt_to.year and dt_from.month == dt_to.month:
                    _, calendar_gun_sayisi = calendar.monthrange(dt_from.year, dt_from.month)
                else:
                    _, calendar_gun_sayisi = calendar.monthrange(dt_from.year, dt_from.month)
            except Exception:
                calendar_gun_sayisi = 30

            # Aylık ekstra map — personel_ekstra_aylik öncelikli, fallback personel.ekstra_odeme
            ekstra_aylik_map = self.db.get_ekstra_aylik_bulk(dt_from.year, dt_from.month)
            tot_pay, tot_avans, tot_ot_saat, tot_ot_yev = 0.0, 0.0, 0.0, 0.0
            for p in persons:
                with self.db.get_connection() as conn:
                    c = conn.cursor()
                    c.execute("SELECT maas, ekip_adi, ekstra_odeme, COALESCE(yevmiyeci_mi,0) FROM personel WHERE ad_soyad=? AND firma_id=?", (p, firma_id))
                    inf = c.fetchone() or (0, '', 0.0, 0)
                maas = float(inf[0] or 0)
                ekip = inf[1]
                ekstra = ekstra_aylik_map[p][0] if p in ekstra_aylik_map else float(inf[2] or 0.0)
                yevmiyeci_mi = bool(inf[3])
                top_normal = df[df['Personel'] == p]['Normal'].sum()
                top_mesai = df[df['Personel'] == p]['Mesai'].sum()

                if yevmiyeci_mi:
                    gunluk_yevmiye = maas
                    total_raw = top_normal + top_mesai
                    total_yevmiye_rounded = round(total_raw * 4) / 4.0
                    brut = total_yevmiye_rounded * gunluk_yevmiye
                    saat_ucreti = gunluk_yevmiye
                else:
                    calisan_gun_sayisi_row = len(df[(df['Personel'] == p) & (df['Normal'] > 0)])
                    maktu = hesapla_maktu_hakedis(dt_from.year, dt_from.month, calisan_gun_sayisi_row, maas)
                    hakedis = maktu['hakedis']
                    saat_ucreti = maktu['gunluk_ucret']
                    mesai_saat_ucreti = maas / 225.0 if maas > 0 else 0
                    mesai_bonus = top_mesai * mesai_saat_ucreti
                    brut = hakedis + mesai_bonus

                avans = float(avans_map.get(p, 0.0) or 0.0)
                net = brut + ekstra - avans

                tot_pay += net
                if yevmiyeci_mi:
                    tot_ot_yev += top_mesai
                else:
                    tot_ot_saat += top_mesai
                tot_avans += avans

                summary_rows.append({
                    'Personel': p,
                    'Ekip': ekip,
                    'Maaş': maas,
                    'Normal': top_normal,
                    'Mesai': top_mesai,
                    'Birim Üc.': saat_ucreti,
                    'Brüt': brut,
                    'Ekstra': ekstra,
                    'Avans': avans,
                    'NET': net,
                })

            columns = ["Personel", "Ekip", "Maaş", "Normal", "Mesai", "Birim Üc.", "Brüt", "Ekstra", "Avans", "NET"]
            df_sum = pd.DataFrame(summary_rows, columns=columns)

            # Save with a styled, professional report layout
            out_name = f"Rapor_{vals['date_from']}_{vals['date_to']}.xlsx"
            cfg = load_config()
            last_dir = cfg.get("last_export_dir", "")
            default_path = os.path.join(last_dir, out_name) if last_dir else out_name
            path, _ = QFileDialog.getSaveFileName(self, "Raporu Kaydet", default_path, "Excel (*.xlsx)")
            if not path:
                return
            try:
                cfg["last_export_dir"] = os.path.dirname(path)
                save_config(cfg)
            except Exception:
                pass

            # Firma adı
            firma_adi = "GENEL"
            try:
                with self.db.get_connection() as conn:
                    row = conn.execute("SELECT ad FROM firma WHERE id=?", (firma_id,)).fetchone()
                    if row and row[0]:
                        firma_adi = row[0]
            except Exception:
                pass

            from datetime import datetime as dt
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                title_row = 1
                meta_row = 2
                box_title_row = 4
                box_value_row = 5
                table_header_row = 7

                df_sum.to_excel(writer, index=False, sheet_name='Rapor', startrow=table_header_row - 1)
                ws = writer.sheets['Rapor']

                # Title + meta
                ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=ws.max_column)
                tcell = ws.cell(row=title_row, column=1)
                tcell.value = "Puantaj Raporu"
                tcell.font = Font(bold=True, size=16)
                tcell.alignment = Alignment(horizontal='center')

                ws.merge_cells(start_row=meta_row, start_column=1, end_row=meta_row, end_column=ws.max_column)
                ws.cell(row=meta_row, column=1).value = f"Firma: {firma_adi}   |   Tarih Aralığı: {vals['date_from']} - {vals['date_to']}   |   Rapor Tarihi: {dt.now().strftime('%Y-%m-%d %H:%M')}"
                ws.cell(row=meta_row, column=1).font = Font(size=10, color="FF455A64")
                ws.cell(row=meta_row, column=1).alignment = Alignment(horizontal='center')

                # Summary boxes
                box_border = Border(
                    left=Side(border_style="thin", color="FFB0BEC5"),
                    right=Side(border_style="thin", color="FFB0BEC5"),
                    top=Side(border_style="thin", color="FFB0BEC5"),
                    bottom=Side(border_style="thin", color="FFB0BEC5"),
                )
                box_title_fill = PatternFill(start_color='FFECEFF1', end_color='FFECEFF1', fill_type='solid')
                box_value_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
                value_font = Font(bold=True, size=16, color="FF1B5E20")
                title_font = Font(bold=True, size=11, color="FF263238")

                if tot_ot_yev > 0 and tot_ot_saat > 0:
                    mesai_box = f"{tot_ot_saat:,.1f} Saat / {tot_ot_yev:,.1f} Yev."
                elif tot_ot_yev > 0:
                    mesai_box = f"{tot_ot_yev:,.1f} Yev."
                else:
                    mesai_box = f"{tot_ot_saat:,.1f} Saat"
                boxes = [
                    ("Toplam Ödenecek", f"{tot_pay:,.2f} ₺", ("A", "C")),
                    ("Toplam Avans", f"{tot_avans:,.2f} ₺", ("D", "F")),
                    ("Toplam Mesai", mesai_box, ("G", "H")),
                    ("Personel", f"{len(persons)} Kişi", ("I", "J")),
                ]

                for title, value, (c1, c2) in boxes:
                    ws.merge_cells(f"{c1}{box_title_row}:{c2}{box_title_row}")
                    ws.merge_cells(f"{c1}{box_value_row}:{c2}{box_value_row}")
                    tc = ws[f"{c1}{box_title_row}"]
                    vc = ws[f"{c1}{box_value_row}"]
                    tc.value = title
                    vc.value = value
                    tc.font = title_font
                    vc.font = value_font
                    tc.alignment = Alignment(horizontal='center', vertical='center')
                    vc.alignment = Alignment(horizontal='center', vertical='center')
                    for r in (box_title_row, box_value_row):
                        for col in range(ord(c1), ord(c2) + 1):
                            cell = ws[f"{chr(col)}{r}"]
                            cell.fill = box_title_fill if r == box_title_row else box_value_fill
                            cell.border = box_border

                ws.row_dimensions[box_title_row].height = 22
                ws.row_dimensions[box_value_row].height = 26

                header_row = table_header_row
                header_fill = PatternFill(start_color='FF263238', end_color='FF263238', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFFFF', size=12)
                thin = Side(border_style="thin", color="FFB0BEC5")
                for col_cell in ws[header_row]:
                    col_cell.fill = header_fill
                    col_cell.font = header_font
                    col_cell.alignment = Alignment(horizontal='center', vertical='center')
                    col_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                # Column widths
                dims = {}
                for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row):
                    for cell in row:
                        if cell.value is not None:
                            dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len(str(cell.value)))
                for col, value in dims.items():
                    ws.column_dimensions[col].width = min(max(value + 2, 10), 40)
                ws.column_dimensions['A'].width = max(ws.column_dimensions['A'].width, 24)

                # Freeze + filter
                ws.freeze_panes = f"A{header_row+1}"
                last_data_row = ws.max_row
                ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{last_data_row}"

                # Alternating row colors
                fill_gray = PatternFill(start_color='FFF7F7F7', end_color='FFF7F7F7', fill_type='solid')
                for idx, row in enumerate(ws.iter_rows(min_row=header_row+1, max_row=last_data_row), start=0):
                    if idx % 2 == 0:
                        for cell in row:
                            cell.fill = fill_gray

                # Number formats + alignment
                currency_cols = {"Maaş", "Birim Üc.", "Brüt", "Ekstra", "Avans", "NET"}
                hour_cols = {"Normal", "Mesai"}
                header_labels = [cell.value for cell in ws[header_row]]
                for col_idx, label in enumerate(header_labels, start=1):
                    col_letter = get_column_letter(col_idx)
                    if label in currency_cols:
                        fmt = '#,##0.00 "₺"'
                        align = Alignment(horizontal='right')
                    elif label in hour_cols:
                        fmt = '0.00'
                        align = Alignment(horizontal='right')
                    elif label in ("Personel", "Ekip"):
                        fmt = None
                        align = Alignment(horizontal='left')
                    else:
                        fmt = None
                        align = Alignment(horizontal='center')

                    for r in range(header_row + 1, last_data_row + 1):
                        cell = ws[f"{col_letter}{r}"]
                        if fmt:
                            cell.number_format = fmt
                        cell.alignment = align

                # NET column highlight
                if "NET" in header_labels:
                    net_col_idx = header_labels.index("NET") + 1
                    net_col_letter = get_column_letter(net_col_idx)
                    net_fill = PatternFill(start_color='FF1B5E20', end_color='FF1B5E20', fill_type='solid')
                    net_font = Font(bold=True, color='FFFFFFFF')
                    ws[f"{net_col_letter}{header_row}"].fill = net_fill
                    ws[f"{net_col_letter}{header_row}"].font = net_font
                    for r in range(header_row + 1, last_data_row + 1):
                        cell = ws[f"{net_col_letter}{r}"]
                        cell.fill = PatternFill(start_color='FFE8F5E9', end_color='FFE8F5E9', fill_type='solid')

                # Totals row
                total_row = last_data_row + 1
                ws.cell(row=total_row, column=1, value="TOPLAM")
                ws.cell(row=total_row, column=1).font = Font(bold=True)
                ws.cell(row=total_row, column=1).fill = PatternFill(start_color='FFE3F2FD', end_color='FFE3F2FD', fill_type='solid')

                total_cols = {"Normal", "Mesai", "Brüt", "Ekstra", "Avans", "NET"}
                for col_idx, label in enumerate(header_labels, start=1):
                    if label in total_cols:
                        col_letter = get_column_letter(col_idx)
                        ws.cell(row=total_row, column=col_idx).value = f"=SUM({col_letter}{header_row+1}:{col_letter}{last_data_row})"
                        ws.cell(row=total_row, column=col_idx).font = Font(bold=True)
                        ws.cell(row=total_row, column=col_idx).fill = PatternFill(start_color='FFE3F2FD', end_color='FFE3F2FD', fill_type='solid')
                        ws.cell(row=total_row, column=col_idx).alignment = Alignment(horizontal='right')

                for cell in ws[total_row]:
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Offer to open the generated file or show folder
            if os.path.exists(path):
                if QMessageBox.question(self, "Aç", "Rapor kaydedildi. Dosyayı şimdi açmak ister misiniz?", QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
                    try:
                        os.startfile(os.path.abspath(path))
                    except Exception:
                        QMessageBox.information(self, "Bilgi", f"Dosya kaydedildi: {path}")
                else:
                    if QMessageBox.question(self, "Klasöre Git", "Klasörü açmak ister misiniz?", QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
                        try:
                            os.startfile(os.path.dirname(os.path.abspath(path)))
                        except Exception:
                            QMessageBox.information(self, "Bilgi", f"Dosya kaydedildi: {path}")

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Export hatası: {e}")

    def _start_export_worker(self, task_fn, done_cb=None, label="Rapor hazırlanıyor..."):  # WHY: shared export runner to keep UI responsive.
        if self._export_thread and self._export_thread.isRunning():  # WHY: avoid overlapping exports that could lock files.
            QMessageBox.information(self, "Bilgi", "Devam eden bir dışa aktarma var.")  # WHY: inform user without starting another thread.
            return
        self._export_done_cb = done_cb  # WHY: keep per-export UI completion handler.
        self._export_cancelled = False  # WHY: reset cancel state for each new export.
        self._export_dialog = QProgressDialog(label, None, 0, 0, self)  # WHY: show indeterminate progress during export.
        self._export_dialog.setWindowModality(Qt.WindowModal)  # WHY: keep modal behavior consistent with other dialogs.
        self._export_dialog.setAutoClose(False)  # WHY: we close explicitly on signals to avoid stuck dialogs.
        self._export_dialog.setAutoReset(False)  # WHY: prevent auto-reset from hiding progress early.
        self._export_dialog.setMinimumDuration(0)  # WHY: show immediately to avoid perceived freeze.
        self._export_dialog.setAttribute(Qt.WA_DeleteOnClose, False)  # WHY: keep dialog alive for late signals.
        self._export_dialog.canceled.connect(self._on_export_dialog_canceled)  # WHY: allow safe cancel without crashing.
        self._export_dialog.rejected.connect(self._on_export_dialog_canceled)  # WHY: handle window close (X) safely.
        self._export_dialog.show()  # WHY: show progress feedback during background work.

        self._export_thread = QThread()  # WHY: run heavy export in background.
        worker = ExportWorker(task_fn)  # WHY: reuse generic export worker for different tasks.
        self._export_worker = worker  # WHY: keep a strong reference to prevent GC.
        worker.moveToThread(self._export_thread)  # WHY: execute worker in background thread.
        self._export_thread.started.connect(worker.run)  # WHY: start export when thread starts.
        worker.finished.connect(self._on_export_finished)  # WHY: handle success payload in UI thread.
        worker.finished.connect(self._export_dialog.accept)  # WHY: close dialog on normal completion.
        worker.finished.connect(self._export_thread.quit)  # WHY: stop thread event loop after completion.
        worker.finished.connect(worker.deleteLater)  # WHY: free worker safely.
        worker.error.connect(self._on_export_error)  # WHY: surface errors without freezing UI.
        worker.error.connect(self._export_thread.quit)  # WHY: stop thread on error.
        worker.error.connect(worker.deleteLater)  # WHY: free worker on error path.
        self._export_thread.finished.connect(self._on_export_thread_finished)  # WHY: clear refs after thread stops.
        self._export_thread.finished.connect(self._export_thread.deleteLater)  # WHY: free thread object after finish.
        self._export_thread.start()  # WHY: kick off background export.

    def _on_export_finished(self, result):  # WHY: centralize export completion handling.
        if self._export_dialog:  # WHY: close dialog if still alive.
            try:
                self._export_dialog.close()  # WHY: ensure dialog closes on completion.
            except RuntimeError:
                pass  # WHY: dialog already deleted; ignore safely.
        self._export_dialog = None  # WHY: release UI reference after safe close.
        if self._export_cancelled:  # WHY: skip success dialog when user cancelled.
            return
        if self._export_done_cb:  # WHY: allow per-export success handling.
            self._export_done_cb(result)  # WHY: call UI completion handler with payload.
            self._export_done_cb = None  # WHY: clear handler after use.

    def _on_export_error(self, msg):  # WHY: handle export errors uniformly.
        if self._export_dialog:  # WHY: close dialog on error.
            try:
                self._export_dialog.close()  # WHY: avoid stuck progress dialog on failure.
            except RuntimeError:
                pass  # WHY: dialog already deleted; ignore safely.
        self._export_dialog = None  # WHY: release UI reference after safe close.
        QMessageBox.critical(self, "Hata", f"Dışa aktarma sırasında hata: {msg}")  # WHY: show error without crashing UI.

    def _on_export_dialog_canceled(self):  # WHY: safe cancel handling for export progress dialog.
        self._export_cancelled = True  # WHY: mark cancel to skip success toast later.
        if self._export_worker:  # WHY: request cooperative stop for worker.
            self._export_worker.request_stop()  # WHY: avoid hard thread termination.
        if self._export_thread:  # WHY: set interruption flag for worker to observe.
            self._export_thread.requestInterruption()  # WHY: allow cooperative stop in loops.
        if self._export_dialog:  # WHY: update dialog text to show canceling.
            try:
                self._export_dialog.setLabelText("İptal ediliyor...")  # WHY: immediate feedback on cancel.
                self._export_dialog.setCancelButtonText("")  # WHY: prevent repeated cancel clicks.
            except RuntimeError:
                pass  # WHY: dialog already deleted; ignore safely.

    def _on_export_thread_finished(self):  # WHY: clean up thread refs safely after export.
        self._export_thread = None  # WHY: clear thread ref after it stops.
        self._export_worker = None  # WHY: clear worker ref after thread completion.

    def export_excel(self):  # WHY: threaded export to keep UI responsive.
        # Use ExportDialog to pick date range / team / person and options
        from page_records import ExportDialog  # WHY: reuse existing dialog without duplication.
        dlg = ExportDialog(self.db, self)  # WHY: keep same filter UI as before.
        if dlg.exec() != QDialog.Accepted:
            return
        vals = dlg.get_values()

        try:
            import openpyxl  # required by export formatting
        except Exception:
            QMessageBox.critical(self, "Hata", "openpyxl yüklü değil. Lütfen openpyxl yükleyin.")  # WHY: fail fast if missing.
            return

        tersane_id = self.tersane_id or 0  # WHY: normalize to global (0) if no tersane selected.
        tersane_label = self._get_active_tersane_label()  # WHY: include tersane in report title.
        firma_id = self.firma_id  # WHY: keep firma context consistent with existing logic.

        out_name = f"Rapor_{vals['date_from']}_{vals['date_to']}.xlsx"
        cfg = load_config()
        last_dir = cfg.get("last_export_dir", "")
        default_path = os.path.join(last_dir, out_name) if last_dir else out_name
        path, _ = QFileDialog.getSaveFileName(self, "Raporu Kaydet", default_path, "Excel (*.xlsx)")
        if not path:
            return
        try:
            cfg["last_export_dir"] = os.path.dirname(path)
            save_config(cfg)
        except Exception:
            pass

        def _task(worker):  # WHY: run export off the UI thread.
            if worker.should_stop():
                return {"status": "cancelled"}  # WHY: allow user-initiated cancel.
            db = Database()  # WHY: use thread-local DB handle for safe background access.
            raw = db.get_records_between(vals['date_from'], vals['date_to'], team=vals['team'], person=vals['person'], tersane_id=tersane_id)  # WHY: filter by active tersane.
            rows = []
            for r in raw:
                if worker.should_stop():
                    return {"status": "cancelled"}  # WHY: allow safe cancel during loops.
                rows.append({
                    'Tarih': r[1],
                    'Personel': r[2],
                    'Ekip': r[9] if r[9] else '',
                    'Giriş': r[3] or '',
                    'Çıkış': r[4] or '',
                    'Kayıp': r[5] or '',
                    'Normal': float(r[6] or 0),
                    'Mesai': float(r[7] or 0),
                    'Açıklama': r[8] or ''
                })
            if not rows:
                return {"status": "empty"}  # WHY: report empty data back to UI.

            df = pd.DataFrame(rows)

            # Summary per person (dashboard style)
            persons = sorted(set(df['Personel'].tolist()))
            allowed_persons = set(persons)  # WHY: filter avans to exported personnel only.
            summary_rows = []
            avans_map = {}
            with db.get_connection() as conn:
                c = conn.cursor()
                c.execute("SELECT ad_soyad, SUM(CASE WHEN tur IN ('Avans', 'Kesinti') THEN tutar ELSE 0 END) FROM avans_kesinti WHERE tarih BETWEEN ? AND ? GROUP BY ad_soyad", (vals['date_from'], vals['date_to']))
                for r in c.fetchall():
                    if not allowed_persons or r[0] in allowed_persons:
                        avans_map[r[0]] = r[1]

            # Month context for maktu calculation (fallback to start date month)
            try:
                dt_from = datetime.strptime(vals['date_from'], "%Y-%m-%d")
                dt_to = datetime.strptime(vals['date_to'], "%Y-%m-%d")
                if dt_from.year == dt_to.year and dt_from.month == dt_to.month:
                    _, calendar_gun_sayisi = calendar.monthrange(dt_from.year, dt_from.month)
                else:
                    _, calendar_gun_sayisi = calendar.monthrange(dt_from.year, dt_from.month)
            except Exception:
                calendar_gun_sayisi = 30

            # Aylık ekstra map — personel_ekstra_aylik öncelikli, fallback personel.ekstra_odeme
            ekstra_aylik_map = db.get_ekstra_aylik_bulk(dt_from.year, dt_from.month)
            tot_pay, tot_avans, tot_ot_saat, tot_ot_yev = 0.0, 0.0, 0.0, 0.0
            for p in persons:
                with db.get_connection() as conn:
                    c = conn.cursor()
                    c.execute("SELECT maas, ekip_adi, ekstra_odeme, COALESCE(yevmiyeci_mi,0) FROM personel WHERE ad_soyad=? AND firma_id=?", (p, firma_id))
                    inf = c.fetchone() or (0, '', 0.0, 0)
                maas = float(inf[0] or 0)
                ekip = inf[1]
                ekstra = ekstra_aylik_map[p][0] if p in ekstra_aylik_map else float(inf[2] or 0.0)
                yevmiyeci_mi = bool(inf[3])
                top_normal = df[df['Personel'] == p]['Normal'].sum()
                top_mesai = df[df['Personel'] == p]['Mesai'].sum()

                if yevmiyeci_mi:
                    gunluk_yevmiye = maas
                    total_raw = top_normal + top_mesai
                    total_yevmiye_rounded = round(total_raw * 4) / 4.0
                    brut = total_yevmiye_rounded * gunluk_yevmiye
                    saat_ucreti = gunluk_yevmiye
                else:
                    calisan_gun_sayisi_row = len(df[(df['Personel'] == p) & (df['Normal'] > 0)])
                    maktu = hesapla_maktu_hakedis(dt_from.year, dt_from.month, calisan_gun_sayisi_row, maas)
                    hakedis = maktu['hakedis']
                    saat_ucreti = maktu['gunluk_ucret']
                    mesai_saat_ucreti = maas / 225.0 if maas > 0 else 0
                    mesai_bonus = top_mesai * mesai_saat_ucreti
                    brut = hakedis + mesai_bonus

                avans = float(avans_map.get(p, 0.0) or 0.0)
                net = brut + ekstra - avans

                tot_pay += net
                if yevmiyeci_mi:
                    tot_ot_yev += top_mesai
                else:
                    tot_ot_saat += top_mesai
                tot_avans += avans

                summary_rows.append({
                    'Personel': p,
                    'Ekip': ekip,
                    'Maaş': maas,
                    'Normal': top_normal,
                    'Mesai': top_mesai,
                    'Birim Üc.': saat_ucreti,
                    'Brüt': brut,
                    'Ekstra': ekstra,
                    'Avans': avans,
                    'NET': net,
                })

            columns = ["Personel", "Ekip", "Maaş", "Normal", "Mesai", "Birim Üc.", "Brüt", "Ekstra", "Avans", "NET"]
            df_sum = pd.DataFrame(summary_rows, columns=columns)

            # Firma adı
            firma_adi = "GENEL"
            try:
                with db.get_connection() as conn:
                    row = conn.execute("SELECT ad FROM firma WHERE id=?", (firma_id,)).fetchone()
                    if row and row[0]:
                        firma_adi = row[0]
            except Exception:
                pass

            from datetime import datetime as dt
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                title_row = 1
                meta_row = 2
                box_title_row = 4
                box_value_row = 5
                table_header_row = 7

                df_sum.to_excel(writer, index=False, sheet_name='Rapor', startrow=table_header_row - 1)
                ws = writer.sheets['Rapor']

                # Title + meta
                ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=ws.max_column)
                tcell = ws.cell(row=title_row, column=1)
                tcell.value = f"Puantaj Raporu - {tersane_label}"
                tcell.font = Font(bold=True, size=16)
                tcell.alignment = Alignment(horizontal='center')

                ws.merge_cells(start_row=meta_row, start_column=1, end_row=meta_row, end_column=ws.max_column)
                ws.cell(row=meta_row, column=1).value = f"Firma: {firma_adi}   |   Tersane: {tersane_label}   |   Tarih Aralığı: {vals['date_from']} - {vals['date_to']}   |   Rapor Tarihi: {dt.now().strftime('%Y-%m-%d %H:%M')}"
                ws.cell(row=meta_row, column=1).font = Font(size=10, color="FF455A64")
                ws.cell(row=meta_row, column=1).alignment = Alignment(horizontal='center')

                # Summary boxes
                box_border = Border(
                    left=Side(border_style="thin", color="FFB0BEC5"),
                    right=Side(border_style="thin", color="FFB0BEC5"),
                    top=Side(border_style="thin", color="FFB0BEC5"),
                    bottom=Side(border_style="thin", color="FFB0BEC5"),
                )
                box_title_fill = PatternFill(start_color='FFECEFF1', end_color='FFECEFF1', fill_type='solid')
                box_value_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
                value_font = Font(bold=True, size=16, color="FF1B5E20")
                title_font = Font(bold=True, size=11, color="FF263238")

                if tot_ot_yev > 0 and tot_ot_saat > 0:
                    mesai_box = f"{tot_ot_saat:,.1f} Saat / {tot_ot_yev:,.1f} Yev."
                elif tot_ot_yev > 0:
                    mesai_box = f"{tot_ot_yev:,.1f} Yev."
                else:
                    mesai_box = f"{tot_ot_saat:,.1f} Saat"
                boxes = [
                    ("Toplam Ödenecek", f"{tot_pay:,.2f} ₺", ("A", "C")),
                    ("Toplam Avans", f"{tot_avans:,.2f} ₺", ("D", "F")),
                    ("Toplam Mesai", mesai_box, ("G", "H")),
                    ("Personel", f"{len(persons)} Kişi", ("I", "J")),
                ]

                for title, value, (c1, c2) in boxes:
                    ws.merge_cells(f"{c1}{box_title_row}:{c2}{box_title_row}")
                    ws.merge_cells(f"{c1}{box_value_row}:{c2}{box_value_row}")
                    tc = ws[f"{c1}{box_title_row}"]
                    vc = ws[f"{c1}{box_value_row}"]
                    tc.value = title
                    vc.value = value
                    tc.font = title_font
                    vc.font = value_font
                    tc.alignment = Alignment(horizontal='center', vertical='center')
                    vc.alignment = Alignment(horizontal='center', vertical='center')
                    for r in (box_title_row, box_value_row):
                        for col in range(ord(c1), ord(c2) + 1):
                            cell = ws[f"{chr(col)}{r}"]
                            cell.fill = box_title_fill if r == box_title_row else box_value_fill
                            cell.border = box_border

                ws.row_dimensions[box_title_row].height = 22
                ws.row_dimensions[box_value_row].height = 26

                # Table formatting
                header_row = table_header_row
                header_fill = PatternFill(start_color='FF424242', end_color='FF424242', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFFFF')
                thin = Side(border_style="thin", color="FFAAAAAA")
                for col_cell in ws[header_row]:
                    col_cell.fill = header_fill
                    col_cell.font = header_font
                    col_cell.alignment = Alignment(horizontal='center')
                    col_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                # Column widths
                dims = {}
                for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row):
                    for cell in row:
                        if cell.value is not None:
                            dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len(str(cell.value)))
                for col, value in dims.items():
                    ws.column_dimensions[col].width = min(max(value + 2, 10), 40)

                # Freeze header
                ws.freeze_panes = f'A{header_row+1}'

                # Alternating row colors
                fill_gray = PatternFill(start_color='FFF5F5F5', end_color='FFF5F5F5', fill_type='solid')
                for idx, row in enumerate(ws.iter_rows(min_row=header_row+1, max_row=ws.max_row), start=0):
                    if idx % 2 == 0:
                        for cell in row:
                            cell.fill = fill_gray

                # Totals row with formulas
                last_data_row = ws.max_row
                total_row = last_data_row + 1
                ws.cell(row=total_row, column=1).value = "Toplam"
                ws.cell(row=total_row, column=1).font = Font(bold=True)
                ws.cell(row=total_row, column=1).fill = PatternFill(start_color='FFE3F2FD', end_color='FFE3F2FD', fill_type='solid')

                total_cols = {"Normal", "Mesai", "Brüt", "Ekstra", "Avans", "NET"}
                for col_idx, cell in enumerate(ws[header_row], start=1):
                    label = cell.value
                    if label in total_cols:
                        col_letter = get_column_letter(col_idx)
                        ws.cell(row=total_row, column=col_idx).value = f"=SUM({col_letter}{header_row+1}:{col_letter}{last_data_row})"
                        ws.cell(row=total_row, column=col_idx).font = Font(bold=True)
                        ws.cell(row=total_row, column=col_idx).fill = PatternFill(start_color='FFE3F2FD', end_color='FFE3F2FD', fill_type='solid')
                        ws.cell(row=total_row, column=col_idx).alignment = Alignment(horizontal='right')

                for cell in ws[total_row]:
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            return {"status": "ok", "path": path}  # WHY: signal success with output path.

        def _done(result):  # WHY: handle export completion on UI thread.
            if not result:
                return  # WHY: no payload, nothing to show.
            if result.get("status") == "empty":
                QMessageBox.information(self, "Bilgi", "Seçilen aralıkta kayıt bulunamadı.")  # WHY: preserve prior empty-data message.
                return
            if result.get("status") == "cancelled":
                return  # WHY: skip success dialog on cancel.

            # Offer to open the generated file or show folder
            if os.path.exists(path):
                if QMessageBox.question(self, "Aç", "Rapor kaydedildi. Dosyayı şimdi açmak ister misiniz?", QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
                    try:
                        os.startfile(os.path.abspath(path))
                    except Exception:
                        QMessageBox.information(self, "Bilgi", f"Dosya kaydedildi: {path}")
                else:
                    if QMessageBox.question(self, "Klasöre Git", "Klasörü açmak ister misiniz?", QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
                        try:
                            os.startfile(os.path.dirname(os.path.abspath(path)))
                        except Exception:
                            QMessageBox.information(self, "Bilgi", f"Dosya kaydedildi: {path}")

        self._start_export_worker(_task, done_cb=_done, label="Rapor hazırlanıyor...")  # WHY: run export in background.
