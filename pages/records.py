from datetime import datetime
from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, 
                             QTableWidgetItem, QHeaderView, QPushButton, 
                             QLabel, QComboBox, QLineEdit, QApplication, QMenu, QCheckBox, QFrame, QMessageBox, QFileDialog, QDialog, QDialogButtonBox, QFormLayout, QDateEdit, QAbstractItemView)
from PySide6.QtWidgets import QProgressDialog  # WHY: show non-blocking export progress for long tasks.
from PySide6.QtCore import Qt, QPoint, QTimer
from PySide6.QtCore import QThread, Signal, Slot, QObject  # WHY: background export workers keep UI responsive.
from PySide6.QtGui import QColor, QKeySequence, QAction

from core.database import Database
from core.hesaplama import parse_time_to_minutes, SABAH_TOLERANS_DK
from core.user_config import load_config, save_config
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import pandas as pd
import calendar
import os

# optional: openpyxl is used for Excel styling; the code will advise if it's missing
try:
    import openpyxl
except Exception:
    openpyxl = None

class ExportDialog(QDialog):
    def __init__(self, db: Database, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Excel Dışa Aktar — Ayarlar")
        self.setFixedSize(420, 260)
        layout = QVBoxLayout(self)

        # Date range
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        today = datetime.now()
        self.date_from.setDate(today)
        self.date_to.setDate(today)

        # Team and person
        self.combo_team = QComboBox()
        self.combo_team.addItem("Tüm Ekipler")
        teams = db.get_unique_teams()
        self.combo_team.addItems(teams)

        self.combo_person = QComboBox()
        self.combo_person.addItem("Tüm Personel")
        # Fill with all personnel
        persons = [p[0] for p in db.get_all_personnel_detailed()]
        for p in persons:
            self.combo_person.addItem(p)

        # Options
        self.chk_formulas = QCheckBox("Toplamlarda Excel Formülü Kullan")
        self.chk_logo = QCheckBox("Başlığa logo ekle")
        self.btn_logo = QPushButton("Logo Seç")
        self.logo_path = None
        def pick_logo():
            cfg = load_config()
            last_dir = cfg.get("last_logo_dir", "")
            path, _ = QFileDialog.getOpenFileName(self, "Logo Seç", last_dir, "Image (*.png *.jpg *.jpeg)")
            if path:
                self.logo_path = path
                self.btn_logo.setText("Logo Seçildi")
                try:
                    cfg["last_logo_dir"] = os.path.dirname(path)
                    save_config(cfg)
                except Exception:
                    pass
        self.btn_logo.clicked.connect(pick_logo)

        form_layout = QFormLayout()
        form_layout.addRow("Başlangıç:", self.date_from)
        form_layout.addRow("Bitiş:", self.date_to)
        form_layout.addRow("Ekip:", self.combo_team)
        form_layout.addRow("Personel:", self.combo_person)
        form_layout.addRow(self.chk_formulas)
        form_layout.addRow(self.chk_logo, self.btn_logo)
        layout.addLayout(form_layout)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def get_values(self):
        return {
            'date_from': self.date_from.date().toPython().strftime('%Y-%m-%d'),
            'date_to': self.date_to.date().toPython().strftime('%Y-%m-%d'),
            'team': None if self.combo_team.currentText() == 'Tüm Ekipler' else self.combo_team.currentText(),
            'person': None if self.combo_person.currentText() == 'Tüm Personel' else self.combo_person.currentText(),
            'formulas': self.chk_formulas.isChecked(),
            'logo': self.logo_path
        }


def get_pdf_fonts():
    """Register a Unicode font for Turkish characters if available."""
    registered = pdfmetrics.getRegisteredFontNames()
    if "AppFont" in registered:
        if "AppFont-Bold" in registered:
            return "AppFont", "AppFont-Bold"
        return "AppFont", "AppFont"
    candidates = [
        ("AppFont", "AppFont-Bold", os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "arial.ttf"), os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "arialbd.ttf")),
        ("AppFont", "AppFont-Bold", os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "segoeui.ttf"), os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "segoeuib.ttf")),
        ("AppFont", "AppFont-Bold", os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "tahoma.ttf"), os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "tahomabd.ttf")),
        ("AppFont", "AppFont-Bold", os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "calibri.ttf"), os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "calibrib.ttf")),
        ("AppFont", "AppFont-Bold", os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "DejaVuSans.ttf"), os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "DejaVuSans-Bold.ttf")),
    ]
    for reg_name, bold_name, reg_path, bold_path in candidates:
        if os.path.exists(reg_path):
            try:
                pdfmetrics.registerFont(TTFont(reg_name, reg_path))
                if os.path.exists(bold_path):
                    pdfmetrics.registerFont(TTFont(bold_name, bold_path))
                    return reg_name, bold_name
                return reg_name, reg_name
            except Exception:
                continue
    return "Helvetica", "Helvetica-Bold"


class ExportWorker(QObject):  # WHY: run export tasks off the UI thread without changing existing logic.
    finished = Signal(object)  # WHY: return payload (path/status) back to UI safely.
    error = Signal(str)  # WHY: surface export errors without crashing the UI thread.

    def __init__(self, task_fn):  # WHY: keep worker generic for different export tasks.
        super().__init__()  # WHY: initialize QObject for Qt signal/slot handling.
        self._task_fn = task_fn  # WHY: store export task to run in background.
        self._stop_requested = False  # WHY: allow safe cancel handling from dialog.

    def request_stop(self):  # WHY: allow UI to request cooperative stop.
        self._stop_requested = True  # WHY: set stop flag without killing thread.

    def should_stop(self):  # WHY: provide a shared stop check for export loops.
        return self._stop_requested or QThread.currentThread().isInterruptionRequested()  # WHY: honor both custom and Qt interruption flags.

    @Slot()
    def run(self):  # WHY: entry point for QThread.start().
        try:
            result = self._task_fn(self)  # WHY: execute export task with stop-aware worker handle.
            self.finished.emit(result)  # WHY: notify UI on completion with payload.
        except Exception as e:
            self.error.emit(str(e))  # WHY: forward exception text to UI thread safely.


class BulkEditDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Toplu Düzenleme")
        self.setFixedSize(380, 300)
        layout = QFormLayout(self)

        self.chk_giris = QCheckBox("Giriş")
        self.input_giris = QLineEdit()
        layout.addRow(self.chk_giris, self.input_giris)

        self.chk_cikis = QCheckBox("Çıkış")
        self.input_cikis = QLineEdit()
        layout.addRow(self.chk_cikis, self.input_cikis)

        self.chk_kayip = QCheckBox("Kayıp")
        self.input_kayip = QLineEdit()
        layout.addRow(self.chk_kayip, self.input_kayip)

        self.chk_normal = QCheckBox("Normal")
        self.input_normal = QLineEdit()
        layout.addRow(self.chk_normal, self.input_normal)

        self.chk_mesai = QCheckBox("Mesai")
        self.input_mesai = QLineEdit()
        layout.addRow(self.chk_mesai, self.input_mesai)

        self.chk_aciklama = QCheckBox("Açıklama")
        self.input_aciklama = QLineEdit()
        layout.addRow(self.chk_aciklama, self.input_aciklama)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def get_values(self):
        vals = {}
        if self.chk_giris.isChecked():
            vals['giris_saati'] = self.input_giris.text()
        if self.chk_cikis.isChecked():
            vals['cikis_saati'] = self.input_cikis.text()
        if self.chk_kayip.isChecked():
            vals['kayip_sure_saat'] = self.input_kayip.text()
        if self.chk_normal.isChecked():
            vals['hesaplanan_normal'] = self.input_normal.text()
        if self.chk_mesai.isChecked():
            vals['hesaplanan_mesai'] = self.input_mesai.text()
        if self.chk_aciklama.isChecked():
            vals['aciklama'] = self.input_aciklama.text()
        return vals


class MonthDeleteDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Ayı Sil")
        self.setFixedSize(320, 160)
        layout = QVBoxLayout(self)

        form = QFormLayout()
        self.combo_year = QComboBox()
        self.combo_year.addItems([str(y) for y in range(2022, 2031)])
        self.combo_month = QComboBox()
        self.combo_month.addItems(["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran", "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"])
        self.chk_backup = QCheckBox("Silmeden önce yedek al")
        form.addRow("Yıl:", self.combo_year)
        form.addRow("Ay:", self.combo_month)
        form.addRow(self.chk_backup)
        layout.addLayout(form)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def get_values(self):
        return {
            'year': int(self.combo_year.currentText()),
            'month': self.combo_month.currentIndex() + 1,
            'backup': self.chk_backup.isChecked()
        }


class RestoreDialog(QDialog):
    def __init__(self, db: Database, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Geri Al - Silinen Kayıtlar")
        self.setFixedSize(420, 200)
        self.db = db
        layout = QVBoxLayout(self)

        self.combo_batches = QComboBox()
        batches = self.db.get_trash_batches()
        self.batch_map = {}
        for b in batches:
            bid, s, e, created_at, d_daily, d_avans = b
            label = f"ID:{bid} {s} - {e} ({d_daily} günlük, {d_avans} avans)"
            self.batch_map[label] = bid
            self.combo_batches.addItem(label)

        layout.addWidget(QLabel("Geri alınacak batch'i seçin:"))
        layout.addWidget(self.combo_batches)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def get_values(self):
        label = self.combo_batches.currentText()
        return {'batch_id': self.batch_map.get(label, None)}


class RecordsPage(QWidget):
    def __init__(self, signal_manager):
        super().__init__()
        self.db = Database()
        self.signal_manager = signal_manager
        self.tersane_id = 0
        self.aktif_firma_id = 1
        self._needs_refresh = False  # NEW: lazy-load flag to avoid heavy refresh on hidden tabs.
        self._period_initialized = False  # NEW: set default year/month only once (avoid jumps on refresh).
        self._period_signals_connected = False  # NEW: avoid duplicate signal connections.
        self._data_signal_connected = False  # NEW: avoid duplicate data_updated connections.
        self._export_thread = None  # WHY: keep background export thread reference alive.
        self._export_worker = None  # WHY: keep background export worker alive during runs.
        self._export_dialog = None  # WHY: progress dialog for export operations.
        self._export_done_cb = None  # WHY: store per-export completion callback for UI.
        self._export_cancelled = False  # WHY: track user cancel to skip success dialogs.
        self.clipboard_data = None
        self.setup_ui()

    def set_tersane_id(self, tersane_id, refresh=True):
        """Global tersane seçiciden gelen tersane_id'yi set eder ve verileri yeniler."""
        self.tersane_id = tersane_id
        self._needs_refresh = True  # NEW: mark dirty; refresh can be deferred.
        if refresh:
            self.update_view()  # WHY: only visible page refreshes to keep UI smooth.

    def update_view(self):
        """Görünür sayfa için güncel tersane verilerini yükle."""
        self._needs_refresh = False  # WHY: clear dirty flag after refresh.
        if not self._period_initialized:
            today = datetime.now()
            self.combo_year.setCurrentText(str(today.year))
            self.combo_month.setCurrentIndex(today.month - 1)
            self._period_initialized = True
        self.load_data()
        if not self._period_signals_connected:
            self.combo_year.currentTextChanged.connect(self._on_period_changed)
            self.combo_month.currentIndexChanged.connect(self._on_period_changed)
            self._period_signals_connected = True  # WHY: avoid duplicate signal connections.
        if not self._data_signal_connected:
            self.signal_manager.data_updated.connect(self._on_data_updated)  # NEW: lazy refresh to avoid hidden-tab work.
            self._data_signal_connected = True  # WHY: avoid duplicate data_updated connections.

    def refresh_if_needed(self):
        """Lazy-load için: sayfa görünür olduğunda gerekiyorsa güncelle."""
        if self._needs_refresh:
            self.update_view()

    def _on_data_updated(self):
        """Veri değiştiğinde sadece görünürsek yenile (lazy)."""
        if not self.isVisible():
            self._needs_refresh = True  # WHY: defer heavy refresh until tab is visible.
            return
        self.update_view()

    def _get_active_tersane_label(self):  # WHY: centralize export titles with active tersane name.
        if self.tersane_id and self.tersane_id > 0:  # WHY: include selected tersane in export metadata.
            tersane = self.db.get_tersane(self.tersane_id)  # WHY: fetch tersane name for display.
            return tersane['ad'] if tersane else f"ID {self.tersane_id}"  # WHY: fallback keeps exports usable if name missing.
        return "Tüm Tersaneler"  # WHY: preserve global mode labeling when no tersane selected.

    def _set_period_relative(self, offset_months):
        today = datetime.now()
        y = today.year
        m = today.month + int(offset_months)
        while m <= 0:
            m += 12
            y -= 1
        while m > 12:
            m -= 12
            y += 1
        self.combo_year.blockSignals(True)
        self.combo_month.blockSignals(True)
        self.combo_year.setCurrentText(str(y))
        self.combo_month.setCurrentIndex(m - 1)
        self.combo_year.blockSignals(False)
        self.combo_month.blockSignals(False)
        self._on_period_changed()

    def _on_period_changed(self):
        year = int(self.combo_year.currentText())
        month = self.combo_month.currentIndex() + 1
        locked = self.db.is_month_locked(year, month, self.aktif_firma_id)
        self._set_month_locked_ui(locked)
        self.load_data()

    def _set_month_locked_ui(self, locked):
        # Tüm düzenleme, silme, toplu işlem ve sağ tık işlemlerini devre dışı bırak
        self.btn_bulk_edit.setEnabled(not locked)
        self.btn_delete_month.setEnabled(not locked)
        self.btn_restore.setEnabled(not locked)
        self.btn_export.setEnabled(not locked)
        self.btn_summary_pdf.setEnabled(not locked)
        self.btn_monthly_excel.setEnabled(not locked)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers if locked else QTableWidget.AllEditTriggers)
        self.lock_banner.setVisible(locked)

    def eventFilter(self, obj, event):
        """Ctrl+C ve Ctrl+V için event filter"""
        if obj == self.table and event.type() == event.Type.KeyPress:
            if event.matches(QKeySequence.Copy):
                self.copy_selection()
                return True
            elif event.matches(QKeySequence.Paste):
                self.paste_selection()
                return True
        return super().eventFilter(obj, event)

    def copy_selection(self):
        """Seçili hücreyi kopyala"""
        selected = self.table.selectedItems()
        if selected:
            item = selected[0]
            self.clipboard_data = item.text()
            QApplication.clipboard().setText(self.clipboard_data)

    def paste_selection(self):
        """Kopyalanan veriyi seçili hücrelere yapıştır"""
        if not self.clipboard_data:
            return
        
        selected = self.table.selectedItems()
        if not selected:
            return
        
        for item in selected:
            row, col = item.row(), item.column()
            if col in [6, 7, 8, 9]:  # Sadece düzenlenebilir kolonlar
                self.table.setItem(row, col, QTableWidgetItem(self.clipboard_data))
                self.on_cell_changed(row, col)

    def setup_ui(self):
        layout = QVBoxLayout(self)

        title = QLabel("✏️ Günlük Kayıtlar")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #fff; margin-bottom: 2px;")
        layout.addWidget(title)

        desc = QLabel("Günlük giriş/çıkış, normal çalışma ve mesai kayıtlarını görüntüle ve düzenle.")
        desc.setStyleSheet("color: #999; font-size: 12px; margin-bottom: 10px;")
        layout.addWidget(desc)

        self.lock_banner = QFrame()
        self.lock_banner.setStyleSheet("""
            QFrame {
                background-color: #E65100;
                border-radius: 6px;
                padding: 8px 16px;
            }
        """)
        self.lock_banner.setVisible(False)
        banner_layout = QHBoxLayout(self.lock_banner)
        banner_layout.setContentsMargins(8, 4, 8, 4)
        self.lbl_lock_msg = QLabel("🔒 Bu ay kilitlidir. Değişiklik yapılamaz.")
        self.lbl_lock_msg.setStyleSheet("color: white; font-weight: bold; font-size: 12px;")
        banner_layout.addWidget(self.lbl_lock_msg)
        banner_layout.addStretch()
        layout.addWidget(self.lock_banner)

        # FİLTRE PANELİ
        period_bar = QHBoxLayout()
        filter_bar = QHBoxLayout()
        action_bar = QHBoxLayout()
        
        self.combo_year = QComboBox()
        self.combo_year.addItems([str(y) for y in range(2024, 2030)])
        self.combo_year.setFixedWidth(70)
        
        self.combo_month = QComboBox()
        self.combo_month.addItems(["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran", 
                                   "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"])
        self.combo_month.setFixedWidth(100)

        # YENİ FİLTRELER
        filter_frame = QFrame()
        filter_frame.setStyleSheet("background-color: #2b2b2b; padding: 8px; border-radius: 5px;")
        filter_layout = QHBoxLayout(filter_frame)
        filter_layout.setSpacing(4)
        
        self.chk_show_empty = QCheckBox("Boş Kayıtları Göster")
        self.chk_show_empty.setChecked(True)
        self.chk_show_empty.setStyleSheet("font-size: 11px;")
        self.chk_show_empty.stateChanged.connect(self.filter_table)
        
        self.chk_only_empty = QCheckBox("Sadece Boş")
        self.chk_only_empty.setStyleSheet("font-size: 11px;")
        self.chk_only_empty.stateChanged.connect(self.filter_table)
        
        self.chk_only_weekend = QCheckBox("Sadece Haftasonu")
        self.chk_only_weekend.setStyleSheet("font-size: 11px;")
        self.chk_only_weekend.stateChanged.connect(self.filter_table)
        
        self.chk_only_special = QCheckBox("Sadece Özel Durum")
        self.chk_only_special.setStyleSheet("font-size: 11px;")
        self.chk_only_special.stateChanged.connect(self.filter_table)
        
        filter_layout.addWidget(self.chk_show_empty)
        filter_layout.addWidget(self.chk_only_empty)
        filter_layout.addWidget(self.chk_only_weekend)
        filter_layout.addWidget(self.chk_only_special)
        filter_layout.addStretch()

        # EKİP FİLTRESİ
        self.combo_team = QComboBox()
        self.combo_team.addItem("Tüm Ekipler")
        self.combo_team.setFixedWidth(120)
        self.combo_team.currentTextChanged.connect(self.filter_table)

        # İSİM ARAMA
        self.search_name = QLineEdit()
        self.search_name.setPlaceholderText("👤 İsim Ara...")
        self.search_name.setStyleSheet("padding: 5px; color: white; background-color: #424242;")
        self.search_name.textChanged.connect(self._on_search_changed)

        # GÜN ARAMA
        self.search_date = QLineEdit()
        self.search_date.setPlaceholderText("📅 Gün (01, 15..)")
        self.search_date.setStyleSheet("padding: 5px; color: white; background-color: #424242;")
        self.search_date.setFixedWidth(100)
        self._search_timer = QTimer(self)
        self._search_timer.setSingleShot(True)
        self._search_timer.setInterval(250)
        self._search_timer.timeout.connect(self.filter_table)
        self.search_date.textChanged.connect(self._on_search_changed)

        # Günlük sayaçlar
        self.lbl_daily_count = QLabel("📋 Yevmiye: 0")
        self.lbl_daily_count.setStyleSheet("color: #ddd; padding: 4px 8px; background-color: #333; border-radius: 4px;")
        self.lbl_warn = QLabel("⚠️ Geç: 0 | Eksik Çıkış: 0")
        self.lbl_warn.setStyleSheet("color: #ddd; padding: 4px 8px; background-color: #333; border-radius: 4px;")

        period_bar.addWidget(QLabel("Dönem:"))
        period_bar.addWidget(self.combo_month)
        period_bar.addWidget(self.combo_year)
        self.btn_this_month = QPushButton("Bu Ay")
        self.btn_this_month.setFixedWidth(70)
        self.btn_this_month.setStyleSheet("""
            QPushButton {
                background-color: #333; color: #ccc; border: 1px solid #555;
                border-radius: 4px; padding: 4px 12px;
            }
            QPushButton:hover { background-color: #444; color: white; }
        """)
        self.btn_this_month.clicked.connect(lambda: self._set_period_relative(0))
        period_bar.addWidget(self.btn_this_month)
        self.btn_prev_month = QPushButton("Geçen Ay")
        self.btn_prev_month.setFixedWidth(80)
        self.btn_prev_month.setStyleSheet("""
            QPushButton {
                background-color: #333; color: #ccc; border: 1px solid #555;
                border-radius: 4px; padding: 4px 12px;
            }
            QPushButton:hover { background-color: #444; color: white; }
        """)
        self.btn_prev_month.clicked.connect(lambda: self._set_period_relative(-1))
        period_bar.addWidget(self.btn_prev_month)
        period_bar.addStretch()

        filter_bar.addWidget(self.combo_team)
        filter_bar.addWidget(self.search_name)
        filter_bar.addWidget(self.search_date)
        filter_bar.addStretch()
        filter_bar.addWidget(self.lbl_daily_count)
        filter_bar.addWidget(self.lbl_warn)

        # Export button
        self.btn_export = QPushButton("📤 Excel Olarak Dışa Aktar")
        self.btn_export.setFixedWidth(180)
        self.btn_export.clicked.connect(self.export_to_excel)
        action_bar.addWidget(self.btn_export)

        # Aylık özet PDF
        self.btn_summary_pdf = QPushButton("📄 Aylık Özet PDF")
        self.btn_summary_pdf.setFixedWidth(150)
        self.btn_summary_pdf.clicked.connect(self.export_monthly_summary_pdf)
        action_bar.addWidget(self.btn_summary_pdf)

        # Aylık Excel
        self.btn_monthly_excel = QPushButton("📊 Aylık Excel")
        self.btn_monthly_excel.setFixedWidth(130)
        self.btn_monthly_excel.clicked.connect(self.export_monthly_excel)
        action_bar.addWidget(self.btn_monthly_excel)

        # Toplu düzenleme
        self.btn_bulk_edit = QPushButton("🧮 Toplu Düzenle")
        self.btn_bulk_edit.setFixedWidth(140)
        self.btn_bulk_edit.clicked.connect(self.open_bulk_edit)
        action_bar.addWidget(self.btn_bulk_edit)

        # Delete month button
        self.btn_delete_month = QPushButton("🗑️ Ayı Sil")
        self.btn_delete_month.setFixedWidth(120)
        self.btn_delete_month.clicked.connect(self.delete_month_dialog)
        action_bar.addWidget(self.btn_delete_month)

        # Restore (undo) button
        self.btn_restore = QPushButton("↺ Geri Al")
        self.btn_restore.setFixedWidth(100)
        self.btn_restore.clicked.connect(self.restore_dialog)
        action_bar.addWidget(self.btn_restore)
        action_bar.addStretch()

        layout.addLayout(period_bar)
        layout.addLayout(filter_bar)
        layout.addLayout(action_bar)
        layout.addWidget(filter_frame)

        # RENKLER AÇIKLAMASI
        legend_label = QLabel("🟢 Normal Çalışma Günü (Pzt-Cum)  |  ⚫ Cumartesi  |  🟡 Pazar & Resmi Tatiller")
        legend_label.setStyleSheet("color: #FFFFFF; padding: 8px; background-color: #2a2a2a; border-radius: 4px; font-size: 11px;")
        layout.addWidget(legend_label)

        # TABLO
        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels(["ID", "Tarih", "Personel", "Ekip", "Giriş", "Çıkış", 
                                            "Kayıp", "Normal", "Mesai", "Açıklama"])
        self.table.setColumnHidden(0, True)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.table.setSortingEnabled(True)  # Sütunlara tıklayarak sırala
        self.table.setAlternatingRowColors(True)
        self.table.installEventFilter(self)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        
        self.table.setStyleSheet("""
            QTableWidget { 
                background-color: #212121; 
                color: white; 
                gridline-color: #424242; 
                border: none;
                alternate-background-color: #2a2a2a;
            }
            QHeaderView::section { 
                background-color: #424242; 
                color: white; 
                padding: 4px; 
                border: 1px solid #616161; 
            }
            QTableWidget::item:selected { 
                background-color: #1565C0; 
            }
        """)
        self.table.cellChanged.connect(self.on_cell_changed)
        layout.addWidget(self.table)
        
        info_lbl = QLabel("💡 Sağ Tık: Hızlı düzenleme | Ctrl+C/V: Kopyala/Yapıştır | Shift/Ctrl+Tıkla: Çoklu seçim")
        info_lbl.setStyleSheet("color: #888; font-size: 11px;")
        layout.addWidget(info_lbl)

    def show_context_menu(self, pos: QPoint):
        year = int(self.combo_year.currentText())
        month = self.combo_month.currentIndex() + 1
        if self.db.is_month_locked(year, month, self.aktif_firma_id):
            QMessageBox.warning(self, "Ay Kilitli", "Bu ay kilitlidir. Sağ tık işlemleri devre dışı.")
            return
        # ...eski context menu kodu...
        from PySide6.QtWidgets import QMenu, QMessageBox
        from PySide6.QtGui import QAction
        item = self.table.itemAt(pos)
        if not item:
            return
        row = item.row()
        self.table.clearSelection()
        self.table.selectRow(row)
        id_item = self.table.item(row, 0)
        if not id_item or not id_item.text():
            QMessageBox.warning(self, "Hata", "Seçili satırın ID bilgisi bulunamadı.")
            return
        rec_id = id_item.text()
        date_item = self.table.item(row, 1)
        ad_soyad_item = self.table.item(row, 2)
        if not date_item or not ad_soyad_item:
            return
        tarih_iso = date_item.data(Qt.UserRole)
        ad_soyad = ad_soyad_item.text()
        import re
        if not tarih_iso or not re.match(r"^\d{4}-\d{2}-\d{2}$", str(tarih_iso)):
            QMessageBox.warning(self, "Hatalı Tarih", "Tarih formatı geçersiz veya boş. Sağ tık işlemi uygulanamaz.")
            return
        menu = QMenu(self)
        act_full_day = QAction("📅 Tam Gün Uygula", self)
        act_reset = QAction("⚡ Sıfırla", self)
        def do_full_day():
            try:
                pinfo = self.db.get_personnel(ad_soyad)
                yevmiyeci = pinfo['yevmiyeci_mi'] if pinfo and 'yevmiyeci_mi' in pinfo else 0
                hesaplanan_normal = 1.0 if yevmiyeci else 7.5
                with self.db.get_connection() as conn:
                    cur = conn.cursor()
                    prev = cur.execute(
                        "SELECT giris_saati, cikis_saati, kayip_sure_saat, aciklama FROM gunluk_kayit WHERE id=?",
                        (rec_id,)
                    ).fetchone()
                    prev_giris, prev_cikis, prev_kayip, prev_aciklama = prev or ("", "", "", "")

                    # Mesai'ye dokunma: sadece normal'i tam gun yap, eksik zaman alanlarini doldur.
                    giris = prev_giris or "08:20"
                    cikis = prev_cikis or "17:30"
                    kayip = prev_kayip or "00:00"

                    aciklama = "Tam Gün (Manuel)"
                    if prev_aciklama and str(prev_aciklama).strip() and str(prev_aciklama).strip() != aciklama:
                        aciklama = f"{prev_aciklama} | {aciklama}"

                    cur.execute(
                        "UPDATE gunluk_kayit SET giris_saati=?, cikis_saati=?, kayip_sure_saat=?, hesaplanan_normal=?, aciklama=? WHERE id=?",
                        (giris, cikis, kayip, hesaplanan_normal, aciklama, rec_id)
                    )
                    conn.commit()
                self.load_data()
                self.signal_manager.data_updated.emit()
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Tam Gün Uygula işleminde hata: {e}")
        def do_reset():
            try:
                self.table.setItem(row, 4, QTableWidgetItem(""))
                self.table.setItem(row, 5, QTableWidgetItem(""))
                self.table.setItem(row, 6, QTableWidgetItem(""))
                self.table.setItem(row, 7, QTableWidgetItem("0.0"))
                self.table.setItem(row, 8, QTableWidgetItem("0.0"))
                self.table.setItem(row, 9, QTableWidgetItem("Sıfırlandı"))
                with self.db.get_connection() as conn:
                    conn.execute(
                        "UPDATE gunluk_kayit SET giris_saati=?, cikis_saati=?, kayip_sure_saat=?, "
                        "hesaplanan_normal=?, hesaplanan_mesai=?, aciklama=? WHERE id=?",
                        ("", "", "", 0.0, 0.0, "Sıfırlandı", rec_id)
                    )
                    conn.commit()
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Sıfırla işleminde hata: {e}")
        act_full_day.triggered.connect(do_full_day)
        act_reset.triggered.connect(do_reset)
        menu.addAction(act_full_day)
        menu.addAction(act_reset)
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def delete_month_dialog(self):
        dlg = MonthDeleteDialog(self)
        if dlg.exec_() != QDialog.Accepted:
            return
        vals = dlg.get_values()
        year, month, do_backup = vals['year'], vals['month'], vals['backup']
        # compute date range
        last_day = calendar.monthrange(year, month)[1]
        start_date = f"{year}-{month:02d}-01"
        end_date = f"{year}-{month:02d}-{last_day:02d}"
        # Onay icin etkilenecek kayit sayilari
        try:
            with self.db.get_connection() as conn:
                c = conn.cursor()
                daily_count = c.execute("SELECT COUNT(*) FROM gunluk_kayit WHERE tarih BETWEEN ? AND ?", (start_date, end_date)).fetchone()[0]
                avans_count = c.execute("SELECT COUNT(*) FROM avans_kesinti WHERE tarih BETWEEN ? AND ?", (start_date, end_date)).fetchone()[0]
        except Exception:
            daily_count, avans_count = None, None

        # Otomatik yedek (her zaman merkezi yola)
        from core.database import get_default_db_path
        backup_folder = str(get_default_db_path().parent)
        ok, path_or_err = self.db.backup_db(backup_folder)
        if ok:
            QMessageBox.information(self, "Yedekleme", f"Otomatik yedek alındı: {path_or_err}")
        else:
            QMessageBox.warning(self, "Yedekleme Hatası", f"Otomatik yedek başarısız: {path_or_err}")

        # Ekstra: Kullanıcı isterse farklı klasöre de yedek alabilir
        if do_backup:
            cfg = load_config()
            last_dir = cfg.get("last_backup_dir", "")
            folder = QFileDialog.getExistingDirectory(self, "Ekstra Yedek Klasoru Sec", last_dir)
            if folder:
                try:
                    cfg["last_backup_dir"] = folder
                    save_config(cfg)
                except Exception:
                    pass
                ok2, path2 = self.db.backup_db(folder)
                if ok2:
                    QMessageBox.information(self, "Ekstra Yedek", f"Ekstra yedek alindi: {path2}")
                else:
                    QMessageBox.warning(self, "Ekstra Yedek Hatasi", f"Ekstra yedek basarisiz: {path2}")
        # Final confirmation
        msg = f"{year} {month} ayindaki tum kayitlar silinecek."
        if daily_count is not None and avans_count is not None:
            msg += f"\nGunluk kayit: {daily_count}, Avans/Kesinti: {avans_count}"
        msg += "\nDevam edilsin mi?"
        if QMessageBox.question(self, "Onay", msg, QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return

        try:
            tersane_filter = self.tersane_id if self.tersane_id and self.tersane_id > 0 else None
            batch_id, deleted_daily, deleted_avans = self.db.move_records_to_trash(
                start_date, end_date,
                firma_id=self.aktif_firma_id,
                tersane_id=tersane_filter
            )
            QMessageBox.information(self, "Silindi", f"Silme başarılı: {deleted_daily} günlük kayıt, {deleted_avans} avans/kesinti silindi. (Batch {batch_id})\nGeri almak için 'Geri Al' butonunu kullanın.")
            # Refresh
            self.signal_manager.data_updated.emit()
            self.load_data()
        except Exception as e:
            from core.app_logger import log_error
            import traceback
            tb = traceback.format_exc()
            log_error(f"Silme Hatası: {e}\n{tb}")
            QMessageBox.critical(self, "Hata", f"Silme sırasında hata: {e}")

    def restore_dialog(self):
        dlg = RestoreDialog(self.db, self)
        if dlg.exec_() != QDialog.Accepted:
            return
        vals = dlg.get_values()
        batch_id = vals.get('batch_id')
        if not batch_id:
            QMessageBox.warning(self, "Hata", "Geri alınacak batch seçilmedi.")
            return
        # Onay icin batch detaylari
        try:
            with self.db.get_connection() as conn:
                row = conn.execute("SELECT deleted_daily, deleted_avans FROM trash_batches WHERE id=?", (batch_id,)).fetchone()
                d_daily = row[0] if row else None
                d_avans = row[1] if row else None
        except Exception:
            d_daily, d_avans = None, None
        msg = f"Batch {batch_id} geri yuklenecek."
        if d_daily is not None and d_avans is not None:
            msg += f"\nGunluk kayit: {d_daily}, Avans/Kesinti: {d_avans}"
        msg += "\nDevam edilsin mi?"
        if QMessageBox.question(self, "Onay", msg, QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return
            return
        try:
            restored_daily, restored_avans = self.db.restore_trash_batch(batch_id)
            QMessageBox.information(self, "Geri Yüklendi", f"Geri yükleme tamamlandı: {restored_daily} günlük kayıt, {restored_avans} avans/kesinti geri yüklendi.")
            self.signal_manager.data_updated.emit()
            self.load_data()
        except Exception as e:
            from core.app_logger import log_error
            import traceback
            tb = traceback.format_exc()
            log_error(f"Geri Yükleme Hatası: {e}\n{tb}")
            QMessageBox.critical(self, "Hata", f"Geri yükleme sırasında hata: {e}")

    def apply_to_selected(self, action_type):
        """Seçili satırlara işlem uygula (transactional, hesaplama.py ile)"""
        try:
            selected_rows = set(idx.row() for idx in self.table.selectionModel().selectedRows())
            self.table.blockSignals(True)
            from core.hesaplama import hesapla_hakedis
            updates = []
            holiday_set = set(self.db.get_holidays())
            # NEW: use active tersane settings to keep calculations consistent across shipyards.
            try:
                settings_cache = self.db.get_settings_cache(tersane_id=self.tersane_id) if self.tersane_id else self.db.get_settings_cache()
            except Exception:
                settings_cache = None  # SAFEGUARD: fall back to legacy behavior if cache fails.
            for row in selected_rows:
                id_item = self.table.item(row, 0)
                if not id_item or not id_item.text():
                    continue
                rec_id = id_item.text()
                tarih = self.table.item(row, 1).text()
                ad_soyad = self.table.item(row, 2).text()
                # Varsayılan olarak mevcut değerleri oku
                giris = self.table.item(row, 4).text() if self.table.item(row, 4) else ""
                cikis = self.table.item(row, 5).text() if self.table.item(row, 5) else ""
                kayip = self.table.item(row, 6).text() if self.table.item(row, 6) else ""
                # DB'den yevmiyeci ve özel_durum çek
                pinfo = self.db.get_personnel(ad_soyad)
                yevmiyeci = pinfo['yevmiyeci_mi'] if pinfo and 'yevmiyeci_mi' in pinfo else 0
                ozel_durum = pinfo['ozel_durum'] if pinfo and 'ozel_durum' in pinfo else None
                def holiday_info_func(tarih_str):
                    return self.db.get_holiday_info(tarih_str)
                def special_status_func(ad):
                    return ozel_durum
                # Hangi işlem?
                if action_type == "full_day":
                    normal = 1.0 if yevmiyeci else 7.5
                    try:
                        mesai = float(self.table.item(row, 8).text() or 0)
                    except Exception:
                        mesai = 0.0
                    desc = "Tam Gün (Merkezi)"
                    self.table.setItem(row, 7, QTableWidgetItem(str(normal)))
                    self.table.setItem(row, 8, QTableWidgetItem(str(mesai)))
                    self.table.setItem(row, 9, QTableWidgetItem(desc))
                    updates.append((normal, mesai, desc, rec_id))
                    continue
                elif action_type == "sunday":
                    giris, cikis, kayip = "", "", ""
                elif action_type == "holiday":
                    giris, cikis, kayip = "", "", ""
                elif action_type == "half_day":
                    giris, cikis, kayip = "", "", ""
                elif action_type == "reset":
                    normal, mesai, desc = 0.0, 0.0, "Sıfırlandı"
                    updates.append((normal, mesai, desc, rec_id))
                    self.table.setItem(row, 7, QTableWidgetItem(str(normal)))
                    self.table.setItem(row, 8, QTableWidgetItem(str(mesai)))
                    self.table.setItem(row, 9, QTableWidgetItem(desc))
                    continue
                # Hesaplama
                normal, mesai, desc = hesapla_hakedis(
                    tarih, giris, cikis, kayip, holiday_set, holiday_info_func, special_status_func, ad_soyad, yevmiyeci, db=self.db,
                    settings_cache=settings_cache.get('shipyard_rules', settings_cache) if settings_cache else None)  # NEW: shipyard_rules dict.
                if action_type == "half_day":
                    normal, mesai = normal/2, mesai/2
                    desc = "Yarım Gün (Merkezi)"
                # Tabloyu güncelle
                self.table.setItem(row, 7, QTableWidgetItem(str(normal)))
                self.table.setItem(row, 8, QTableWidgetItem(str(mesai)))
                self.table.setItem(row, 9, QTableWidgetItem(desc))
                updates.append((normal, mesai, desc, rec_id))
            # Toplu DB güncelle
            if updates:
                self.db.bulk_update_hakedis(updates)
            self.table.blockSignals(False)
            self.signal_manager.data_updated.emit()
        except Exception as e:
            import traceback
            from core.app_logger import log_error
            log_error(f"apply_to_selected hata: {e}\n{traceback.format_exc()}")

    def load_data(self):
        self.table.blockSignals(True)
        # Sorting while populating can move rows mid-fill (e.g. sorted by "Çıkış"),
        # causing later columns to appear empty/misaligned. Populate with sorting off.
        _sorting_enabled = self.table.isSortingEnabled()
        if _sorting_enabled:
            self.table.setSortingEnabled(False)
        self.table.setRowCount(0)
        
        # Ekip listesini güncelle
        current_team = self.combo_team.currentText()
        self.combo_team.blockSignals(True)
        self.combo_team.clear()
        self.combo_team.addItem("Tüm Ekipler")
        teams = self.db.get_unique_teams()
        self.combo_team.addItems(teams)
        self.combo_team.setCurrentText(current_team)
        self.combo_team.blockSignals(False)

        try:
            year = int(self.combo_year.currentText())
            month = self.combo_month.currentIndex() + 1
            records = self.db.get_records_by_month(year, month, tersane_id=self.tersane_id)
            self.table.setRowCount(len(records))
            
            with self.db.get_connection() as conn:
                c = conn.cursor()
                c.execute("SELECT tarih FROM resmi_tatiller")
                holiday_set = {row[0] for row in c.fetchall()}
            
            for r, row in enumerate(records):
                self.table.setItem(r, 0, QTableWidgetItem(str(row[0])))
                try:
                    dt = datetime.strptime(row[1], "%Y-%m-%d")
                    gun_adi = dt.strftime("%d %b %a")
                    is_weekend = dt.weekday() >= 5
                except (ValueError, TypeError):
                    dt, gun_adi, is_weekend = None, row[1], False
                is_holiday = False
                try:
                    dt = datetime.strptime(row[1], "%Y-%m-%d")
                    ay_gun = dt.strftime("%m-%d")
                    is_holiday = ay_gun in holiday_set
                except (ValueError, TypeError):
                    is_holiday = False
                if is_holiday or "Pazar" in (row[8] or ""):
                    bg_color = QColor("#FF6F00")
                elif is_weekend:
                    bg_color = QColor("#424242")
                else:
                    bg_color = QColor("#1B5E20")
                for col in range(1, 10):
                    if col == 1:
                        t_item = QTableWidgetItem(gun_adi)
                        t_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                        # ISO tarihi UserRole ile sakla
                        t_item.setData(Qt.UserRole, row[1])
                    elif col == 2:
                        t_item = QTableWidgetItem(row[2])
                        t_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                    elif col == 3:
                        t_item = QTableWidgetItem(row[9] if row[9] else "-")
                        t_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                    elif col == 4:
                        t_item = QTableWidgetItem(str(row[3] or ""))
                        t_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                    elif col == 5:
                        t_item = QTableWidgetItem(str(row[4] or ""))
                        t_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                    elif col == 6:
                        t_item = QTableWidgetItem(str(row[5] or ""))
                    elif col == 7:
                        t_item = QTableWidgetItem(str(row[6]))
                    elif col == 8:
                        t_item = QTableWidgetItem(str(row[7]))
                    elif col == 9:
                        t_item = QTableWidgetItem(str(row[8] or ""))
                    
                    t_item.setBackground(bg_color)
                    self.table.setItem(r, col, t_item)
                    
        except Exception as e:
            from core.app_logger import log_error
            log_error(f"Tablo yükleme hatası: {e}")
        self.table.blockSignals(False)
        if _sorting_enabled:
            self.table.setSortingEnabled(True)
        self.filter_table()

    def _on_search_changed(self):
        if hasattr(self, '_search_timer') and self._search_timer:
            self._search_timer.start()

    def filter_table(self):
        name_text = self.search_name.text().lower()
        date_text = self.search_date.text().lower()
        team_filter = self.combo_team.currentText()
        
        show_empty = self.chk_show_empty.isChecked()
        only_empty = self.chk_only_empty.isChecked()
        only_weekend = self.chk_only_weekend.isChecked()
        only_special = self.chk_only_special.isChecked()
        
        yevmiye_count = 0
        late_count = 0
        missing_exit_count = 0
        for i in range(self.table.rowCount()):
            row_name = self.table.item(i, 2).text().lower()
            row_date = self.table.item(i, 1).text().lower()
            row_team = self.table.item(i, 3).text()
            row_giris = self.table.item(i, 4).text()
            row_cikis = self.table.item(i, 5).text()
            row_aciklama = self.table.item(i, 9).text()
            try:
                row_normal = float(self.table.item(i, 7).text() or 0)
            except Exception:
                row_normal = 0.0
            
            # Boş mu?
            # Boş kayıt: giriş/çıkış boş VE normal saat boş/0
            is_empty = (not row_giris or not row_cikis) and row_normal == 0
            
            # Haftasonu mu?
            try:
                tarih_str = self.table.item(i, 1).text().split()[0]
                year = int(self.combo_year.currentText())
                dt = datetime.strptime(f"{year}-{self.combo_month.currentIndex()+1:02d}-{tarih_str}", "%Y-%m-%d")
                is_weekend = dt.weekday() >= 5
            except (ValueError, TypeError):
                is_weekend = False
            
            # Özel durum mu?
            is_special = "Özel Durum" in row_aciklama
            
            # Filtre kontrolleri
            match_name = name_text in row_name
            match_date = date_text in row_date
            match_team = (team_filter == "Tüm Ekipler") or (row_team == team_filter)
            
            match_empty = True
            if only_empty:
                match_empty = is_empty
            elif not show_empty:
                match_empty = not is_empty
            
            match_weekend = (not only_weekend) or is_weekend
            match_special = (not only_special) or is_special
            
            if match_name and match_date and match_team and match_empty and match_weekend and match_special:
                self.table.setRowHidden(i, False)
                if row_normal > 0:
                    if not date_text or date_text in row_date:
                        yevmiye_count += 1
                if row_giris and not row_cikis:
                    missing_exit_count += 1
                if row_giris:
                    giris_dk = parse_time_to_minutes(row_giris)
                    if giris_dk is not None and giris_dk > SABAH_TOLERANS_DK:
                        late_count += 1
            else:
                self.table.setRowHidden(i, True)

        if date_text:
            self.lbl_daily_count.setText(f"📋 Yevmiye (gün): {yevmiye_count}")
        else:
            self.lbl_daily_count.setText(f"📋 Yevmiye: {yevmiye_count}")
        self.lbl_warn.setText(f"⚠️ Geç: {late_count} | Eksik Çıkış: {missing_exit_count}")

    def open_bulk_edit(self):
        selected_rows = sorted({item.row() for item in self.table.selectedItems()})
        if not selected_rows:
            QMessageBox.information(self, "Bilgi", "Toplu düzenleme için satır seçin.")
            return

        dlg = BulkEditDialog(self)
        if dlg.exec() != QDialog.Accepted:
            return
        vals = dlg.get_values()
        if not vals:
            QMessageBox.information(self, "Bilgi", "Uygulanacak alan seçilmedi.")
            return

        from core.hesaplama import hesapla_hakedis
        updates = []
        holiday_set = set(self.db.get_holidays())
        # NEW: use active tersane settings to keep calculations consistent across shipyards.
        try:
            settings_cache = self.db.get_settings_cache(tersane_id=self.tersane_id) if self.tersane_id else self.db.get_settings_cache()
        except Exception:
            settings_cache = None  # SAFEGUARD: fall back to legacy behavior if cache fails.
        for row in selected_rows:
            rec_id = self.table.item(row, 0).text()
            tarih = self.table.item(row, 1).text()
            ad_soyad = self.table.item(row, 2).text()
            # Varsayılan olarak mevcut değerleri oku
            giris = self.table.item(row, 4).text() if self.table.item(row, 4) else ""
            cikis = self.table.item(row, 5).text() if self.table.item(row, 5) else ""
            kayip = self.table.item(row, 6).text() if self.table.item(row, 6) else ""
            # Uygulanacak alanlar güncelleniyor
            if 'giris_saati' in vals:
                giris = vals['giris_saati']
                self.table.setItem(row, 4, QTableWidgetItem(giris))
            if 'cikis_saati' in vals:
                cikis = vals['cikis_saati']
                self.table.setItem(row, 5, QTableWidgetItem(cikis))
            if 'kayip_sure_saat' in vals:
                kayip = vals['kayip_sure_saat']
                self.table.setItem(row, 6, QTableWidgetItem(kayip))
            # DB'den yevmiyeci ve özel_durum çek
            pinfo = self.db.get_personnel(ad_soyad)
            yevmiyeci = pinfo['yevmiyeci_mi'] if pinfo and 'yevmiyeci_mi' in pinfo else 0
            ozel_durum = pinfo['ozel_durum'] if pinfo and 'ozel_durum' in pinfo else None
            def holiday_info_func(tarih_str):
                return self.db.get_holiday_info(tarih_str)
            def special_status_func(ad):
                return ozel_durum
            # Hesaplama
            normal, mesai, desc = hesapla_hakedis(
                tarih, giris, cikis, kayip, holiday_set, holiday_info_func, special_status_func, ad_soyad, yevmiyeci, db=self.db,
                settings_cache=settings_cache.get('shipyard_rules', settings_cache) if settings_cache else None)  # NEW: shipyard_rules dict.
            # Manuel override
            if 'hesaplanan_normal' in vals:
                try:
                    normal = float(vals['hesaplanan_normal'])
                except (ValueError, TypeError):
                    pass
            if 'hesaplanan_mesai' in vals:
                try:
                    mesai = float(vals['hesaplanan_mesai'])
                except (ValueError, TypeError):
                    pass
            if 'aciklama' in vals:
                desc = vals['aciklama']
            # Tabloyu güncelle
            self.table.setItem(row, 7, QTableWidgetItem(str(normal)))
            self.table.setItem(row, 8, QTableWidgetItem(str(mesai)))
            self.table.setItem(row, 9, QTableWidgetItem(desc))
            updates.append((normal, mesai, desc, rec_id))
        # Toplu DB güncelle
        if updates:
            self.db.bulk_update_hakedis(updates)
        self.signal_manager.data_updated.emit()

    def _start_export_worker(self, task_fn, done_cb=None, label="DÄ±ÅŸa aktarÄ±lÄ±yor..."):  # WHY: shared export runner to keep UI responsive.
        if self._export_thread and self._export_thread.isRunning():  # WHY: avoid overlapping exports that could lock files.
            QMessageBox.information(self, "Bilgi", "Devam eden bir dÄ±ÅŸa aktarma var.")  # WHY: inform user without starting another thread.
            return
        self._export_done_cb = done_cb  # WHY: keep per-export UI completion handler.
        self._export_cancelled = False  # WHY: reset cancel state for each new export.
        self._export_dialog = QProgressDialog(label, None, 0, 0, self)  # WHY: show indeterminate progress during export.
        self._export_dialog.setWindowModality(Qt.WindowModal)  # WHY: keep modal behavior consistent with other dialogs.
        self._export_dialog.setAutoClose(False)  # WHY: we close explicitly on signals to avoid stuck dialogs.
        self._export_dialog.setAutoReset(False)  # WHY: prevent auto-reset from hiding progress early.
        self._export_dialog.setMinimumDuration(0)  # WHY: show immediately to avoid perceived freeze.
        self._export_dialog.setAttribute(Qt.WA_DeleteOnClose, False)  # WHY: keep dialog alive for late signals.
        self._export_dialog.canceled.connect(self._on_export_dialog_canceled)  # WHY: allow safe cancel without crashing.
        self._export_dialog.rejected.connect(self._on_export_dialog_canceled)  # WHY: handle window close (X) safely.
        self._export_dialog.show()  # WHY: show progress feedback during background work.

        self._export_thread = QThread()  # WHY: run heavy export in background.
        worker = ExportWorker(task_fn)  # WHY: reuse generic export worker for different tasks.
        self._export_worker = worker  # WHY: keep a strong reference to prevent GC.
        worker.moveToThread(self._export_thread)  # WHY: execute worker in background thread.
        self._export_thread.started.connect(worker.run)  # WHY: start export when thread starts.
        worker.finished.connect(self._on_export_finished)  # WHY: handle success payload in UI thread.
        worker.finished.connect(self._export_dialog.accept)  # WHY: close dialog on normal completion.
        worker.finished.connect(self._export_thread.quit)  # WHY: stop thread event loop after completion.
        worker.finished.connect(worker.deleteLater)  # WHY: free worker safely.
        worker.error.connect(self._on_export_error)  # WHY: surface errors without freezing UI.
        worker.error.connect(self._export_thread.quit)  # WHY: stop thread on error.
        worker.error.connect(worker.deleteLater)  # WHY: free worker on error path.
        self._export_thread.finished.connect(self._on_export_thread_finished)  # WHY: clear refs after thread stops.
        self._export_thread.finished.connect(self._export_thread.deleteLater)  # WHY: free thread object after finish.
        self._export_thread.start()  # WHY: kick off background export.

    def _on_export_finished(self, result):  # WHY: centralize export completion handling.
        if self._export_dialog:  # WHY: close dialog if still alive.
            try:
                self._export_dialog.close()  # WHY: ensure dialog closes on completion.
            except RuntimeError:
                pass  # WHY: dialog already deleted; ignore safely.
        self._export_dialog = None  # WHY: release UI reference after safe close.
        if self._export_cancelled:  # WHY: skip success dialog when user cancelled.
            return
        if self._export_done_cb:  # WHY: allow per-export success handling.
            self._export_done_cb(result)  # WHY: call UI completion handler with payload.
            self._export_done_cb = None  # WHY: clear handler after use.

    def _on_export_error(self, msg):  # WHY: handle export errors uniformly.
        if self._export_dialog:  # WHY: close dialog on error.
            try:
                self._export_dialog.close()  # WHY: avoid stuck progress dialog on failure.
            except RuntimeError:
                pass  # WHY: dialog already deleted; ignore safely.
        self._export_dialog = None  # WHY: release UI reference after safe close.
        QMessageBox.critical(self, "Hata", f"DÄ±ÅŸa aktarma sÄ±rasÄ±nda hata: {msg}")  # WHY: show error without crashing UI.

    def _on_export_dialog_canceled(self):  # WHY: safe cancel handling for export progress dialog.
        self._export_cancelled = True  # WHY: mark cancel to skip success toast later.
        if self._export_worker:  # WHY: request cooperative stop for worker.
            self._export_worker.request_stop()  # WHY: avoid hard thread termination.
        if self._export_thread:  # WHY: set interruption flag for worker to observe.
            self._export_thread.requestInterruption()  # WHY: allow cooperative stop in loops.
        if self._export_dialog:  # WHY: update dialog text to show canceling.
            try:
                self._export_dialog.setLabelText("Ä°ptal ediliyor...")  # WHY: immediate feedback on cancel.
                self._export_dialog.setCancelButtonText("")  # WHY: prevent repeated cancel clicks.
            except RuntimeError:
                pass  # WHY: dialog already deleted; ignore safely.

    def _on_export_thread_finished(self):  # WHY: clean up thread refs safely after export.
        self._export_thread = None  # WHY: clear thread ref after it stops.
        self._export_worker = None  # WHY: clear worker ref after thread completion.

    def _pick_save_path(self, title, default_name, file_filter):
        cfg = load_config()
        last_dir = cfg.get("last_export_dir", "")
        default_path = os.path.join(last_dir, default_name) if last_dir else default_name
        path, _ = QFileDialog.getSaveFileName(self, title, default_path, file_filter)
        if path:
            try:
                cfg["last_export_dir"] = os.path.dirname(path)
                save_config(cfg)
            except Exception:
                pass
        return path

    def _next_available_path(self, path):  # WHY: avoid Windows file-lock overwrite errors by picking a free filename.
        if not path:
            return path
        if not os.path.exists(path):
            return path
        root, ext = os.path.splitext(path)
        idx = 1
        candidate = f"{root}_{idx}{ext}"
        while os.path.exists(candidate):
            idx += 1
            candidate = f"{root}_{idx}{ext}"
        return candidate

    def export_monthly_summary_pdf(self):
        try:
            from reportlab.lib.pagesizes import A4  # WHY: keep PDF generation unchanged, just threaded.
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer  # WHY: reuse existing PDF layout.
            from reportlab.lib.styles import getSampleStyleSheet  # WHY: preserve prior style handling.
            from reportlab.lib import colors  # WHY: keep table styling consistent with current output.
        except Exception:
            QMessageBox.warning(self, "Eksik Paket", "PDF oluşturmak için reportlab yüklü olmalı.")  # WHY: fail fast if dependency missing.
            return

        year = int(self.combo_year.currentText())  # WHY: keep current period selection intact.
        month = self.combo_month.currentIndex() + 1  # WHY: keep current period selection intact.
        tersane_id = self.tersane_id or 0  # WHY: normalize to global (0) if no tersane selected.
        tersane_label = self._get_active_tersane_label()  # WHY: show selected tersane in report title.

        path = self._pick_save_path("PDF Kaydet", f"Aylik_Ozet_{year}_{month:02d}.pdf", "PDF Files (*.pdf)")
        if not path:
            return

        def _task(worker):  # WHY: run PDF generation off the UI thread.
            if worker.should_stop():  # WHY: allow safe cancel before heavy work.
                return {"status": "cancelled"}  # WHY: inform UI about cancellation.
            db = Database()  # WHY: use a fresh DB handle inside worker thread.
            records = db.get_records_by_month(year, month, tersane_id)  # WHY: scope to active tersane for export.
            if not records:
                return {"status": "empty"}  # WHY: allow UI to show "no records" message.

            # Günlük yevmiye sayıları
            day_counts = {}  # WHY: compute daily headcount for summary table.
            total_normal = 0.0  # WHY: accumulate normal hours without changing formulas.
            total_mesai = 0.0  # WHY: accumulate overtime hours without changing formulas.
            person_set = set()  # WHY: compute unique person count.
            for r in records:
                if worker.should_stop():  # WHY: allow cooperative cancel mid-loop.
                    return {"status": "cancelled"}  # WHY: stop safely on cancel.
                tarih = r[1]  # WHY: keep existing record tuple positions.
                normal = float(r[6] or 0)  # WHY: preserve numeric casting logic.
                mesai = float(r[7] or 0)  # WHY: preserve numeric casting logic.
                if normal > 0:
                    day_counts[tarih] = day_counts.get(tarih, 0) + 1  # WHY: count daily yevmiye recipients.
                    person_set.add(r[2])  # WHY: track unique personnel names.
                total_normal += normal  # WHY: keep totals identical to previous logic.
                total_mesai += mesai  # WHY: keep totals identical to previous logic.

            data = [["Tarih", "Yevmiye Alan Kişi"]]  # WHY: preserve existing PDF header columns.
            for tarih in sorted(day_counts.keys()):
                data.append([tarih, str(day_counts[tarih])])  # WHY: keep same row structure for PDF table.

            styles = getSampleStyleSheet()  # WHY: reuse reportlab default styles.
            font_regular, font_bold = get_pdf_fonts()  # WHY: preserve Turkish-capable font selection.
            styles['Title'].fontName = font_bold  # WHY: keep bold title style.
            styles['Normal'].fontName = font_regular  # WHY: keep normal font for body text.
            doc = SimpleDocTemplate(path, pagesize=A4)  # WHY: keep existing page size/layout.
            story = []  # WHY: build PDF story elements as before.
            story.append(Paragraph(f"Aylık Özet — {year}/{month:02d} — {tersane_label}", styles['Title']))  # WHY: include tersane in title.
            story.append(Spacer(1, 12))  # WHY: keep spacing consistent with prior layout.
            story.append(Paragraph(f"Toplam Normal: {total_normal:.2f} | Toplam Mesai: {total_mesai:.2f}", styles['Normal']))  # WHY: keep totals visible.
            story.append(Paragraph(f"Yevmiye Alan Kişi Sayısı (benzersiz): {len(person_set)}", styles['Normal']))  # WHY: keep unique count line.
            story.append(Spacer(1, 12))  # WHY: preserve spacing before table.
            table = Table(data)  # WHY: keep same table structure.
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), font_bold),
                ('FONTNAME', (0, 1), (-1, -1), font_regular)
            ]))  # WHY: preserve table styling to match existing look.
            story.append(table)  # WHY: append summary table to PDF.
            doc.build(story)  # WHY: write the PDF file in worker thread.
            return {"status": "ok", "path": path}  # WHY: signal success to UI.

        def _done(result):  # WHY: handle export completion on UI thread.
            if not result:
                return  # WHY: no payload means nothing to show.
            if result.get("status") == "empty":
                QMessageBox.information(self, "Bilgi", "Seçilen ayda kayıt bulunamadı.")  # WHY: preserve prior empty-data message.
                return
            if result.get("status") == "cancelled":
                return  # WHY: skip success message on cancel.
            QMessageBox.information(self, "Başarılı", f"PDF oluşturuldu: {result.get('path', path)}")  # WHY: confirm successful export.

        self._start_export_worker(_task, done_cb=_done, label="PDF hazırlanıyor...")  # WHY: run export in background.

    def export_monthly_excel(self):
        """Seçili ayın filtreli görünümünü detay + özet Excel olarak dışa aktar."""
        rows = self._gather_visible_rows()
        if not rows:
            QMessageBox.information(self, "Bilgi", "Seçilen filtrelerde dışa aktarılacak kayıt yok.")  # WHY: keep prior empty-data behavior.
            return

        if openpyxl is None:
            QMessageBox.critical(self, "Hata", "openpyxl yüklü değil. Lütfen openpyxl yükleyin.")  # WHY: fail fast when Excel engine missing.
            return

        year = int(self.combo_year.currentText())
        month = self.combo_month.currentIndex() + 1
        month_name = self.combo_month.currentText()
        tersane_label = self._get_active_tersane_label()  # WHY: include active tersane in Excel titles.

        cleaned = []
        person_info = {}
        try:
            with self.db.get_connection() as conn:
                c = conn.cursor()
                c.execute("SELECT ad_soyad, maas, COALESCE(yevmiyeci_mi,0) FROM personel")
                for row in c.fetchall():
                    person_info[row[0]] = {
                        "maas": float(row[1] or 0),
                        "yevmiyeci_mi": int(row[2] or 0)
                    }
        except Exception:
            person_info = {}
        for r in rows:
            raw_date = r.get('Tarih') or ''
            parts = raw_date.split()
            day_str = parts[0] if parts else ''
            day_name = parts[-1] if len(parts) >= 3 else ''
            if day_str.isdigit():
                date_full = f"{year}-{month:02d}-{int(day_str):02d}"
            else:
                date_full = raw_date

            giris = r.get('Giriş') or "-"
            cikis = r.get('Çıkış') or "-"
            kayip = r.get('Kayıp') or "00:00:00"
            normal = float(r.get('Normal') or 0)
            mesai = float(r.get('Mesai') or 0)
            aciklama = r.get('Açıklama') or "-"

            if (giris == "-" or cikis == "-") and (normal > 0 or mesai > 0):
                durum = "Eksik"
            elif normal == 0 and mesai == 0 and ("Gelmedi" in aciklama or (giris == "-" and cikis == "-")):
                durum = "Gelmedi"
            else:
                durum = "Tam"

            person_name = r.get('Personel') or "-"
            birim = "Yevmiye" if person_info.get(person_name, {}).get("yevmiyeci_mi") else "Saat"
            cleaned.append({
                "Tarih": date_full,
                "Gün": day_name,
                "Personel": person_name,
                "Ekip": r.get('Ekip') or "-",
                "Giriş": giris,
                "Çıkış": cikis,
                "Kayıp": kayip,
                "Normal": normal,
                "Mesai": mesai,
                "Birim": birim,
                "Durum": durum,
                "Açıklama": aciklama
            })

        df_detail = pd.DataFrame(cleaned)

        df_summary = df_detail.groupby('Personel', as_index=False).agg(
            CalisilanGun=('Normal', lambda s: int(sum(1 for v in s if v > 0))),
            ToplamNormal=('Normal', 'sum'),
            ToplamMesai=('Mesai', 'sum'),
            EksikKayit=('Durum', lambda s: int(sum(1 for v in s if v == 'Eksik'))),
            Gelmedi=('Durum', lambda s: int(sum(1 for v in s if v == 'Gelmedi'))),
        )
        if not df_summary.empty:
            df_summary['Birim'] = df_summary['Personel'].apply(
                lambda p: "Yevmiye" if person_info.get(p, {}).get("yevmiyeci_mi") else "Saat"
            )
            df_summary = df_summary[[
                'Personel', 'Birim', 'CalisilanGun', 'ToplamNormal', 'ToplamMesai', 'EksikKayit', 'Gelmedi'
            ]]

        filters = [f"Tersane: {tersane_label}"]  # WHY: show active tersane in filter header.
        team = self.combo_team.currentText()
        if team and team != "Tüm Ekipler":
            filters.append(f"Ekip: {team}")
        name_f = self.search_name.text().strip()
        if name_f:
            filters.append(f"İsim: {name_f}")
        day_f = self.search_date.text().strip()
        if day_f:
            filters.append(f"Gün: {day_f}")
        if self.chk_only_empty.isChecked():
            filters.append("Sadece Boş")
        if self.chk_show_empty.isChecked():
            filters.append("Boş Kayıtlar Dahil")
        if self.chk_only_weekend.isChecked():
            filters.append("Sadece Haftasonu")
        if self.chk_only_special.isChecked():
            filters.append("Sadece Özel Durum")
        filter_text = " | ".join(filters) if filters else "Filtre: Yok"

        default_name = f"puantaj_{year}_{month:02d}.xlsx"
        path = self._pick_save_path("Aylik Excel Kaydet", default_name, "Excel (*.xlsx)")
        if not path:
            return
        safe_path = self._next_available_path(path)  # WHY: if selected file exists/open, export to a non-conflicting name.

        def _task(worker):  # WHY: move Excel creation off the UI thread.
            if worker.should_stop():
                return {"status": "cancelled"}  # WHY: allow user-initiated cancel.
            db = Database()  # WHY: use thread-local DB handle for safe background access.
            with pd.ExcelWriter(safe_path, engine='openpyxl') as writer:
                df_detail.to_excel(writer, index=False, sheet_name='Detay', startrow=5)
                df_summary.to_excel(writer, index=False, sheet_name='Özet', startrow=3)

                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                # Detay sheet
                ws = writer.sheets['Detay']
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
                ws.cell(row=1, column=1).value = f"Aylık Puantaj — {month_name} {year} — {tersane_label}"  # WHY: include tersane in title.
                ws.cell(row=1, column=1).font = Font(bold=True, size=14)
                ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ws.max_column)
                ws.cell(row=2, column=1).value = filter_text
                ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')

                header_row = 6
                header_fill = PatternFill(start_color='FF424242', end_color='FF424242', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFFFF')
                thin = Side(border_style="thin", color="FFAAAAAA")
                for col_cell in ws[header_row]:
                    col_cell.fill = header_fill
                    col_cell.font = header_font
                    col_cell.alignment = Alignment(horizontal='center', vertical='center')
                    col_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                dims = {}
                for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row):
                    for cell in row:
                        if cell.value is not None:
                            dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len(str(cell.value)))
                for col, value in dims.items():
                    ws.column_dimensions[col].width = min(max(value + 2, 10), 35)

                ws.freeze_panes = f"A{header_row+1}"

                fill_gray = PatternFill(start_color='FFF5F5F5', end_color='FFF5F5F5', fill_type='solid')
                for idx, row in enumerate(ws.iter_rows(min_row=header_row+1, max_row=ws.max_row), start=0):
                    if idx % 2 == 0:
                        for cell in row:
                            cell.fill = fill_gray

                ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

                # Detay sayfası sayı formatları
                detail_cols = {cell.value: cell.column for cell in ws[header_row]}  # WHY: map header names to columns.
                normal_col = detail_cols.get('Normal')  # WHY: format normal hours as numeric.
                mesai_col = detail_cols.get('Mesai')  # WHY: format overtime hours as numeric.
                for r_idx in range(header_row + 1, ws.max_row + 1):
                    if worker.should_stop():
                        return {"status": "cancelled"}  # WHY: allow safe cancel in long loops.
                    if normal_col:
                        ws.cell(row=r_idx, column=normal_col).number_format = '0.00'  # WHY: keep hours readable with decimals.
                    if mesai_col:
                        ws.cell(row=r_idx, column=mesai_col).number_format = '0.00'  # WHY: keep hours readable with decimals.

                # Özet sheet
                ws2 = writer.sheets['Özet']
                header_row2 = 4
                for col_cell in ws2[header_row2]:
                    col_cell.fill = header_fill
                    col_cell.font = header_font
                    col_cell.alignment = Alignment(horizontal='center', vertical='center')
                    col_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                dims2 = {}
                for row in ws2.iter_rows(min_row=header_row2, max_row=ws2.max_row):
                    for cell in row:
                        if cell.value is not None:
                            dims2[cell.column_letter] = max(dims2.get(cell.column_letter, 0), len(str(cell.value)))
                for col, value in dims2.items():
                    ws2.column_dimensions[col].width = min(max(value + 2, 10), 30)

                ws2.freeze_panes = f"A{header_row2+1}"

                for idx, row in enumerate(ws2.iter_rows(min_row=header_row2+1, max_row=ws2.max_row), start=0):
                    if idx % 2 == 0:
                        for cell in row:
                            cell.fill = fill_gray

                ws2.auto_filter.ref = f"A{header_row2}:{get_column_letter(ws2.max_column)}{ws2.max_row}"

                # Özet sayfası sayı formatları
                summary_cols = {cell.value: cell.column for cell in ws2[header_row2]}  # WHY: map summary headers.
                for r_idx in range(header_row2 + 1, ws2.max_row + 1):
                    if worker.should_stop():
                        return {"status": "cancelled"}  # WHY: allow safe cancel in long loops.
                    if summary_cols.get('ToplamNormal'):
                        ws2.cell(row=r_idx, column=summary_cols['ToplamNormal']).number_format = '0.00'  # WHY: numeric hours format.
                    if summary_cols.get('ToplamMesai'):
                        ws2.cell(row=r_idx, column=summary_cols['ToplamMesai']).number_format = '0.00'  # WHY: numeric hours format.
                    if summary_cols.get('CalisilanGun'):
                        ws2.cell(row=r_idx, column=summary_cols['CalisilanGun']).number_format = '0'  # WHY: integer day counts.
                    if summary_cols.get('EksikKayit'):
                        ws2.cell(row=r_idx, column=summary_cols['EksikKayit']).number_format = '0'  # WHY: integer counts.
                    if summary_cols.get('Gelmedi'):
                        ws2.cell(row=r_idx, column=summary_cols['Gelmedi']).number_format = '0'  # WHY: integer counts.

                # Çarşaf sayfa (aylık grid) - ekip bazlı tek sheet
                from calendar import monthrange
                turk_days = ["Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz"]
                ws3 = writer.book.create_sheet('Çarşaf')

                holiday_set = db.get_holidays()  # WHY: use worker-thread DB to avoid UI thread access.
                days_in_month = monthrange(year, month)[1]

                # Personel ücret/ekip/yevmiyeci bilgileri
                person_salary = {}
                person_yevmiyeci = {}
                person_team = {}
                leave_fill_map = {
                    "Hasta": PatternFill(start_color='FFFFF9C4', end_color='FFFFF9C4', fill_type='solid'),
                    "Raporlu": PatternFill(start_color='FFE1F5FE', end_color='FFE1F5FE', fill_type='solid'),
                    "\u00d6z\u00fcr": PatternFill(start_color='FFEDE7F6', end_color='FFEDE7F6', fill_type='solid'),
                    "Y\u0131ll\u0131k \u0130zin": PatternFill(start_color='FFFFFDE7', end_color='FFFFFDE7', fill_type='solid'),
                    "Do\u011fum \u0130zni": PatternFill(start_color='FFFFEBEE', end_color='FFFFEBEE', fill_type='solid'),
                    "\u0130dari \u0130zin": PatternFill(start_color='FFE0F2F1', end_color='FFE0F2F1', fill_type='solid'),
                    "Evlilik \u0130zni": PatternFill(start_color='FFD1C4E9', end_color='FFD1C4E9', fill_type='solid'),
                    "\u00c7ocuk \u0130zni": PatternFill(start_color='FFC8E6C9', end_color='FFC8E6C9', fill_type='solid'),
                    "\u0130\u015f Kazas\u0131 \u0130zni": PatternFill(start_color='FFFFCDD2', end_color='FFFFCDD2', fill_type='solid'),
                    "Di\u011fer": PatternFill(start_color='FFF0F4C3', end_color='FFF0F4C3', fill_type='solid'),
                }
                used_leave_types = set()
                with db.get_connection() as conn:
                    c = conn.cursor()
                    c.execute("SELECT ad_soyad, maas, COALESCE(yevmiyeci_mi,0), COALESCE(NULLIF(TRIM(ekip_adi),''),'Diğer') FROM personel")
                    for row in c.fetchall():
                        person_salary[row[0]] = float(row[1] or 0)
                        person_yevmiyeci[row[0]] = bool(row[2])
                        person_team[row[0]] = row[3] or "Diğer"

                # Map: person->day->(normal, mesai) ve ekip eşlemesi
                day_map = {}
                person_order = []
                seen_person = set()
                for r in rows:
                    raw_date = r.get('Tarih') or ''
                    parts = raw_date.split()
                    if not parts:
                        continue
                    if parts[0].isdigit():
                        day = int(parts[0])
                    else:
                        try:
                            day = int(raw_date.split("-")[-1])
                        except Exception:
                            continue
                    person = (r.get('Personel') or "-").strip() or "-"
                    ekip = (r.get('Ekip') or "").strip()
                    if ekip:
                        person_team[person] = ekip
                    elif person not in person_team:
                        person_team[person] = "Diğer"

                    if person not in seen_person:
                        person_order.append(person)
                        seen_person.add(person)

                    normal = float(r.get('Normal') or 0)
                    mesai = float(r.get('Mesai') or 0)
                    aciklama = str(
                        r.get('Açýklama')
                        or r.get('AÃ§Ä±klama')
                        or r.get('Açıklama')
                        or ""
                    ).strip()
                    izin_turu = db._canonicalize_izin_turu(aciklama) if aciklama else ""
                    day_map.setdefault(person, {}).setdefault(day, {"normal": 0.0, "mesai": 0.0, "izin_turu": ""})
                    day_map[person][day]["normal"] += normal
                    day_map[person][day]["mesai"] += mesai
                    if izin_turu in leave_fill_map:
                        used_leave_types.add(izin_turu)
                        if not day_map[person][day].get("izin_turu"):
                            day_map[person][day]["izin_turu"] = izin_turu

                # Ekip sırası: görünüm sırasını koru
                team_order = []
                for person in person_order:
                    team = (person_team.get(person) or "Diğer").strip() or "Diğer"
                    if team not in team_order:
                        team_order.append(team)

                team_groups = {}
                for team in team_order:
                    team_groups[team] = {"maasli": [], "yevmiyeci": []}
                for person in person_order:
                    team = (person_team.get(person) or "Diğer").strip() or "Diğer"
                    if team not in team_groups:
                        team_groups[team] = {"maasli": [], "yevmiyeci": []}
                        team_order.append(team)
                    if person_yevmiyeci.get(person, False):
                        team_groups[team]["yevmiyeci"].append(person)
                    else:
                        team_groups[team]["maasli"].append(person)

                total_col = 1 + days_in_month + 1
                last_col = total_col + 5

                # Başlık
                ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
                ws3.cell(row=1, column=1).value = f"Çarşaf Puantaj — {month_name} {year} — {tersane_label}"  # WHY: include tersane in sheet title.
                ws3.cell(row=1, column=1).font = Font(bold=True, size=14)
                ws3.cell(row=1, column=1).alignment = Alignment(horizontal='center')

                # Ortak stiller
                header_fill_main = PatternFill(start_color='FF1565C0', end_color='FF1565C0', fill_type='solid')
                header_fill_totals = PatternFill(start_color='FF283593', end_color='FF283593', fill_type='solid')
                weekend_fill = PatternFill(start_color='FFFFEBEE', end_color='FFFFEBEE', fill_type='solid')
                holiday_fill = PatternFill(start_color='FFFFF3E0', end_color='FFFFF3E0', fill_type='solid')
                total_fill = PatternFill(start_color='FFE8EAF6', end_color='FFE8EAF6', fill_type='solid')
                name_fill_normal = PatternFill(start_color='FFE3F2FD', end_color='FFE3F2FD', fill_type='solid')
                name_fill_mesai = PatternFill(start_color='FFF3E5F5', end_color='FFF3E5F5', fill_type='solid')
                data_fill = PatternFill(start_color='FFFAFAFA', end_color='FFFAFAFA', fill_type='solid')
                team_fill = PatternFill(start_color='FFCFD8DC', end_color='FFCFD8DC', fill_type='solid')
                section_fill = PatternFill(start_color='FFE1F5FE', end_color='FFE1F5FE', fill_type='solid')
                currency_format = '#,##0.00 ₺'
                person_sep = Side(border_style="medium", color="FF90A4AE")  # WHY: thick bottom border visually separates each employee block.

                def _day_fill(day_num):
                    dt = datetime(year, month, day_num)
                    key = dt.strftime("%m-%d")
                    if key in holiday_set:
                        return holiday_fill
                    if dt.weekday() >= 5:
                        return weekend_fill
                    return data_fill

                def _write_header(row_no):
                    ws3.cell(row=row_no, column=1, value="ADI SOYADI")
                    for d in range(1, days_in_month + 1):
                        day_label = f"{d:02d} {turk_days[datetime(year, month, d).weekday()]}"
                        ws3.cell(row=row_no, column=1 + d, value=day_label)
                    ws3.cell(row=row_no, column=total_col, value="TOPLAM GÜN")
                    ws3.cell(row=row_no, column=total_col + 1, value="TOPLAM NORMAL")
                    ws3.cell(row=row_no, column=total_col + 2, value="BİRİM ÜC.")
                    ws3.cell(row=row_no, column=total_col + 3, value="HAKEDİŞ")
                    ws3.cell(row=row_no, column=total_col + 4, value="MAAŞ")
                    ws3.cell(row=row_no, column=total_col + 5, value="TOPLAM HAKEDİŞ")

                    for col_idx in range(1, 1 + days_in_month + 1):
                        cell = ws3.cell(row=row_no, column=col_idx)
                        cell.fill = header_fill_main
                        cell.font = Font(bold=True, color='FFFFFFFF', size=9)
                        cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
                        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    for col_idx in range(total_col, total_col + 6):
                        cell = ws3.cell(row=row_no, column=col_idx)
                        cell.fill = header_fill_totals
                        cell.font = Font(bold=True, color='FFFFFFFF', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90, wrap_text=True)
                        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    ws3.row_dimensions[row_no].height = 60

                legend_items = [("Hafta Sonu", weekend_fill), ("Resmi Tatil", holiday_fill)]
                for leave_type in leave_fill_map.keys():
                    if leave_type in used_leave_types:
                        legend_items.append((leave_type, leave_fill_map[leave_type]))
                ws3.cell(row=2, column=1, value="Lejant:")
                ws3.cell(row=2, column=1).font = Font(bold=True, size=10)
                ws3.cell(row=2, column=1).alignment = Alignment(horizontal='left', vertical='center')
                legend_col = 2
                for label, fill in legend_items:
                    lcell = ws3.cell(row=2, column=legend_col, value=label)
                    lcell.fill = fill
                    lcell.font = Font(bold=True, size=9, color='FF263238')
                    lcell.alignment = Alignment(horizontal='center', vertical='center')
                    lcell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    legend_col += 1
                ws3.row_dimensions[2].height = 20

                current_row = 4
                first_data_row = None

                for team in team_order:
                    sections = [
                        ("Maaşlılar", sorted(team_groups.get(team, {}).get("maasli", []))),
                        ("Yevmiyeciler", sorted(team_groups.get(team, {}).get("yevmiyeci", []))),
                    ]
                    if not any(people for _, people in sections):
                        continue

                    ws3.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=last_col)
                    tcell = ws3.cell(row=current_row, column=1)
                    tcell.value = f"EKİP: {team}"
                    tcell.font = Font(bold=True, size=12, color='FF263238')
                    tcell.alignment = Alignment(horizontal='left', vertical='center')
                    tcell.fill = team_fill
                    for col_idx in range(1, last_col + 1):
                        ws3.cell(row=current_row, column=col_idx).border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    current_row += 1

                    for section_name, people in sections:
                        if not people:
                            continue

                        ws3.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=last_col)
                        scell = ws3.cell(row=current_row, column=1)
                        scell.value = section_name
                        scell.font = Font(bold=True, size=11, color='FF0D47A1')
                        scell.alignment = Alignment(horizontal='left', vertical='center')
                        scell.fill = section_fill
                        for col_idx in range(1, last_col + 1):
                            ws3.cell(row=current_row, column=col_idx).border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        current_row += 1

                        header_row3 = current_row
                        _write_header(header_row3)
                        current_row += 1
                        if first_data_row is None:
                            first_data_row = current_row

                        for person in people:
                            row_normal = current_row
                            row_mesai = row_normal + 1

                            salary = person_salary.get(person, 0)
                            is_yevmiyeci = person_yevmiyeci.get(person, False)
                            if is_yevmiyeci:
                                birim_ucret = salary
                            else:
                                yevmiye = salary / 30.0 if salary > 0 else 0
                                birim_ucret = round(yevmiye / 7.5, 2) if yevmiye > 0 else 0

                            ws3.cell(row=row_normal, column=1, value=person)
                            ws3.cell(row=row_mesai, column=1, value=f"{person} (Mesai)")
                            ws3.cell(row=row_normal, column=1).fill = name_fill_normal
                            ws3.cell(row=row_mesai, column=1).fill = name_fill_mesai
                            ws3.cell(row=row_normal, column=1).font = Font(bold=True, size=10)
                            ws3.cell(row=row_mesai, column=1).font = Font(italic=True, size=9)
                            ws3.cell(row=row_normal, column=1).alignment = Alignment(horizontal='left', vertical='center')
                            ws3.cell(row=row_mesai, column=1).alignment = Alignment(horizontal='left', vertical='center')
                            ws3.cell(row=row_normal, column=1).border = Border(left=thin, right=thin, top=thin, bottom=thin)
                            ws3.cell(row=row_mesai, column=1).border = Border(left=thin, right=thin, top=thin, bottom=person_sep)

                            toplam_gun = 0
                            toplam_normal = 0.0
                            toplam_mesai = 0.0
                            day_start_col_letter = get_column_letter(2)
                            day_end_col_letter = get_column_letter(1 + days_in_month)
                            toplam_gun_col_letter = get_column_letter(total_col)
                            toplam_normal_col_letter = get_column_letter(total_col + 1)
                            birim_col_letter = get_column_letter(total_col + 2)
                            hakedis_col_letter = get_column_letter(total_col + 3)
                            maas_col_letter = get_column_letter(total_col + 4)

                            for d in range(1, days_in_month + 1):
                                entry = day_map.get(person, {}).get(d)
                                cell_val_normal = ""
                                cell_val_mesai = ""
                                if entry:
                                    n = entry.get("normal", 0.0)
                                    m = entry.get("mesai", 0.0)
                                    if n > 0:
                                        toplam_gun += 1
                                        cell_val_normal = n
                                    toplam_normal += n
                                    toplam_mesai += m
                                    if m > 0:
                                        cell_val_mesai = m

                                c_norm = ws3.cell(row=row_normal, column=1 + d, value=cell_val_normal)
                                c_mes = ws3.cell(row=row_mesai, column=1 + d, value=cell_val_mesai)
                                fill = _day_fill(d)
                                if entry and entry.get("izin_turu") in leave_fill_map:
                                    fill = leave_fill_map[entry.get("izin_turu")]
                                for cc in (c_norm, c_mes):
                                    cc.fill = fill
                                    bot = person_sep if cc is c_mes else thin
                                    cc.border = Border(left=thin, right=thin, top=thin, bottom=bot)
                                    cc.alignment = Alignment(horizontal='center', vertical='center')
                                    if isinstance(cc.value, (int, float)):
                                        cc.number_format = '0.00'

                            ws3.cell(
                                row=row_normal,
                                column=total_col,
                                value=f'=COUNTIF({day_start_col_letter}{row_normal}:{day_end_col_letter}{row_normal},">0")'
                            )
                            ws3.cell(
                                row=row_normal,
                                column=total_col + 1,
                                value=f"=SUM({day_start_col_letter}{row_normal}:{day_end_col_letter}{row_normal})"
                            )
                            ws3.cell(row=row_normal, column=total_col + 2, value=birim_ucret)
                            ws3.cell(row=row_normal, column=total_col + 4, value=salary)
                            ws3.cell(row=row_normal, column=total_col + 5, value="")
                            if is_yevmiyeci:
                                ws3.cell(
                                    row=row_normal,
                                    column=total_col + 3,
                                    value=f"=ROUND({toplam_normal_col_letter}{row_normal}*{birim_col_letter}{row_normal},2)"
                                )
                            else:
                                ws3.cell(
                                    row=row_normal,
                                    column=total_col + 3,
                                    value=(
                                        f"=ROUND((MIN(30,MAX(0,30-({days_in_month}-"
                                        f"({toplam_normal_col_letter}{row_normal}/7.5))))/30)*"
                                        f"{maas_col_letter}{row_normal},2)"
                                    )
                                )

                            ws3.cell(row=row_mesai, column=total_col, value="")
                            ws3.cell(
                                row=row_mesai,
                                column=total_col + 1,
                                value=f"=SUM({day_start_col_letter}{row_mesai}:{day_end_col_letter}{row_mesai})"
                            )
                            ws3.cell(row=row_mesai, column=total_col + 2, value=f"={birim_col_letter}{row_normal}")
                            ws3.cell(
                                row=row_mesai,
                                column=total_col + 3,
                                value=f"=ROUND({toplam_normal_col_letter}{row_mesai}*{birim_col_letter}{row_mesai},2)"
                            )
                            ws3.cell(row=row_mesai, column=total_col + 4, value="")
                            ws3.cell(
                                row=row_mesai,
                                column=total_col + 5,
                                value=f"={hakedis_col_letter}{row_normal}+{hakedis_col_letter}{row_mesai}"
                            )

                            for r_idx in (row_normal, row_mesai):
                                for c_idx in range(total_col, total_col + 6):
                                    cell = ws3.cell(row=r_idx, column=c_idx)
                                    cell.fill = total_fill
                                    bot = person_sep if r_idx == row_mesai else thin
                                    cell.border = Border(left=thin, right=thin, top=thin, bottom=bot)
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                                    if c_idx in (total_col + 2, total_col + 3, total_col + 4, total_col + 5):
                                        cell.number_format = currency_format
                                    elif c_idx == total_col + 1:
                                        cell.number_format = '0.00'
                                    elif c_idx == total_col:
                                        cell.number_format = '0'
                            ws3.cell(row=row_mesai, column=total_col + 5).font = Font(bold=True, size=11, color='FF1565C0')
                            ws3.cell(row=row_mesai, column=total_col + 5).fill = PatternFill(start_color='FFBBDEFB', end_color='FFBBDEFB', fill_type='solid')

                            ws3.row_dimensions[row_normal].height = 20
                            ws3.row_dimensions[row_mesai].height = 18
                            current_row += 2

                        current_row += 1  # blank line between sections

                    current_row += 1  # blank line between teams

                # Column widths — autofit (formüller ve gün sütunları hariç)
                day_col_indices = set(range(2, 2 + days_in_month))
                col_widths: dict[int, float] = {}
                for _row in ws3.iter_rows():
                    for _cell in _row:
                        ci = _cell.column
                        if ci in day_col_indices:
                            col_widths[ci] = 4.5  # WHY: rotated header — narrow is correct.
                            continue
                        if _cell.value is None:
                            continue
                        val_str = str(_cell.value)
                        if val_str.startswith('='):
                            continue  # WHY: formula strings are arbitrarily long; skip them.
                        col_widths[ci] = max(col_widths.get(ci, 8), len(val_str) + 2)
                for ci, w in col_widths.items():
                    ws3.column_dimensions[get_column_letter(ci)].width = min(w, 40)
                ws3.column_dimensions['A'].width = max(col_widths.get(1, 26), 26)  # WHY: name column always at least 26.

                ws3.freeze_panes = f"B{first_data_row or 4}"

            return {"status": "ok", "path": safe_path}  # WHY: report real output path (may include suffix on conflicts).

        def _done(result):  # WHY: show completion message on UI thread.
            if not result:
                return  # WHY: no payload, nothing to show.
            if result.get("status") == "cancelled":
                return  # WHY: skip success message on cancel.
            QMessageBox.information(self, "Başarılı", f"Aylık Excel oluşturuldu: {result.get('path', path)}")  # WHY: confirm export completion.

        self._start_export_worker(_task, done_cb=_done, label="Aylık Excel hazırlanıyor...")  # WHY: run export in background.

    def on_cell_changed(self, row, col):
        # --- Hücre düzenleme güvenliği ---
        rec_id = self.table.item(row, 0).text()
        val = self.table.item(row, col).text()
        col_map = {6: 'kayip_sure_saat', 7: 'hesaplanan_normal', 8: 'hesaplanan_mesai', 9: 'aciklama'}
        try:
            # Tarih hücresi ise: sadece YYYY-MM-DD formatı kabul
            if col == 1:
                from PySide6.QtWidgets import QMessageBox
                import re
                iso_date = self.table.item(row, col).data(Qt.UserRole)
                if not iso_date or not re.match(r"^\d{4}-\d{2}-\d{2}$", str(iso_date)):
                    QMessageBox.warning(self, "Hatalı Tarih", "Tarih formatı geçersiz. Lütfen YYYY-MM-DD formatında bir tarih seçin.")
                    self.load_data()
                    return
            # Saat hücreleri: sadece HH:MM veya HH:MM:SS kabul
            if col in [4, 5]:
                from PySide6.QtWidgets import QMessageBox
                import re
                if val and not re.match(r"^([01]?\d|2[0-3]):[0-5]\d(:[0-5]\d)?$", val):
                    QMessageBox.warning(self, "Hatalı Saat", "Saat formatı geçersiz. Lütfen HH:MM veya HH:MM:SS formatında girin.")
                    self.load_data()
                    return
            # Kayıp süre: sadece HH:MM veya HH:MM:SS kabul
            if col == 6:
                from PySide6.QtWidgets import QMessageBox
                import re
                if val and not re.match(r"^([01]?\d|2[0-3]):[0-5]\d(:[0-5]\d)?$", val):
                    QMessageBox.warning(self, "Hatalı Süre", "Kayıp süre formatı geçersiz. Lütfen HH:MM veya HH:MM:SS formatında girin.")
                    self.load_data()
                    return
            # Sadece izin verilen kolonlar DB'ye yazılır
            if col in col_map:
                self.db.update_single_record(rec_id, col_map[col], val)
                self.signal_manager.data_updated.emit()
        except Exception as e:
            from PySide6.QtWidgets import QMessageBox
            import traceback
            QMessageBox.critical(self, "Hata", f"Hücre güncelleme hatası (satır={row}, sütun={col}):\n{e}\n{traceback.format_exc()}")

    def _gather_visible_rows(self):
        """Return a list of dicts for visible rows in the table."""
        rows = []
        for i in range(self.table.rowCount()):
            if self.table.isRowHidden(i):
                continue
            try:
                rec = {
                    'Tarih': self.table.item(i, 1).text(),
                    'Personel': self.table.item(i, 2).text(),
                    'Ekip': self.table.item(i, 3).text(),
                    'Giriş': self.table.item(i, 4).text(),
                    'Çıkış': self.table.item(i, 5).text(),
                    'Kayıp': self.table.item(i, 6).text(),
                    'Normal': float(self.table.item(i, 7).text() or 0),
                    'Mesai': float(self.table.item(i, 8).text() or 0),
                    'Açıklama': self.table.item(i, 9).text()
                }
                rows.append(rec)
            except Exception:
                continue
        return rows

    def _export_to_excel_legacy(self):  # WHY: keep original sync export as reference; replaced by threaded version below.
        """Export currently visible rows to a styled Excel (Çarşaf puantajı)."""
        # Yeni: Dialog üzerinden tarih aralığı / ekip / personel seçimi ile dışa aktarma
        dlg = ExportDialog(self.db, self)
        if dlg.exec() != QDialog.Accepted:
            return
        vals = dlg.get_values()

        # Prefer a typed name filter (search box) if user typed a name but didn't select in dialog
        person_in_dialog = vals.get('person')
        if not person_in_dialog:
            typed_name = self.search_name.text().strip()
            if typed_name:
                rows_raw = self.db.get_records_between_like(vals['date_from'], vals['date_to'], team=vals['team'], person_like=typed_name)
            else:
                rows_raw = self.db.get_records_between(vals['date_from'], vals['date_to'], team=vals['team'], person=person_in_dialog)
        else:
            rows_raw = self.db.get_records_between(vals['date_from'], vals['date_to'], team=vals['team'], person=person_in_dialog)
        if not rows_raw:
            QMessageBox.information(self, "Bilgi", "Seçilen kriterlerde kayıt bulunamadı.")
            return

        rows = []
        for r in rows_raw:
            rows.append({
                'Tarih': r[1],
                'Personel': r[2],
                'Ekip': r[9] if r[9] else '',
                'Giriş': r[3] or '',
                'Çıkış': r[4] or '',
                'Kayıp': r[5] or '',
                'Normal': float(r[6] or 0),
                'Mesai': float(r[7] or 0),
                'Açıklama': r[8] or ''
            })
        df = pd.DataFrame(rows)

        # Add summary row
        total_normal = df['Normal'].sum()
        total_mesai = df['Mesai'].sum()
        summary = {k: '' for k in df.columns}
        summary['Tarih'] = 'Toplam'
        summary['Normal'] = total_normal
        summary['Mesai'] = total_mesai
        df = pd.concat([df, pd.DataFrame([summary])], ignore_index=True)

        # Ask save path
        path = self._pick_save_path("Excel Kaydet", "puantaj_export.xlsx", "Excel (*.xlsx)")
        if not path:
            return

        try:
            if openpyxl is None:
                raise ImportError("openpyxl yüklü değil. Lütfen 'pip install openpyxl' komutunu çalıştırın.")

            # Write dataframe starting at row 4 (leave space for title and header styling)
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Puantaj', startrow=3)
                wb = writer.book
                ws = writer.sheets['Puantaj']

                # Title row (merged)
                title = f"Puantaj {vals['date_from']} → {vals['date_to']}"
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
                tcell = ws.cell(row=1, column=1)
                tcell.value = title
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                tcell.font = Font(bold=True, size=14)
                tcell.alignment = Alignment(horizontal='center')

                # Insert logo if requested
                if vals.get('logo'):
                    try:
                        from openpyxl.drawing.image import Image as XLImage
                        img = XLImage(vals.get('logo'))
                        img.width = 120
                        img.height = 40
                        ws.add_image(img, 'A1')
                    except Exception:
                        pass

                # Header format (header is at row 4 because of title + blank + header)
                header_row = 4
                header_fill = PatternFill(start_color='FF424242', end_color='FF424242', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFFFF')
                thin = Side(border_style="thin", color="FFAAAAAA")
                for col_cell in ws[header_row]:
                    col_cell.fill = header_fill
                    col_cell.font = header_font
                    col_cell.alignment = Alignment(horizontal='center')
                    col_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                # Column widths
                dims = {}
                for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row):
                    for cell in row:
                        if cell.value is not None:
                            dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len(str(cell.value)))
                for col, value in dims.items():
                    ws.column_dimensions[col].width = min(max(value + 2, 10), 40)

                # Freeze header
                ws.freeze_panes = f'A{header_row+1}'

                # Alternating row colors
                fill_gray = PatternFill(start_color='FFF5F5F5', end_color='FFF5F5F5', fill_type='solid')
                for idx, row in enumerate(ws.iter_rows(min_row=header_row+1, max_row=ws.max_row-1), start=0):
                    if idx % 2 == 0:
                        for cell in row:
                            cell.fill = fill_gray

                # Summary row: last row
                last_row = ws.max_row
                cols = {cell.value: cell.column for cell in ws[header_row]}
                from openpyxl.utils import get_column_letter
                normal_col = get_column_letter(cols.get('Normal')) if cols.get('Normal') else None
                mesai_col = get_column_letter(cols.get('Mesai')) if cols.get('Mesai') else None
                data_start = header_row + 1
                data_end = last_row - 1
                if normal_col:
                    if vals.get('formulas'):
                        ws[f"{normal_col}{last_row}"] = f"=SUM({normal_col}{data_start}:{normal_col}{data_end})"
                    else:
                        total_normal = sum([float(ws[f"{normal_col}{r}"].value or 0) for r in range(data_start, data_end+1)])
                        ws[f"{normal_col}{last_row}"] = total_normal
                if mesai_col:
                    if vals.get('formulas'):
                        ws[f"{mesai_col}{last_row}"] = f"=SUM({mesai_col}{data_start}:{mesai_col}{data_end})"
                    else:
                        total_mesai = sum([float(ws[f"{mesai_col}{r}"].value or 0) for r in range(data_start, data_end+1)])
                        ws[f"{mesai_col}{last_row}"] = total_mesai

                # Apply border to summary row
                for cell in ws[last_row]:
                    cell.font = Font(bold=True)
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Export sonrası ek butonlar ve gelişmiş bilgi
            if os.path.exists(path):
                msg = QMessageBox(self)
                msg.setWindowTitle("Başarılı")
                msg.setText(f"Excel başarıyla kaydedildi:\n{path}")
                msg.setIcon(QMessageBox.Information)
                btn_open = msg.addButton("Dosyayı Aç", QMessageBox.AcceptRole)
                btn_folder = msg.addButton("Klasörü Göster", QMessageBox.ActionRole)
                btn_copy = msg.addButton("Yolu Kopyala", QMessageBox.ActionRole)
                btn_mail = msg.addButton("E-posta ile Gönder", QMessageBox.ActionRole)
                msg.addButton("Kapat", QMessageBox.RejectRole)
                msg.exec()
                clicked = msg.clickedButton()
                if clicked == btn_open:
                    try:
                        os.startfile(os.path.abspath(path))
                    except Exception as e:
                        QMessageBox.warning(self, "Hata", f"Dosya açılamadı: {e}")
                elif clicked == btn_folder:
                    try:
                        os.startfile(os.path.dirname(os.path.abspath(path)))
                    except Exception as e:
                        QMessageBox.warning(self, "Hata", f"Klasör açılamadı: {e}")
                elif clicked == btn_copy:
                    QApplication.clipboard().setText(path)
                    QMessageBox.information(self, "Kopyalandı", "Dosya yolu panoya kopyalandı.")
                elif clicked == btn_mail:
                    import webbrowser
                    import urllib.parse
                    subject = urllib.parse.quote("Puantaj Raporu")
                    body = urllib.parse.quote(f"Rapor dosyası: {path}")
                    webbrowser.open(f"mailto:?subject={subject}&body={body}")
            else:
                QMessageBox.information(self, "Başarılı", f"Excel başarıyla kaydedildi: {path}")
        except Exception as e:
            import traceback
            from core.app_logger import log_error
            tb = traceback.format_exc()
            log_error(f"Export Hatası: {e}\n{tb}")
            QMessageBox.critical(self, "Hata", f"Dışa aktarırken hata: {e}\n\n{tb}")

    def export_to_excel(self):  # WHY: threaded export to keep UI responsive while preserving logic.
        """Export currently visible rows to a styled Excel (Çarşaf puantajı)."""  # WHY: keep original export purpose.
        # Yeni: Dialog üzerinden tarih aralığı / ekip / personel seçimi ile dışa aktarma
        dlg = ExportDialog(self.db, self)  # WHY: reuse existing export dialog for filters.
        if dlg.exec() != QDialog.Accepted:
            return
        vals = dlg.get_values()

        if openpyxl is None:
            QMessageBox.critical(self, "Hata", "openpyxl yüklü değil. Lütfen openpyxl yükleyin.")  # WHY: avoid running export without engine.
            return

        tersane_id = self.tersane_id or 0  # WHY: normalize to global (0) if no tersane selected.
        tersane_label = self._get_active_tersane_label()  # WHY: include tersane in export title.

        # Prefer a typed name filter (search box) if user typed a name but didn't select in dialog
        person_in_dialog = vals.get('person')
        typed_name = self.search_name.text().strip()  # WHY: preserve quick name filter behavior.

        # Ask save path
        path = self._pick_save_path("Excel Kaydet", "puantaj_export.xlsx", "Excel (*.xlsx)")
        if not path:
            return

        def _task(worker):  # WHY: run export off the UI thread.
            if worker.should_stop():
                return {"status": "cancelled"}  # WHY: allow user-initiated cancel.
            db = Database()  # WHY: use thread-local DB handle for safe background access.

            if not person_in_dialog:
                if typed_name:
                    rows_raw = db.get_records_between_like(vals['date_from'], vals['date_to'], team=vals['team'], person_like=typed_name, tersane_id=tersane_id)  # WHY: include tersane filter.
                else:
                    rows_raw = db.get_records_between(vals['date_from'], vals['date_to'], team=vals['team'], person=person_in_dialog, tersane_id=tersane_id)  # WHY: include tersane filter.
            else:
                rows_raw = db.get_records_between(vals['date_from'], vals['date_to'], team=vals['team'], person=person_in_dialog, tersane_id=tersane_id)  # WHY: include tersane filter.
            if not rows_raw:
                return {"status": "empty"}  # WHY: report empty data back to UI.

            rows = []
            for r in rows_raw:
                if worker.should_stop():
                    return {"status": "cancelled"}  # WHY: allow safe cancel during loops.
                rows.append({  # WHY: map DB rows to export columns without changing content.
                    'Tarih': r[1],
                    'Personel': r[2],
                    'Ekip': r[9] if r[9] else '',
                    'Giriş': r[3] or '',
                    'Çıkış': r[4] or '',
                    'Kayıp': r[5] or '',
                    'Normal': float(r[6] or 0),
                    'Mesai': float(r[7] or 0),
                    'Açıklama': r[8] or ''
                })
            df = pd.DataFrame(rows)

            # Add summary row (unchanged logic)
            total_normal = df['Normal'].sum()
            total_mesai = df['Mesai'].sum()
            summary = {k: '' for k in df.columns}
            summary['Tarih'] = 'Toplam'
            summary['Normal'] = total_normal
            summary['Mesai'] = total_mesai
            df = pd.concat([df, pd.DataFrame([summary])], ignore_index=True)

            # Write dataframe starting at row 4 (leave space for title and header styling)
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Puantaj', startrow=3)
                ws = writer.sheets['Puantaj']

                # Title row (merged)
                title = f"Puantaj {vals['date_from']} → {vals['date_to']} — {tersane_label}"  # WHY: include tersane name in export title.
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
                tcell = ws.cell(row=1, column=1)
                tcell.value = title
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                tcell.font = Font(bold=True, size=14)
                tcell.alignment = Alignment(horizontal='center')

                # Insert logo if requested
                if vals.get('logo'):
                    try:
                        from openpyxl.drawing.image import Image as XLImage
                        img = XLImage(vals.get('logo'))
                        img.width = 120
                        img.height = 40
                        ws.add_image(img, 'A1')
                    except Exception:
                        pass

                # Header format (header is at row 4 because of title + blank + header)
                header_row = 4
                header_fill = PatternFill(start_color='FF424242', end_color='FF424242', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFFFF')
                thin = Side(border_style="thin", color="FFAAAAAA")
                for col_cell in ws[header_row]:
                    col_cell.fill = header_fill
                    col_cell.font = header_font
                    col_cell.alignment = Alignment(horizontal='center')
                    col_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                # Column widths
                dims = {}
                for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row):
                    for cell in row:
                        if cell.value is not None:
                            dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len(str(cell.value)))
                for col, value in dims.items():
                    ws.column_dimensions[col].width = min(max(value + 2, 10), 40)

                # Freeze header
                ws.freeze_panes = f'A{header_row+1}'

                # Alternating row colors
                fill_gray = PatternFill(start_color='FFF5F5F5', end_color='FFF5F5F5', fill_type='solid')
                for idx, row in enumerate(ws.iter_rows(min_row=header_row+1, max_row=ws.max_row-1), start=0):
                    if idx % 2 == 0:
                        for cell in row:
                            cell.fill = fill_gray

                # Summary row: last row
                last_row = ws.max_row
                cols = {cell.value: cell.column for cell in ws[header_row]}
                from openpyxl.utils import get_column_letter
                normal_col = get_column_letter(cols.get('Normal')) if cols.get('Normal') else None
                mesai_col = get_column_letter(cols.get('Mesai')) if cols.get('Mesai') else None
                data_start = header_row + 1
                data_end = last_row - 1
                if normal_col:
                    if vals.get('formulas'):
                        ws[f"{normal_col}{last_row}"] = f"=SUM({normal_col}{data_start}:{normal_col}{data_end})"
                    else:
                        total_normal = sum([float(ws[f"{normal_col}{r}"].value or 0) for r in range(data_start, data_end+1)])
                        ws[f"{normal_col}{last_row}"] = total_normal
                if mesai_col:
                    if vals.get('formulas'):
                        ws[f"{mesai_col}{last_row}"] = f"=SUM({mesai_col}{data_start}:{mesai_col}{data_end})"
                    else:
                        total_mesai = sum([float(ws[f"{mesai_col}{r}"].value or 0) for r in range(data_start, data_end+1)])
                        ws[f"{mesai_col}{last_row}"] = total_mesai

                # Apply border to summary row
                for cell in ws[last_row]:
                    cell.font = Font(bold=True)
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            return {"status": "ok", "path": path}  # WHY: signal success with output path.

        def _done(result):  # WHY: handle export completion on UI thread.
            if not result:
                return  # WHY: no payload, nothing to show.
            if result.get("status") == "empty":
                QMessageBox.information(self, "Bilgi", "Seçilen kriterlerde kayıt bulunamadı.")  # WHY: preserve prior empty-data message.
                return
            if result.get("status") == "cancelled":
                return  # WHY: skip success dialog on cancel.

            # Export sonrası ek butonlar ve gelişmiş bilgi
            if os.path.exists(path):
                msg = QMessageBox(self)
                msg.setWindowTitle("Başarılı")
                msg.setText(f"Excel başarıyla kaydedildi:\n{path}")
                msg.setIcon(QMessageBox.Information)
                btn_open = msg.addButton("Dosyayı Aç", QMessageBox.AcceptRole)
                btn_folder = msg.addButton("Klasörü Göster", QMessageBox.ActionRole)
                btn_copy = msg.addButton("Yolu Kopyala", QMessageBox.ActionRole)
                btn_mail = msg.addButton("E-posta ile Gönder", QMessageBox.ActionRole)
                msg.addButton("Kapat", QMessageBox.RejectRole)
                msg.exec()
                clicked = msg.clickedButton()
                if clicked == btn_open:
                    try:
                        os.startfile(os.path.abspath(path))
                    except Exception as e:
                        QMessageBox.warning(self, "Hata", f"Dosya açılamadı: {e}")
                elif clicked == btn_folder:
                    try:
                        os.startfile(os.path.dirname(os.path.abspath(path)))
                    except Exception as e:
                        QMessageBox.warning(self, "Hata", f"Klasör açılamadı: {e}")
                elif clicked == btn_copy:
                    QApplication.clipboard().setText(path)
                    QMessageBox.information(self, "Kopyalandı", "Dosya yolu panoya kopyalandı.")
                elif clicked == btn_mail:
                    import webbrowser
                    import urllib.parse
                    subject = urllib.parse.quote("Puantaj Raporu")
                    body = urllib.parse.quote(f"Rapor dosyası: {path}")
                    webbrowser.open(f"mailto:?subject={subject}&body={body}")
            else:
                QMessageBox.information(self, "Başarılı", f"Excel başarıyla kaydedildi: {path}")

        self._start_export_worker(_task, done_cb=_done, label="Excel hazırlanıyor...")  # WHY: run export in background.

    def export_df_to_file(self, df, path, title=None, logo_path=None, use_formulas=False):
        """Programmatic export utility (callable from scripts/tests)."""
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Puantaj', startrow=3)
            wb = writer.book
            ws = writer.sheets['Puantaj']

            # Title row
            if title:
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
                tcell = ws.cell(row=1, column=1)
                tcell.value = title
                tcell.font = Font(bold=True, size=14)
                tcell.alignment = Alignment(horizontal='center')

            # Insert logo if provided
            if logo_path:
                try:
                    img = XLImage(logo_path)
                    img.width = 120
                    img.height = 40
                    ws.add_image(img, 'A1')
                except Exception:
                    pass

            # Header format
            header_row = 4
            header_fill = PatternFill(start_color='FF424242', end_color='FF424242', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFFFF')
            thin = Side(border_style="thin", color="FFAAAAAA")
            for col_cell in ws[header_row]:
                col_cell.fill = header_fill
                col_cell.font = header_font
                col_cell.alignment = Alignment(horizontal='center')
                col_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Column widths
            dims = {}
            for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row):
                for cell in row:
                    if cell.value is not None:
                        dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len(str(cell.value)))
            for col, value in dims.items():
                ws.column_dimensions[col].width = min(max(value + 2, 10), 40)

            # Freeze header
            ws.freeze_panes = f'A{header_row+1}'

            # Alternating row colors
            fill_gray = PatternFill(start_color='FFF5F5F5', end_color='FFF5F5F5', fill_type='solid')
            for idx, row in enumerate(ws.iter_rows(min_row=header_row+1, max_row=ws.max_row-1), start=0):
                if idx % 2 == 0:
                    for cell in row:
                        cell.fill = fill_gray

            # Summary row
            last_row = ws.max_row
            cols = {cell.value: cell.column for cell in ws[header_row]}
            normal_col = get_column_letter(cols.get('Normal')) if cols.get('Normal') else None
            mesai_col = get_column_letter(cols.get('Mesai')) if cols.get('Mesai') else None
            data_start = header_row + 1
            data_end = last_row - 1
            if normal_col:
                if use_formulas:
                    ws[f"{normal_col}{last_row}"] = f"=SUM({normal_col}{data_start}:{normal_col}{data_end})"
                else:
                    total_normal = sum([float(ws[f"{normal_col}{r}"].value or 0) for r in range(data_start, data_end+1)])
                    ws[f"{normal_col}{last_row}"] = total_normal
            if mesai_col:
                if use_formulas:
                    ws[f"{mesai_col}{last_row}"] = f"=SUM({mesai_col}{data_start}:{mesai_col}{data_end})"
                else:
                    total_mesai = sum([float(ws[f"{mesai_col}{r}"].value or 0) for r in range(data_start, data_end+1)])
                    ws[f"{mesai_col}{last_row}"] = total_mesai

            for cell in ws[last_row]:
                cell.font = Font(bold=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
