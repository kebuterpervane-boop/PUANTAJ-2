from datetime import datetime
from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget,
                              QTableWidgetItem, QHeaderView, QPushButton,
                              QLabel, QComboBox, QMessageBox, QSpinBox,
                              QDialog, QDialogButtonBox, QGroupBox, QGridLayout)
from PySide6.QtWidgets import QFileDialog  # WHY: save dialog for export output.
from PySide6.QtWidgets import QProgressDialog  # WHY: show export progress without freezing UI.
from PySide6.QtCore import Qt, QThread, Signal, Slot, QObject  # NEW: threading helpers for smooth UI.
from core.database import Database
import pandas as pd
import os
from core.user_config import load_config, save_config

class RaporlarLoadWorker(QObject):
    """Rapor verilerini arka planda hazirlar."""
    finished = Signal(str, list, list)  # WHY: report_type, headers, rows.
    error = Signal(str)  # WHY: surface errors without blocking UI thread.

    def __init__(self, db, rapor_tur, year, month, tersane_id=0):
        super().__init__()
        self.db = db  # WHY: keep DB access same as before, only off UI thread.
        self.rapor_tur = rapor_tur
        self.year = year
        self.month = month
        self.tersane_id = tersane_id or 0  # WHY: normalize to keep behavior consistent with global (0) mode.

    @Slot()
    def run(self):
        try:
            headers = []
            rows = []
            if self.rapor_tur == "Çalışan Saatleri":
                headers = ["Personel", "Çalışılan Gün", "İzin Günü", "Normal", "Mesai", "Toplam", "Toplam Gün", "Birim", "Açıklamalar"]
                puantaj_data = self.db.get_dashboard_data(self.year, self.month, self.tersane_id)  # WHY: filter by active tersane.
                for item in puantaj_data:
                    ad_soyad = item.get('ad_soyad', '')
                    # Gerçek çalışılan gün sayısı (normal > 0 olan günler, pazar hariç)
                    with self.db.get_connection() as conn:
                        c = conn.cursor()
                        if self.tersane_id and self.tersane_id > 0:
                            c.execute("""
                                SELECT COUNT(*) FROM gunluk_kayit 
                                WHERE ad_soyad=? 
                                AND strftime('%Y', tarih)=? 
                                AND strftime('%m', tarih)=? 
                                AND hesaplanan_normal > 0
                                AND strftime('%w', tarih) != '0'
                                AND tersane_id = ?
                            """, (ad_soyad, str(self.year), f"{self.month:02d}", self.tersane_id))  # WHY: keep tersane scope.
                        else:
                            c.execute("""
                                SELECT COUNT(*) FROM gunluk_kayit 
                                WHERE ad_soyad=? 
                                AND strftime('%Y', tarih)=? 
                                AND strftime('%m', tarih)=? 
                                AND hesaplanan_normal > 0
                                AND strftime('%w', tarih) != '0'
                            """, (ad_soyad, str(self.year), f"{self.month:02d}"))
                        calisan_gun = c.fetchone()[0] or 0
                        # İzin günleri
                        c.execute("""
                            SELECT SUM(gun_sayisi) FROM izin_takip 
                            WHERE ad_soyad=? 
                            AND strftime('%Y', izin_tarihi)=? 
                            AND strftime('%m', izin_tarihi)=?
                        """, (ad_soyad, str(self.year), f"{self.month:02d}"))
                        izin_gun = c.fetchone()[0] or 0
                        # Açıklamalar: Ekstra ödeme notu + Avans notu
                        c.execute("SELECT ekstra_odeme_not, avans_not FROM personel WHERE ad_soyad=?", (ad_soyad,))
                        notes_result = c.fetchone()
                        notes_text = ""
                        if notes_result:
                            ekstra_not = notes_result[0] or ""
                            avans_not = notes_result[1] or ""
                            if ekstra_not:
                                notes_text += f"Ekstra: {ekstra_not}"
                            if avans_not:
                                if notes_text:
                                    notes_text += " | "
                                notes_text += f"Avans: {avans_not}"
                    toplam = (item.get('top_normal', 0) or 0) + (item.get('top_mesai', 0) or 0)
                    toplam_gun = calisan_gun + izin_gun
                    birim = "Yevmiye" if item.get('yevmiyeci_mi', 0) else "Saat"
                    rows.append([
                        ad_soyad,
                        str(calisan_gun),
                        f"{izin_gun:.1f}",
                        f"{item.get('top_normal', 0):.1f}",
                        f"{item.get('top_mesai', 0):.1f}",
                        f"{toplam:.1f}",
                        f"{toplam_gun:.1f}",
                        birim,
                        notes_text
                    ])
            elif self.rapor_tur == "Devamsızlık İstatistikleri":
                headers = ["Personel", "Devamsız Gün", "İzin Günü", "Açıklama"]
                izin_list = self.db.get_izin_list(self.year, self.month, tersane_id=self.tersane_id)  # WHY: scope by active tersane.
                for ad, tarih, tur, gun, aciklama, onay in [(r[1], r[2], r[3], r[4], r[5], r[6]) for r in izin_list]:
                    rows.append([ad, "0", str(int(gun)), f"{tur} - {aciklama}"])
            elif self.rapor_tur == "Ekip Bazında Analiz":
                headers = ["Ekip", "Personel Sayısı", "Saat Toplam", "Saat Ort.", "Yevmiye Toplam", "Yevmiye Ort."]
                puantaj_data = self.db.get_dashboard_data(self.year, self.month, self.tersane_id)  # WHY: filter by active tersane.
                ekip_dict = {}
                for item in puantaj_data:
                    ekip = item.get('ekip', 'Diğer')
                    if ekip not in ekip_dict:
                        ekip_dict[ekip] = {'count': 0, 'count_saat': 0, 'count_yev': 0, 'saat': 0, 'yev': 0}
                    ekip_dict[ekip]['count'] += 1
                    toplam = (item.get('top_normal', 0) or 0) + (item.get('top_mesai', 0) or 0)
                    if item.get('yevmiyeci_mi', 0):
                        ekip_dict[ekip]['count_yev'] += 1
                        ekip_dict[ekip]['yev'] += toplam
                    else:
                        ekip_dict[ekip]['count_saat'] += 1
                        ekip_dict[ekip]['saat'] += toplam
                for ekip, data in ekip_dict.items():
                    avg_saat = data['saat'] / data['count_saat'] if data['count_saat'] > 0 else 0
                    avg_yev = data['yev'] / data['count_yev'] if data['count_yev'] > 0 else 0
                    rows.append([
                        ekip,
                        str(data['count']),
                        f"{data['saat']:.1f}",
                        f"{avg_saat:.1f}",
                        f"{data['yev']:.1f}",
                        f"{avg_yev:.1f}"
                    ])
            elif self.rapor_tur == "Personel Performansı":
                headers = ["Personel", "Çalışılan Gün", "Mesai", "Birim", "Devamsızlık", "Puan"]
                puantaj_data = self.db.get_dashboard_data(self.year, self.month, self.tersane_id)  # WHY: filter by active tersane.
                for item in puantaj_data:
                    puan = 50 + ((item.get('top_mesai', 0) or 0) * 5)
                    rows.append([
                        item.get('ad_soyad', ''),
                        "20",
                        f"{item.get('top_mesai', 0):.1f}",
                        "Yevmiye" if item.get('yevmiyeci_mi', 0) else "Saat",
                        "0",
                        f"{puan:.0f}"
                    ])
            elif self.rapor_tur == "Aylık Özet":
                headers = ["Personel", "Gün", "Normal", "Mesai", "Birim", "İzin", "Avans"]
                puantaj_data = self.db.get_dashboard_data(self.year, self.month, self.tersane_id)  # WHY: filter by active tersane.
                for item in puantaj_data:
                    rows.append([
                        item.get('ad_soyad', ''),
                        "20",
                        f"{item.get('top_normal', 0):.1f}",
                        f"{item.get('top_mesai', 0):.1f}",
                        "Yevmiye" if item.get('yevmiyeci_mi', 0) else "Saat",
                        "0",
                        f"{item.get('avans', 0):.2f}"
                    ])
            self.finished.emit(self.rapor_tur, headers, rows)
        except Exception as e:
            self.error.emit(str(e))


class ExportWorker(QObject):  # WHY: generic worker for background export tasks.
    finished = Signal(object)  # WHY: return payload (path/status) to UI thread.
    error = Signal(str)  # WHY: surface errors safely without crashing UI.

    def __init__(self, task_fn):  # WHY: keep worker reusable for different export tasks.
        super().__init__()  # WHY: initialize QObject for signal/slot usage.
        self._task_fn = task_fn  # WHY: store export task callable.
        self._stop_requested = False  # WHY: allow cooperative cancel handling.

    def request_stop(self):  # WHY: allow UI to request a safe stop.
        self._stop_requested = True  # WHY: set flag without killing thread.

    def should_stop(self):  # WHY: shared stop check for long loops.
        return self._stop_requested or QThread.currentThread().isInterruptionRequested()  # WHY: respect both flags.

    @Slot()
    def run(self):  # WHY: thread entry point.
        try:
            result = self._task_fn(self)  # WHY: execute export task with stop-aware worker.
            self.finished.emit(result)  # WHY: notify UI on completion.
        except Exception as e:
            self.error.emit(str(e))  # WHY: forward exception to UI thread.

class RaporlarPage(QWidget):
    def __init__(self, signal_manager):
        super().__init__()
        self.db = Database()
        self.signal_manager = signal_manager
        self.tersane_id = 0  # NEW: active tersane id for report scoping.
        self._needs_refresh = False  # NEW: lazy-load flag to avoid heavy refresh on hidden tabs.
        self._load_thread = None  # NEW: keep thread reference to avoid premature GC.
        self._load_worker = None  # NEW: keep worker reference to avoid GC while thread runs.
        self._export_thread = None  # WHY: keep background export thread reference.
        self._export_worker = None  # WHY: keep export worker alive during run.
        self._export_dialog = None  # WHY: progress dialog for export operations.
        self._export_done_cb = None  # WHY: store export completion callback.
        self._export_cancelled = False  # WHY: track cancel to suppress success toast.
        self.setup_ui()
        self.load_data()
        self.signal_manager.data_updated.connect(self._on_data_updated)  # NEW: lazy refresh to avoid hidden-tab work.

    def set_tersane_id(self, tersane_id, refresh=True):
        """Global tersane seçiciden gelen tersane_id'yi set eder ve verileri yeniler."""
        self.tersane_id = tersane_id
        self._needs_refresh = True  # WHY: mark dirty; refresh can be deferred.
        if refresh:
            self.update_view()  # WHY: only visible page refreshes to keep UI smooth.

    def update_view(self):
        """Görünür sayfa için güncel tersane verilerini yükle."""
        self._needs_refresh = False  # WHY: clear dirty flag after refresh.
        self.load_data()

    def refresh_if_needed(self):
        """Lazy-load için: sayfa görünür olduğunda gerekiyorsa güncelle."""
        if self._needs_refresh:
            self.update_view()

    def _get_active_tersane_label(self):  # WHY: centralize export titles with active tersane name.
        if self.tersane_id and self.tersane_id > 0:  # WHY: include selected tersane in export metadata.
            tersane = self.db.get_tersane(self.tersane_id)  # WHY: fetch tersane name for display.
            return tersane['ad'] if tersane else f"ID {self.tersane_id}"  # WHY: fallback keeps export usable if name missing.
        return "Tüm Tersaneler"  # WHY: preserve global mode label when no tersane selected.

    def _on_data_updated(self):
        """Veri değiştiğinde sadece görünürsek yenile (lazy)."""
        if not self.isVisible():
            self._needs_refresh = True  # WHY: defer heavy refresh until tab is visible.
            return
        self.update_view()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Başlık
        title = QLabel("📊 Raporlar ve İstatistikler")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #fff; margin-bottom: 2px;")
        layout.addWidget(title)

        desc = QLabel("Rapor türüne göre dönemsel analizleri görüntüleyin ve dışa aktarın.")
        desc.setStyleSheet("color: #999; font-size: 12px; margin-bottom: 10px;")
        layout.addWidget(desc)
        
        # Filtre ve Rapor Seçimi
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Rapor Türü:"))
        self.combo_rapor = QComboBox()
        self.combo_rapor.addItems(["Çalışan Saatleri", "Devamsızlık İstatistikleri", 
                                   "Ekip Bazında Analiz", "Personel Performansı",
                                   "Aylık Özet"])
        self.combo_rapor.currentTextChanged.connect(self.load_data)
        filter_layout.addWidget(self.combo_rapor)
        
        filter_layout.addWidget(QLabel("Ay:"))
        self.combo_month = QComboBox()
        self.combo_month.addItems(["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
                                   "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"])
        today = datetime.now()
        self.combo_month.setCurrentIndex(today.month - 1)
        self.combo_month.currentTextChanged.connect(self.load_data)
        filter_layout.addWidget(self.combo_month)
        
        filter_layout.addWidget(QLabel("Yıl:"))
        self.combo_year = QComboBox()
        self.combo_year.addItems([str(y) for y in range(2024, 2030)])
        self.combo_year.setCurrentText(str(today.year))
        self.combo_year.currentTextChanged.connect(self.load_data)
        filter_layout.addWidget(self.combo_year)
        
        layout.addLayout(filter_layout)
        
        self.lbl_rapor_desc = QLabel("")
        self.lbl_rapor_desc.setStyleSheet("color: #999; font-size: 11px; padding: 4px 8px; background-color: #2a2a2a; border-radius: 4px;")
        self.lbl_rapor_desc.setWordWrap(True)
        layout.addWidget(self.lbl_rapor_desc)
        self.combo_rapor.currentTextChanged.connect(self._update_rapor_desc)
        self._update_rapor_desc()
        
        # Tablo
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("QTableWidget { alternate-background-color: #2a2a2a; }")
        layout.addWidget(self.table)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        btn_excel = QPushButton("📊 Excel'e Aktar")
        btn_excel.setStyleSheet("background-color: #4CAF50; color: white; padding: 8px;")
        btn_excel.clicked.connect(self.export_to_excel)
        btn_layout.addWidget(btn_excel)
        
        btn_yazdir = QPushButton("🖨️ Yazdır")
        btn_yazdir.setStyleSheet("background-color: #2196F3; color: white; padding: 8px;")
        btn_yazdir.clicked.connect(self.print_report)
        btn_layout.addWidget(btn_yazdir)

        btn_icmal = QPushButton("📋 SGK Mesai İcmal")
        btn_icmal.setStyleSheet("background-color: #7B1FA2; color: white; padding: 8px; font-weight: bold;")
        btn_icmal.clicked.connect(self.export_sgk_icmal)
        btn_layout.addWidget(btn_icmal)

        layout.addLayout(btn_layout)

    def load_data(self):
        self._start_load_worker()  # WHY: load report data in background to keep UI responsive.

    def _update_rapor_desc(self):
        descs = {
            "Çalışan Saatleri": "Her personelin çalışılan gün, normal saat, mesai saati ve izin günü toplamlarını gösterir.",
            "Devamsızlık İstatistikleri": "Dönem içindeki izin kayıtlarını ve devamsızlık durumlarını listeler.",
            "Ekip Bazında Analiz": "Ekiplere göre personel sayısı, toplam ve ortalama çalışma saatlerini karşılaştırır.",
            "Personel Performansı": "Her personelin çalışma günü, mesai miktarı ve performans puanını gösterir.",
            "Aylık Özet": "Dönemin genel özeti: gün sayısı, normal, mesai, izin ve avans bilgileri."
        }
        rapor = self.combo_rapor.currentText()
        self.lbl_rapor_desc.setText(descs.get(rapor, ""))

    def _start_load_worker(self):
        """Arka planda rapor verisi hazırlar (UI donmasini engeller)."""
        if self._load_thread and self._load_thread.isRunning():
            return  # WHY: do not start a second load while one is running.
        rapor_tur = self.combo_rapor.currentText()
        month = self.combo_month.currentIndex() + 1
        year = int(self.combo_year.currentText())
        self._load_thread = QThread()  # WHY: run heavy DB work off the UI thread.
        worker = RaporlarLoadWorker(self.db, rapor_tur, year, month, self.tersane_id)
        self._load_worker = worker  # WHY: keep a strong reference to avoid GC while running.
        worker.moveToThread(self._load_thread)  # WHY: execute worker in background thread.
        self._load_thread.started.connect(worker.run)  # WHY: start work when thread starts.
        worker.finished.connect(self._on_load_finished)  # WHY: update UI when data is ready.
        worker.error.connect(self._on_load_error)  # WHY: show errors without crashing UI.
        worker.finished.connect(self._load_thread.quit)  # WHY: stop thread event loop after completion.
        worker.finished.connect(worker.deleteLater)  # WHY: free worker object safely in Qt.
        worker.error.connect(self._load_thread.quit)  # WHY: stop thread on error to avoid orphan threads.
        worker.error.connect(worker.deleteLater)  # WHY: free worker on error path.
        self._load_thread.finished.connect(self._on_load_thread_finished)  # WHY: clear references only after thread stops.
        self._load_thread.finished.connect(self._load_thread.deleteLater)  # WHY: free thread object after finish.
        self._load_thread.start()  # WHY: start background work now that signals are wired.

    def _on_load_finished(self, rapor_tur, headers, rows):
        """Arka plan rapor sonucunu tabloya uygular."""
        try:
            self.table.setColumnCount(len(headers))
            self.table.setHorizontalHeaderLabels(headers)
            self.table.setRowCount(len(rows))
            for r_idx, row in enumerate(rows):
                for c_idx, val in enumerate(row):
                    self.table.setItem(r_idx, c_idx, QTableWidgetItem(str(val)))
            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        except RuntimeError:
            pass  # SAFEGUARD: UI object may be gone; ignore late signals.

    def _on_load_error(self, msg):
        """Arka plan rapor hatasi."""
        QMessageBox.critical(self, "Hata", f"Rapor yuklenemedi: {msg}")

    def _on_load_thread_finished(self):
        """Thread kapaninca referanslari temizle."""
        self._load_thread = None  # WHY: clear thread ref after it has fully stopped.
        self._load_worker = None  # WHY: clear worker ref after thread completion.

    def show_calisan_saatleri(self, year, month):
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels(["Personel", "Çalışılan Gün", "İzin Günü", "Normal", "Mesai", "Toplam", "Toplam Gün", "Birim", "Açıklamalar"])
        self.table.setRowCount(0)
        
        puantaj_data = self.db.get_dashboard_data(year, month, self.tersane_id)  # WHY: scope to active tersane if selected.
        for item in puantaj_data:
            row = self.table.rowCount()
            self.table.insertRow(row)
            ad_soyad = item.get('ad_soyad', '')
            
            # Gerçek çalışılan gün sayısı (normal > 0 olan günler, pazar hariç)
            with self.db.get_connection() as conn:
                c = conn.cursor()
                if self.tersane_id and self.tersane_id > 0:
                    c.execute("""
                        SELECT COUNT(*) FROM gunluk_kayit 
                        WHERE ad_soyad=? 
                        AND strftime('%Y', tarih)=? 
                        AND strftime('%m', tarih)=? 
                        AND hesaplanan_normal > 0
                        AND strftime('%w', tarih) != '0'
                        AND tersane_id = ?
                    """, (ad_soyad, str(year), f"{month:02d}", self.tersane_id))  # WHY: keep tersane scope.
                else:
                    c.execute("""
                        SELECT COUNT(*) FROM gunluk_kayit 
                        WHERE ad_soyad=? 
                        AND strftime('%Y', tarih)=? 
                        AND strftime('%m', tarih)=? 
                        AND hesaplanan_normal > 0
                        AND strftime('%w', tarih) != '0'
                    """, (ad_soyad, str(year), f"{month:02d}"))
                calisan_gun = c.fetchone()[0] or 0
                
                # İzin günleri
                c.execute("""
                    SELECT SUM(gun_sayisi) FROM izin_takip 
                    WHERE ad_soyad=? 
                    AND strftime('%Y', izin_tarihi)=? 
                    AND strftime('%m', izin_tarihi)=?
                """, (ad_soyad, str(year), f"{month:02d}"))
                izin_gun = c.fetchone()[0] or 0
                
                # Açıklamalar: Ekstra ödeme notu + Avans notu
                c.execute("SELECT ekstra_odeme_not, avans_not FROM personel WHERE ad_soyad=?", (ad_soyad,))
                notes_result = c.fetchone()
                notes_text = ""
                if notes_result:
                    ekstra_not = notes_result[0] or ""
                    avans_not = notes_result[1] or ""
                    if ekstra_not:
                        notes_text += f"Ekstra: {ekstra_not}"
                    if avans_not:
                        if notes_text:
                            notes_text += " | "
                        notes_text += f"Avans: {avans_not}"
            
            self.table.setItem(row, 0, QTableWidgetItem(ad_soyad))
            self.table.setItem(row, 1, QTableWidgetItem(str(calisan_gun)))
            self.table.setItem(row, 2, QTableWidgetItem(f"{izin_gun:.1f}"))
            self.table.setItem(row, 3, QTableWidgetItem(f"{item.get('top_normal', 0):.1f}"))
            self.table.setItem(row, 4, QTableWidgetItem(f"{item.get('top_mesai', 0):.1f}"))
            toplam = item.get('top_normal', 0) + item.get('top_mesai', 0)
            toplam_gun = calisan_gun + izin_gun
            birim = "Yevmiye" if item.get('yevmiyeci_mi', 0) else "Saat"
            self.table.setItem(row, 5, QTableWidgetItem(f"{toplam:.1f}"))
            self.table.setItem(row, 6, QTableWidgetItem(f"{toplam_gun:.1f}"))
            self.table.setItem(row, 7, QTableWidgetItem(birim))
            self.table.setItem(row, 8, QTableWidgetItem(notes_text))

    def show_devamsizlik(self, year, month):
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Personel", "Devamsız Gün", "İzin Günü", "Açıklama"])
        self.table.setRowCount(0)
        
        izin_list = self.db.get_izin_list(year, month, tersane_id=self.tersane_id)  # WHY: scope to active tersane if selected.
        for ad, tarih, tur, gun, aciklama, onay in izin_list:
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(ad))
            self.table.setItem(row, 1, QTableWidgetItem("0"))
            self.table.setItem(row, 2, QTableWidgetItem(str(int(gun))))
            self.table.setItem(row, 3, QTableWidgetItem(f"{tur} - {aciklama}"))

    def show_ekip_analiz(self, year, month):
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["Ekip", "Personel Sayısı", "Saat Toplam", "Saat Ort.", "Yevmiye Toplam", "Yevmiye Ort."])
        self.table.setRowCount(0)
        
        puantaj_data = self.db.get_dashboard_data(year, month, self.tersane_id)  # WHY: scope to active tersane if selected.
        ekip_dict = {}
        for item in puantaj_data:
            ekip = item.get('ekip', 'Diğer')
            if ekip not in ekip_dict:
                ekip_dict[ekip] = {'count': 0, 'count_saat': 0, 'count_yev': 0, 'saat': 0, 'yev': 0}
            ekip_dict[ekip]['count'] += 1
            toplam = item.get('top_normal', 0) + item.get('top_mesai', 0)
            if item.get('yevmiyeci_mi', 0):
                ekip_dict[ekip]['count_yev'] += 1
                ekip_dict[ekip]['yev'] += toplam
            else:
                ekip_dict[ekip]['count_saat'] += 1
                ekip_dict[ekip]['saat'] += toplam
        
        for ekip, data in ekip_dict.items():
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(ekip))
            self.table.setItem(row, 1, QTableWidgetItem(str(data['count'])))
            avg_saat = data['saat'] / data['count_saat'] if data['count_saat'] > 0 else 0
            avg_yev = data['yev'] / data['count_yev'] if data['count_yev'] > 0 else 0
            self.table.setItem(row, 2, QTableWidgetItem(f"{data['saat']:.1f}"))
            self.table.setItem(row, 3, QTableWidgetItem(f"{avg_saat:.1f}"))
            self.table.setItem(row, 4, QTableWidgetItem(f"{data['yev']:.1f}"))
            self.table.setItem(row, 5, QTableWidgetItem(f"{avg_yev:.1f}"))

    def show_personel_performans(self, year, month):
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["Personel", "Çalışılan Gün", "Mesai", "Birim", "Devamsızlık", "Puan"])
        self.table.setRowCount(0)
        
        puantaj_data = self.db.get_dashboard_data(year, month, self.tersane_id)  # WHY: scope to active tersane if selected.
        for item in puantaj_data:
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(item.get('ad_soyad', '')))
            self.table.setItem(row, 1, QTableWidgetItem("20"))
            self.table.setItem(row, 2, QTableWidgetItem(f"{item.get('top_mesai', 0):.1f}"))
            self.table.setItem(row, 3, QTableWidgetItem("Yevmiye" if item.get('yevmiyeci_mi', 0) else "Saat"))
            self.table.setItem(row, 4, QTableWidgetItem("0"))
            # Puan hesaplama: daha fazla mesai = daha yüksek puan
            puan = 50 + (item.get('top_mesai', 0) * 5)
            self.table.setItem(row, 5, QTableWidgetItem(f"{puan:.0f}"))

    def show_aylik_ozet(self, year, month):
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["Personel", "Gün", "Normal", "Mesai", "Birim", "İzin", "Avans"])
        self.table.setRowCount(0)
        
        puantaj_data = self.db.get_dashboard_data(year, month, self.tersane_id)  # WHY: scope to active tersane if selected.
        for item in puantaj_data:
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(item.get('ad_soyad', '')))
            self.table.setItem(row, 1, QTableWidgetItem("20"))
            self.table.setItem(row, 2, QTableWidgetItem(f"{item.get('top_normal', 0):.1f}"))
            self.table.setItem(row, 3, QTableWidgetItem(f"{item.get('top_mesai', 0):.1f}"))
            self.table.setItem(row, 4, QTableWidgetItem("Yevmiye" if item.get('yevmiyeci_mi', 0) else "Saat"))
            self.table.setItem(row, 5, QTableWidgetItem("0"))
            self.table.setItem(row, 6, QTableWidgetItem(f"{item.get('avans', 0):.2f}"))

    def _export_to_excel_legacy(self):  # WHY: keep original sync export as reference; replaced by threaded version below.
        rapor_tur = self.combo_rapor.currentText()
        month = self.combo_month.currentIndex() + 1
        year = int(self.combo_year.currentText())
        
        try:
            # Tablo verilerini DataFrame'e dönüştür
            data = []
            for row in range(self.table.rowCount()):
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)

            headers = []
            for col in range(self.table.columnCount()):
                header_item = self.table.horizontalHeaderItem(col)
                headers.append(header_item.text() if header_item else f"Kolon {col+1}")

            df = pd.DataFrame(data, columns=headers)
            filename = f"Rapor_{rapor_tur}_{year}-{month:02d}.xlsx"
            
            from PySide6.QtWidgets import QFileDialog
            cfg = load_config()
            last_dir = cfg.get("last_export_dir", "")
            default_path = os.path.join(last_dir, filename) if last_dir else filename
            path, _ = QFileDialog.getSaveFileName(self, "Raporu Kaydet", default_path, "Excel (*.xlsx)")
            if not path:
                return
            try:
                cfg["last_export_dir"] = os.path.dirname(path)
                save_config(cfg)
            except Exception:
                pass

            try:
                import openpyxl
                with pd.ExcelWriter(path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name=rapor_tur, startrow=3)
                    ws = writer.sheets[rapor_tur]

                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter

                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
                    tcell = ws.cell(row=1, column=1)
                    tcell.value = f"{rapor_tur} - {year}-{month:02d}"
                    tcell.font = Font(bold=True, size=14)
                    tcell.alignment = Alignment(horizontal='center')

                    header_row = 4
                    header_fill = PatternFill(start_color='FF424242', end_color='FF424242', fill_type='solid')
                    header_font = Font(bold=True, color='FFFFFFFF')
                    thin = Side(border_style="thin", color="FFAAAAAA")
                    for col_cell in ws[header_row]:
                        col_cell.fill = header_fill
                        col_cell.font = header_font
                        col_cell.alignment = Alignment(horizontal='center', vertical='center')
                        col_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                    dims = {}
                    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row):
                        for cell in row:
                            if cell.value is not None:
                                dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len(str(cell.value)))
                    for col, value in dims.items():
                        ws.column_dimensions[col].width = min(max(value + 2, 10), 35)

                    ws.freeze_panes = f"A{header_row+1}"

                    fill_gray = PatternFill(start_color='FFF5F5F5', end_color='FFF5F5F5', fill_type='solid')
                    for idx, row in enumerate(ws.iter_rows(min_row=header_row+1, max_row=ws.max_row), start=0):
                        if idx % 2 == 0:
                            for cell in row:
                                cell.fill = fill_gray

                    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

                QMessageBox.information(self, "Başarılı", f"Rapor kaydedildi: {path}")
            except ImportError:
                df.to_excel(path, index=False, sheet_name=rapor_tur)
                QMessageBox.information(self, "Başarılı", f"Rapor kaydedildi (stilsiz): {path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel export hatası: {e}")

    def _start_export_worker(self, task_fn, done_cb=None, label="Rapor hazırlanıyor..."):  # WHY: shared export runner to keep UI responsive.
        if self._export_thread and self._export_thread.isRunning():  # WHY: avoid overlapping exports that could lock files.
            QMessageBox.information(self, "Bilgi", "Devam eden bir dışa aktarma var.")  # WHY: inform user without starting another thread.
            return
        self._export_done_cb = done_cb  # WHY: keep per-export UI completion handler.
        self._export_cancelled = False  # WHY: reset cancel state for each new export.
        self._export_dialog = QProgressDialog(label, None, 0, 0, self)  # WHY: show indeterminate progress during export.
        self._export_dialog.setWindowModality(Qt.WindowModal)  # WHY: keep modal behavior consistent with other dialogs.
        self._export_dialog.setAutoClose(False)  # WHY: we close explicitly on signals to avoid stuck dialogs.
        self._export_dialog.setAutoReset(False)  # WHY: prevent auto-reset from hiding progress early.
        self._export_dialog.setMinimumDuration(0)  # WHY: show immediately to avoid perceived freeze.
        self._export_dialog.setAttribute(Qt.WA_DeleteOnClose, False)  # WHY: keep dialog alive for late signals.
        self._export_dialog.canceled.connect(self._on_export_dialog_canceled)  # WHY: allow safe cancel without crashing.
        self._export_dialog.rejected.connect(self._on_export_dialog_canceled)  # WHY: handle window close (X) safely.
        self._export_dialog.show()  # WHY: show progress feedback during background work.

        self._export_thread = QThread()  # WHY: run heavy export in background.
        worker = ExportWorker(task_fn)  # WHY: reuse generic export worker for different tasks.
        self._export_worker = worker  # WHY: keep a strong reference to prevent GC.
        worker.moveToThread(self._export_thread)  # WHY: execute worker in background thread.
        self._export_thread.started.connect(worker.run)  # WHY: start export when thread starts.
        worker.finished.connect(self._on_export_finished)  # WHY: handle success payload in UI thread.
        worker.finished.connect(self._export_dialog.accept)  # WHY: close dialog on normal completion.
        worker.finished.connect(self._export_thread.quit)  # WHY: stop thread event loop after completion.
        worker.finished.connect(worker.deleteLater)  # WHY: free worker safely.
        worker.error.connect(self._on_export_error)  # WHY: surface errors without freezing UI.
        worker.error.connect(self._export_thread.quit)  # WHY: stop thread on error.
        worker.error.connect(worker.deleteLater)  # WHY: free worker on error path.
        self._export_thread.finished.connect(self._on_export_thread_finished)  # WHY: clear refs after thread stops.
        self._export_thread.finished.connect(self._export_thread.deleteLater)  # WHY: free thread object after finish.
        self._export_thread.start()  # WHY: kick off background export.

    def _on_export_finished(self, result):  # WHY: centralize export completion handling.
        if self._export_dialog:  # WHY: close dialog if still alive.
            try:
                self._export_dialog.close()  # WHY: ensure dialog closes on completion.
            except RuntimeError:
                pass  # WHY: dialog already deleted; ignore safely.
        self._export_dialog = None  # WHY: release UI reference after safe close.
        if self._export_cancelled:  # WHY: skip success dialog when user cancelled.
            return
        if self._export_done_cb:  # WHY: allow per-export success handling.
            self._export_done_cb(result)  # WHY: call UI completion handler with payload.
            self._export_done_cb = None  # WHY: clear handler after use.

    def _on_export_error(self, msg):  # WHY: handle export errors uniformly.
        if self._export_dialog:  # WHY: close dialog on error.
            try:
                self._export_dialog.close()  # WHY: avoid stuck progress dialog on failure.
            except RuntimeError:
                pass  # WHY: dialog already deleted; ignore safely.
        self._export_dialog = None  # WHY: release UI reference after safe close.
        QMessageBox.critical(self, "Hata", f"Dışa aktarma sırasında hata: {msg}")  # WHY: show error without crashing UI.

    def _on_export_dialog_canceled(self):  # WHY: safe cancel handling for export progress dialog.
        self._export_cancelled = True  # WHY: mark cancel to skip success toast later.
        if self._export_worker:  # WHY: request cooperative stop for worker.
            self._export_worker.request_stop()  # WHY: avoid hard thread termination.
        if self._export_thread:  # WHY: set interruption flag for worker to observe.
            self._export_thread.requestInterruption()  # WHY: allow cooperative stop in loops.
        if self._export_dialog:  # WHY: update dialog text to show canceling.
            try:
                self._export_dialog.setLabelText("İptal ediliyor...")  # WHY: immediate feedback on cancel.
                self._export_dialog.setCancelButtonText("")  # WHY: prevent repeated cancel clicks.
            except RuntimeError:
                pass  # WHY: dialog already deleted; ignore safely.

    def _on_export_thread_finished(self):  # WHY: clean up thread refs safely after export.
        self._export_thread = None  # WHY: clear thread ref after it stops.
        self._export_worker = None  # WHY: clear worker ref after thread completion.

    def export_to_excel(self):  # WHY: threaded export to keep UI responsive.
        rapor_tur = self.combo_rapor.currentText()
        month = self.combo_month.currentIndex() + 1
        year = int(self.combo_year.currentText())
        tersane_label = self._get_active_tersane_label()  # WHY: include active tersane in title.

        # Tablo verilerini DataFrame'e dönüştür
        data = []
        for row in range(self.table.rowCount()):
            row_data = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else "")
            data.append(row_data)

        if not data:
            QMessageBox.information(self, "Bilgi", "Dışa aktarılacak veri yok.")  # WHY: avoid empty exports.
            return

        headers = []
        for col in range(self.table.columnCount()):
            header_item = self.table.horizontalHeaderItem(col)
            headers.append(header_item.text() if header_item else f"Kolon {col+1}")

        df = pd.DataFrame(data, columns=headers)
        filename = f"Rapor_{rapor_tur}_{year}-{month:02d}.xlsx"
        cfg = load_config()
        last_dir = cfg.get("last_export_dir", "")
        default_path = os.path.join(last_dir, filename) if last_dir else filename
        path, _ = QFileDialog.getSaveFileName(self, "Raporu Kaydet", default_path, "Excel (*.xlsx)")
        if not path:
            return
        try:
            cfg["last_export_dir"] = os.path.dirname(path)
            save_config(cfg)
        except Exception:
            pass

        def _task(worker):  # WHY: run export off the UI thread.
            if worker.should_stop():
                return {"status": "cancelled"}  # WHY: allow user-initiated cancel.
            try:
                import openpyxl  # noqa: F401  # WHY: ensure styled export is possible.
                with pd.ExcelWriter(path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name=rapor_tur, startrow=3)
                    ws = writer.sheets[rapor_tur]

                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter

                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
                    tcell = ws.cell(row=1, column=1)
                    tcell.value = f"{rapor_tur} - {year}-{month:02d} - {tersane_label}"  # WHY: include tersane in title.
                    tcell.font = Font(bold=True, size=14)
                    tcell.alignment = Alignment(horizontal='center')

                    header_row = 4
                    header_fill = PatternFill(start_color='FF424242', end_color='FF424242', fill_type='solid')
                    header_font = Font(bold=True, color='FFFFFFFF')
                    thin = Side(border_style="thin", color="FFAAAAAA")
                    for col_cell in ws[header_row]:
                        col_cell.fill = header_fill
                        col_cell.font = header_font
                        col_cell.alignment = Alignment(horizontal='center', vertical='center')
                        col_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                    dims = {}
                    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row):
                        for cell in row:
                            if cell.value is not None:
                                dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len(str(cell.value)))
                    for col, value in dims.items():
                        ws.column_dimensions[col].width = min(max(value + 2, 10), 35)

                    ws.freeze_panes = f"A{header_row+1}"

                    fill_gray = PatternFill(start_color='FFF5F5F5', end_color='FFF5F5F5', fill_type='solid')
                    for idx, row in enumerate(ws.iter_rows(min_row=header_row+1, max_row=ws.max_row), start=0):
                        if idx % 2 == 0:
                            for cell in row:
                                cell.fill = fill_gray

                    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

                return {"status": "ok", "path": path, "plain": False}  # WHY: styled export success.
            except ImportError:
                df.to_excel(path, index=False, sheet_name=rapor_tur)
                return {"status": "ok", "path": path, "plain": True}  # WHY: fallback to plain export when openpyxl missing.

        def _done(result):  # WHY: handle export completion on UI thread.
            if not result:
                return  # WHY: no payload, nothing to show.
            if result.get("status") == "cancelled":
                return  # WHY: skip success dialog on cancel.
            if result.get("plain"):
                QMessageBox.information(self, "Başarılı", f"Rapor kaydedildi (stilsiz): {result.get('path', path)}")  # WHY: preserve legacy fallback message.
            else:
                QMessageBox.information(self, "Başarılı", f"Rapor kaydedildi: {result.get('path', path)}")  # WHY: confirm styled export.

        self._start_export_worker(_task, done_cb=_done, label="Rapor hazırlanıyor...")  # WHY: run export in background.

    def print_report(self):
        QMessageBox.information(self, "Yazdırma", "Yazdırma özelliği yakında gelecek.")

    # ------------------------------------------------------------------
    # SGK MESAİ İCMAL
    # ------------------------------------------------------------------

    def export_sgk_icmal(self):
        """SGK için Luca formatında aylık mesai icmal tablosunu Excel'e aktarır."""
        dlg = _SgkIcmalDialog(self)
        if dlg.exec() != QDialog.Accepted:
            return

        bas_yil, bas_ay, bit_yil, bit_ay = dlg.get_period()
        tersane_id = self.tersane_id
        tersane_label = self._get_active_tersane_label()

        # Dönem içindeki ay listesini oluştur
        ay_listesi = []
        y, m = bas_yil, bas_ay
        while (y, m) <= (bit_yil, bit_ay):
            ay_listesi.append((y, m))
            m += 1
            if m > 12:
                m = 1
                y += 1

        ay_etiketler = [f"{_ay_adi(m)} {y}" for y, m in ay_listesi]

        cfg = load_config()
        last_dir = cfg.get("last_export_dir", "")
        filename = f"SGK_Mesai_Icmal_{bas_yil}{bas_ay:02d}-{bit_yil}{bit_ay:02d}.xlsx"
        default_path = os.path.join(last_dir, filename) if last_dir else filename
        path, _ = QFileDialog.getSaveFileName(self, "İcmali Kaydet", default_path, "Excel (*.xlsx)")
        if not path:
            return
        try:
            cfg["last_export_dir"] = os.path.dirname(path)
            save_config(cfg)
        except Exception:
            pass

        db = self.db

        def _task(worker):
            # Personel listesini çek
            with db.get_connection() as conn:
                c = conn.cursor()
                if tersane_id and tersane_id > 0:
                    personel_rows = c.execute(
                        "SELECT ad_soyad, ise_baslangic, "
                        "CASE WHEN gorevi IS NOT NULL AND TRIM(gorevi)!='' THEN gorevi ELSE COALESCE(ekip_adi,'') END "
                        "FROM personel WHERE tersane_id=? ORDER BY ad_soyad",
                        (tersane_id,)
                    ).fetchall()
                else:
                    personel_rows = c.execute(
                        "SELECT ad_soyad, ise_baslangic, "
                        "CASE WHEN gorevi IS NOT NULL AND TRIM(gorevi)!='' THEN gorevi ELSE COALESCE(ekip_adi,'') END "
                        "FROM personel ORDER BY ad_soyad"
                    ).fetchall()

                # Her personel için aylık FM toplamlarını çek
                # SGK/denetim formatı — iki ayrı durum:
                #   1. Normal mesai günleri: hesaplanan_mesai / 1.5  (sistem 1.5x uygulamış)
                #   2. Tatil/Pazar çalışma:  7.5 saat sabit          (gerçek bir günlük çalışma)
                TATIL_SQL = """
                    SELECT
                        COALESCE(SUM(CASE
                            WHEN (aciklama LIKE '%Tatil%' OR aciklama LIKE '%Pazar%')
                                 AND hesaplanan_mesai > 0
                            THEN 7.5 ELSE 0 END), 0),
                        COALESCE(SUM(CASE
                            WHEN (aciklama NOT LIKE '%Tatil%' AND aciklama NOT LIKE '%Pazar%')
                            THEN hesaplanan_mesai ELSE 0 END), 0)
                    FROM gunluk_kayit
                    WHERE ad_soyad=? AND tarih LIKE ?
                """
                TATIL_SQL_T = TATIL_SQL.replace(
                    "WHERE ad_soyad=? AND tarih LIKE ?",
                    "WHERE ad_soyad=? AND tarih LIKE ? AND tersane_id=?"
                )
                tablo = []
                for idx, (ad_soyad, ise_bas, ekip) in enumerate(personel_rows, start=1):
                    if worker.should_stop():
                        return {"status": "cancelled"}
                    aylik_fm = []
                    toplam_fm = 0.0
                    for yil, ay in ay_listesi:
                        ay_str = f"{yil}-{ay:02d}"
                        if tersane_id and tersane_id > 0:
                            row = c.execute(TATIL_SQL_T, (ad_soyad, f"{ay_str}%", tersane_id)).fetchone()
                        else:
                            row = c.execute(TATIL_SQL, (ad_soyad, f"{ay_str}%")).fetchone()
                        tatil_saat = row[0] if row else 0.0
                        normal_raw = row[1] if row else 0.0
                        fm_val = round(tatil_saat + (normal_raw / 1.5), 2) if (tatil_saat or normal_raw) else 0.0
                        aylik_fm.append(fm_val)
                        toplam_fm += fm_val
                    tablo.append({
                        "sira": idx,
                        "ad_soyad": ad_soyad,
                        "ise_baslangic": ise_bas or "",
                        "gorevi": ekip or "",
                        "aylik_fm": aylik_fm,
                        "toplam_fm": toplam_fm,
                    })

            _yaz_icmal_excel(path, tablo, ay_listesi, ay_etiketler, tersane_label)
            return {"status": "ok", "path": path}

        def _done(result):
            if not result or result.get("status") == "cancelled":
                return
            QMessageBox.information(self, "Başarılı", f"SGK Mesai İcmal kaydedildi:\n{result.get('path', path)}")

        self._start_export_worker(_task, done_cb=_done, label="SGK Mesai İcmal hazırlanıyor...")


# ------------------------------------------------------------------
# Yardımcı: ay adları
# ------------------------------------------------------------------

def _ay_adi(ay_no):
    return ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
            "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"][ay_no - 1]


# ------------------------------------------------------------------
# Yardımcı: dönem seçim dialogu
# ------------------------------------------------------------------

class _SgkIcmalDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("SGK Mesai İcmal - Dönem Seçimi")
        self.setMinimumWidth(340)
        layout = QVBoxLayout(self)

        grp = QGroupBox("Dönem")
        grid = QGridLayout(grp)
        aylar = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
                 "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
        yillar = [str(y) for y in range(2023, 2031)]
        bugun = datetime.now()

        grid.addWidget(QLabel("Başlangıç Ay:"), 0, 0)
        self.cb_bas_ay = QComboBox(); self.cb_bas_ay.addItems(aylar)
        self.cb_bas_ay.setCurrentIndex(bugun.month - 2 if bugun.month > 1 else 0)
        grid.addWidget(self.cb_bas_ay, 0, 1)

        grid.addWidget(QLabel("Başlangıç Yıl:"), 0, 2)
        self.cb_bas_yil = QComboBox(); self.cb_bas_yil.addItems(yillar)
        self.cb_bas_yil.setCurrentText(str(bugun.year))
        grid.addWidget(self.cb_bas_yil, 0, 3)

        grid.addWidget(QLabel("Bitiş Ay:"), 1, 0)
        self.cb_bit_ay = QComboBox(); self.cb_bit_ay.addItems(aylar)
        self.cb_bit_ay.setCurrentIndex(bugun.month - 1)
        grid.addWidget(self.cb_bit_ay, 1, 1)

        grid.addWidget(QLabel("Bitiş Yıl:"), 1, 2)
        self.cb_bit_yil = QComboBox(); self.cb_bit_yil.addItems(yillar)
        self.cb_bit_yil.setCurrentText(str(bugun.year))
        grid.addWidget(self.cb_bit_yil, 1, 3)

        layout.addWidget(grp)

        note = QLabel("SGK/denetim formatı: gerçek çalışılan saat (1.5x ücret katsayısı dahil değil).\nÖrn: 3 saat fazla mesai → icmalde 3 yazar.")
        note.setWordWrap(True)
        note.setStyleSheet("color: #aaa; font-size: 11px;")
        layout.addWidget(note)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self._validate_and_accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _validate_and_accept(self):
        bas_y, bas_m = int(self.cb_bas_yil.currentText()), self.cb_bas_ay.currentIndex() + 1
        bit_y, bit_m = int(self.cb_bit_yil.currentText()), self.cb_bit_ay.currentIndex() + 1
        if (bas_y, bas_m) > (bit_y, bit_m):
            QMessageBox.warning(self, "Hata", "Başlangıç tarihi bitiş tarihinden sonra olamaz.")
            return
        if (bit_y * 12 + bit_m) - (bas_y * 12 + bas_m) > 23:
            QMessageBox.warning(self, "Hata", "En fazla 24 aylık dönem seçilebilir.")
            return
        self.accept()

    def get_period(self):
        return (int(self.cb_bas_yil.currentText()), self.cb_bas_ay.currentIndex() + 1,
                int(self.cb_bit_yil.currentText()), self.cb_bit_ay.currentIndex() + 1)


# ------------------------------------------------------------------
# Excel yazıcı: SGK Mesai İcmal
# ------------------------------------------------------------------

def _yaz_icmal_excel(path, tablo, ay_listesi, ay_etiketler, tersane_label):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mesai İcmal"

    # Renk paleti
    LACIVERT   = "1A237E"   # başlık
    KOYU_MOR   = "4A148C"   # sabit sütun başlıkları
    KOYU_MAVI  = "283593"   # ay başlıkları
    TOPLAM_KOL = "880E4F"   # TOPLAM FM sütunu
    SARI_VURGU = "FFF9C4"   # toplam satırı
    ZEBRA_KOYU = "E8EAF6"   # çift satırlar
    BEYAZ      = "FFFFFF"
    YAZI_BEYAZ = "FFFFFF"
    YAZI_KOYU  = "1A1A2E"

    thin  = Side(border_style="thin",   color="BBBBCC")
    med   = Side(border_style="medium", color="555577")
    thick = Side(border_style="medium", color=LACIVERT)

    def stil(cell, bg=None, bold=False, renk=YAZI_KOYU, hizala="center", wrap=False, border=None, size=10):
        if bg:
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.font = Font(bold=bold, color=renk, size=size, name="Calibri")
        cell.alignment = Alignment(horizontal=hizala, vertical="center", wrap_text=wrap)
        b = border or Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.border = b

    n_ay = len(ay_listesi)
    # Sütun düzeni: A=Sıra, B=Ad Soyadı, C=İşe Giriş, D=Görevi, E..E+n-1=Aylar, E+n=TOPLAM
    COL_SIRA   = 1
    COL_AD     = 2
    COL_GIRIS  = 3
    COL_GOREVI = 4
    COL_AY_BAS = 5
    COL_TOPLAM = COL_AY_BAS + n_ay

    TOPLAM_SUTUN = COL_TOPLAM
    SON_SUTUN    = COL_TOPLAM

    # --- Satır 1: Ana başlık ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=SON_SUTUN)
    c = ws.cell(row=1, column=1)
    c.value = f"FAZLA MESAİ İCMAL CETVELİ  —  {tersane_label}"
    stil(c, bg=LACIVERT, bold=True, renk=BEYAZ, size=13)
    ws.row_dimensions[1].height = 26

    # --- Satır 2: Alt bilgi ---
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=SON_SUTUN)
    c = ws.cell(row=2, column=1)
    c.value = f"Dönem: {ay_etiketler[0]} – {ay_etiketler[-1]}   |   Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}   |   SGK Denetim Formatı (gerçek saat)"
    stil(c, bg="303F9F", bold=False, renk="C5CAE9", size=9)
    ws.row_dimensions[2].height = 16

    # --- Satır 3: Boş ayraç ---
    ws.row_dimensions[3].height = 6

    # --- Satır 4: Sütun başlıkları ---
    baslik_data = [
        (COL_SIRA,   "Sıra",            8),
        (COL_AD,     "Ad Soyadı",       28),
        (COL_GIRIS,  "İşe Giriş",       13),
        (COL_GOREVI, "Görevi / Ekip",   18),
    ]
    for col, metin, genislik in baslik_data:
        c = ws.cell(row=4, column=col, value=metin)
        stil(c, bg=KOYU_MOR, bold=True, renk=BEYAZ)
        ws.column_dimensions[get_column_letter(col)].width = genislik

    for i, etiket in enumerate(ay_etiketler):
        col = COL_AY_BAS + i
        c = ws.cell(row=4, column=col, value=etiket)
        stil(c, bg=KOYU_MAVI, bold=True, renk=BEYAZ)
        ws.column_dimensions[get_column_letter(col)].width = 12

    c = ws.cell(row=4, column=COL_TOPLAM, value="TOPLAM FM")
    stil(c, bg=TOPLAM_KOL, bold=True, renk=BEYAZ)
    ws.column_dimensions[get_column_letter(COL_TOPLAM)].width = 13

    ws.row_dimensions[4].height = 22
    ws.freeze_panes = ws.cell(row=5, column=COL_AY_BAS)

    # --- Veri satırları ---
    for satir_idx, kayit in enumerate(tablo):
        row = 5 + satir_idx
        zebra = ZEBRA_KOYU if satir_idx % 2 == 0 else BEYAZ

        vals = [
            (COL_SIRA,   kayit["sira"],         "center"),
            (COL_AD,     kayit["ad_soyad"],      "left"),
            (COL_GIRIS,  kayit["ise_baslangic"], "center"),
            (COL_GOREVI, kayit["gorevi"],        "left"),
        ]
        for col, val, hizala in vals:
            c = ws.cell(row=row, column=col, value=val)
            stil(c, bg=zebra, hizala=hizala)

        for i, fm_val in enumerate(kayit["aylik_fm"]):
            col = COL_AY_BAS + i
            c = ws.cell(row=row, column=col, value=round(fm_val, 2) if fm_val else "")
            c.number_format = "0.00"
            stil(c, bg=zebra)

        c = ws.cell(row=row, column=COL_TOPLAM, value=round(kayit["toplam_fm"], 2))
        c.number_format = "0.00"
        stil(c, bg="FCE4EC", bold=True,
             border=Border(left=med, right=med, top=thin, bottom=thin))

        ws.row_dimensions[row].height = 17

    # --- Toplam satırı ---
    toplam_row = 5 + len(tablo)
    ws.row_dimensions[toplam_row].height = 20

    c = ws.cell(row=toplam_row, column=COL_SIRA, value="")
    stil(c, bg=SARI_VURGU, bold=True)
    ws.merge_cells(start_row=toplam_row, start_column=COL_AD, end_row=toplam_row, end_column=COL_GOREVI)
    c = ws.cell(row=toplam_row, column=COL_AD, value="GENEL TOPLAM")
    stil(c, bg=SARI_VURGU, bold=True, hizala="right")

    genel_toplam = 0.0
    for i in range(n_ay):
        col = COL_AY_BAS + i
        ay_top = sum(k["aylik_fm"][i] for k in tablo)
        c = ws.cell(row=toplam_row, column=col, value=round(ay_top, 2))
        c.number_format = "0.00"
        stil(c, bg=SARI_VURGU, bold=True)
        genel_toplam += ay_top

    c = ws.cell(row=toplam_row, column=COL_TOPLAM, value=round(genel_toplam, 2))
    c.number_format = "0.00"
    stil(c, bg="F8BBD0", bold=True, size=11,
         border=Border(left=med, right=med, top=med, bottom=med))

    # --- Dış çerçeve (tablo etrafı) ---
    from openpyxl.styles import Border as OBorder
    son_veri_satiri = toplam_row
    for row in ws.iter_rows(min_row=4, max_row=son_veri_satiri,
                            min_col=1, max_col=SON_SUTUN):
        for cell in row:
            top_b    = med if cell.row == 4 else cell.border.top
            bottom_b = med if cell.row == son_veri_satiri else cell.border.bottom
            left_b   = med if cell.column == 1 else cell.border.left
            right_b  = med if cell.column == SON_SUTUN else cell.border.right
            cell.border = OBorder(top=top_b, bottom=bottom_b, left=left_b, right=right_b)

    # Yazdırma ayarları
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage  = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    wb.save(path)
